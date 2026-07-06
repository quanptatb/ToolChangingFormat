from datetime import date, datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties

from .common import (
    text_key, parse_date_value, format_date, clean_ingredient_name,
    parse_print_number, is_blank_cell_value, merged_value_lookup,
    effective_cell_value, format_number_clean, concise_customer_label,
    a4_row_empty
)

APPROVAL_OUTPUT_HEADERS = [
    "ID",
    "Trạng thái",
    "Ngày duyệt",
    "Cần duyệt",
    "Mã món",
    "Nhóm món",
    "Tên món ăn",
    "Nhóm NVL",
    "Tên nguyên vật liệu",
    "STT",
    "ĐVT",
    "ĐVT chợ",
    "Định mức theo KH",
    "Version",
]
APPROVAL_LAST_COL = len(APPROVAL_OUTPUT_HEADERS)
APPROVAL_HEADER_ROW = 2
APPROVAL_DATA_START_ROW = 3


def clean_header(value):
    return text_key(value).replace("\n", " ")


def find_approval_source_columns(ws):
    required_keys = {
        "id",
        "status",
        "dish_code",
        "dish_group",
        "dish",
        "ingredient",
        "ingredient_no",
        "unit",
        "purchase_unit",
    }

    for row in range(1, min(ws.max_row, 10) + 1):
        columns = {}
        for col in range(1, ws.max_column + 1):
            header_text = clean_header(ws.cell(row, col).value)
            if not header_text:
                continue
            if header_text == "id":
                columns.setdefault("id", col)
            elif header_text == "trang thai":
                columns.setdefault("status", col)
            elif header_text == "ngay duyet":
                columns.setdefault("approval_date", col)
            elif header_text == "nhom mon an theo nguyen lieu chinh":
                columns.setdefault("main_group", col)
            elif header_text == "nhom mon an theo nvl cap 1":
                columns.setdefault("ingredient_group", col)
            elif header_text == "ma mon an":
                columns.setdefault("dish_code", col)
            elif header_text == "nhom mon":
                columns.setdefault("dish_group", col)
            elif header_text == "ten mon an":
                columns.setdefault("dish", col)
            elif header_text in {"ten nguyen vat lieu", "nguyen vat lieu", "ten nvl"}:
                columns.setdefault("ingredient", col)
            elif header_text in {"stt nvl", "stt nguyen vat lieu"}:
                columns.setdefault("ingredient_no", col)
            elif header_text in {"don vi tinh di cho", "dvt di cho", "don vi tinh mua hang", "dvt mua hang"}:
                columns.setdefault("purchase_unit", col)
            elif header_text in {"don vi tinh", "dvt"}:
                columns.setdefault("unit", col)
            elif header_text == "version":
                columns.setdefault("version", col)
            elif header_text == "can duyet":
                columns.setdefault("needs_review", col)

        if not required_keys.issubset(columns):
            continue

        sentinel_cols = [
            col for col in (columns.get("version"), columns.get("needs_review"))
            if col and col > columns["purchase_unit"]
        ]
        customer_end = min(sentinel_cols) - 1 if sentinel_cols else ws.max_column
        customer_cols = []
        for col in range(columns["purchase_unit"] + 1, customer_end + 1):
            header_value = ws.cell(row, col).value
            header_text = clean_header(header_value)
            if header_text and header_text not in {"version", "can duyet"}:
                customer_cols.append((col, str(header_value).strip()))

        if customer_cols:
            return row, columns, customer_cols

    raise ValueError("Không tìm thấy đủ cột để xuất format Duyệt định mức.")


def approval_norm_summary(ws, lookup, row, customer_cols):
    all_names = [name for _, name in customer_cols]
    grouped = {}
    order = []
    for col, customer_name in customer_cols:
        value = effective_cell_value(ws, lookup, row, col)
        if is_blank_cell_value(value):
            continue
        display_value = format_number_clean(value)
        if display_value not in grouped:
            grouped[display_value] = []
            order.append(display_value)
        grouped[display_value].append(customer_name)

    lines = []
    for display_value in order:
        customer_label = concise_customer_label(grouped[display_value], all_names)
        lines.append(f"{customer_label}: {display_value}")
    return "\n".join(lines)


def collect_approval_records(ws):
    header_row_idx, columns, customer_cols = find_approval_source_columns(ws)
    lookup = merged_value_lookup(ws)
    rows = []

    for source_row in range(header_row_idx + 1, ws.max_row + 1):
        raw = {
            key: effective_cell_value(ws, lookup, source_row, col)
            for key, col in columns.items()
        }
        if all(is_blank_cell_value(value) for value in raw.values()):
            continue

        approval_date = raw.get("approval_date")
        parsed_date = parse_date_value(approval_date)
        ingredient_group = raw.get("ingredient_group") or raw.get("main_group")
        values = [
            raw.get("id"),
            raw.get("status"),
            parsed_date or approval_date,
            raw.get("needs_review"),
            raw.get("dish_code"),
            raw.get("dish_group"),
            raw.get("dish"),
            ingredient_group,
            clean_ingredient_name(raw.get("ingredient")),
            parse_print_number(raw.get("ingredient_no")),
            raw.get("unit"),
            raw.get("purchase_unit"),
            approval_norm_summary(ws, lookup, source_row, customer_cols),
            raw.get("version"),
        ]
        if not a4_row_empty(values):
            rows.append(values)

    return rows


def approval_attention_fill(value, is_status=False):
    key = text_key(value)
    if not key:
        return None
    if "da duyet" in key or "chuan" in key:
        return PatternFill("solid", fgColor="C6EFCE")
    if "chua duyet" in key:
        return PatternFill("solid", fgColor="FFF2CC")
    if "can duyet" in key:
        return PatternFill("solid", fgColor="FCE4D6")
    if not is_status:
        return PatternFill("solid", fgColor="FCE4D6")
    return None


def estimate_wrapped_lines(value, chars_per_line):
    text = "" if value is None else str(value)
    if not text:
        return 1
    return sum(
        max(1, (len(line) + chars_per_line - 1) // chars_per_line)
        for line in text.splitlines()
    )


def apply_approval_base_format(ws, last_row):
    thin_gray = Side(style="thin", color="D9D9D9")
    medium_gray = Side(style="medium", color="6B7280")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_border = Border(left=thin_gray, right=thin_gray, top=medium_gray, bottom=medium_gray)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="DDEBF7")

    ws.title = "Duyệt định mức"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    widths = [8, 11, 10, 12, 10, 11, 24, 13, 25, 5.5, 6, 8, 38, 8]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for index in range(APPROVAL_LAST_COL + 1, 35):
        ws.column_dimensions[get_column_letter(index)].hidden = True

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=APPROVAL_LAST_COL)
    title_cell = ws.cell(1, 1)
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = header_border
    ws.row_dimensions[1].height = 28

    for col in range(1, APPROVAL_LAST_COL + 1):
        cell = ws.cell(APPROVAL_HEADER_ROW, col)
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=8.5, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    ws.row_dimensions[APPROVAL_HEADER_ROW].height = 32

    center_cols = {1, 2, 3, 4, 5, 6, 10, 11, 12, 14}
    for row in range(APPROVAL_DATA_START_ROW, last_row + 1):
        line_count = max(
            estimate_wrapped_lines(ws.cell(row, 7).value, 24),
            estimate_wrapped_lines(ws.cell(row, 9).value, 25),
            estimate_wrapped_lines(ws.cell(row, 13).value, 38),
        )
        ws.row_dimensions[row].height = min(80, max(20, 13 * line_count + 6))

        for col in range(1, APPROVAL_LAST_COL + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=8.5)
            cell.border = border
            horizontal = "center" if col in center_cols else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

        ws.cell(row, 3).number_format = "dd/mm/yyyy"
        ws.cell(row, 10).number_format = "0"
        status_fill = approval_attention_fill(ws.cell(row, 2).value, is_status=True)
        if status_fill:
            ws.cell(row, 2).fill = status_fill
        review_fill = approval_attention_fill(ws.cell(row, 4).value)
        if review_fill:
            ws.cell(row, 4).fill = review_fill


def approval_dish_key(values):
    if not values:
        return None
    code_key = text_key(values[4])
    dish_key = text_key(values[6])
    if not code_key and not dish_key:
        return None
    return code_key, dish_key


def add_approval_page_breaks(ws, rows, max_body_rows=34):
    ws.row_breaks.brk = []
    if not rows:
        return set()

    last_row = len(rows) + APPROVAL_DATA_START_ROW - 1
    boundaries = [
        APPROVAL_DATA_START_ROW + index
        for index in range(1, len(rows))
        if approval_dish_key(rows[index]) != approval_dish_key(rows[index - 1])
    ]
    break_rows = set()
    page_start = APPROVAL_DATA_START_ROW
    while page_start + max_body_rows <= last_row:
        latest_start = page_start + max_body_rows
        earliest_start = page_start + max(14, max_body_rows // 2)
        candidates = [
            row for row in boundaries
            if earliest_start <= row <= latest_start
        ]
        next_page = max(candidates or [latest_start])
        if next_page <= page_start:
            next_page = latest_start
        break_rows.add(next_page)
        ws.row_breaks.append(Break(id=next_page - 1))
        page_start = next_page
    return break_rows


def merge_approval_runs(ws, rows, page_break_rows=None):
    page_break_rows = set(page_break_rows or ())
    if not rows:
        return

    def can_merge_col(start, end, col):
        first = ws.cell(start, col).value
        return all(text_key(ws.cell(row, col).value) == text_key(first) for row in range(start + 1, end + 1))

    def merge_col(start, end, col):
        if end <= start or not can_merge_col(start, end, col):
            return
        ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
        horizontal = "left" if col == 7 else "center"
        ws.cell(start, col).alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

    start = APPROVAL_DATA_START_ROW
    previous_key = approval_dish_key(rows[0])
    last = len(rows) + APPROVAL_DATA_START_ROW - 1
    for row in range(APPROVAL_DATA_START_ROW + 1, last + 2):
        current_key = approval_dish_key(rows[row - APPROVAL_DATA_START_ROW]) if row <= last else None
        should_close = row in page_break_rows or current_key != previous_key or previous_key is None
        if should_close:
            end = row - 1
            for col in (1, 2, 3, 4, 5, 6, 7, 14):
                merge_col(start, end, col)
            start = row
            previous_key = current_key


def process_sheet_approval(ws_source):
    rows = collect_approval_records(ws_source)
    if not rows:
        raise ValueError("File nguồn không có dữ liệu để xuất format Duyệt định mức.")

    workbook = Workbook()
    ws = workbook.active
    ws.append(["DUYỆT ĐỊNH MỨC"])
    ws.append(APPROVAL_OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)

    last_row = len(rows) + APPROVAL_DATA_START_ROW - 1
    apply_approval_base_format(ws, last_row)
    page_break_rows = add_approval_page_breaks(ws, rows)
    merge_approval_runs(ws, rows, page_break_rows)
    return workbook
