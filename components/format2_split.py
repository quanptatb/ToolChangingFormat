import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .common import (
    clean_ingredient_name,
    parse_date_value,
    format_date,
    to_numeric,
    format_number_clean,
    strip_accents,
    SHORT_NAMES,
    output_dir,
    input_dir,
    save_workbook
)

def get_shopping_short_name(nvl_name):
    n = str(nvl_name).lower().strip()
    n = clean_ingredient_name(n)
    
    # Custom mapping for Format 2
    if any(x in n for x in ["nạc vai", "heo vai", "vai tươi", "vai heo"]):
        return "vai"
    if "trứng" in n:
        return "trứng gà"
    if "củ cải" in n:
        return "củ cải"
    if "má gà" in n or "thịt gà" in n:
        return "gà"
    if "tôm" in n:
        return "tôm"
    if "cá sống" in n:
        return "cá sống"
    if "mọc" in n:
        return "mọc"
    if "xay" in n or "bằm" in n:
        return "thịt xay"
    if "đậu hũ chiên" in n:
        return "đậu hũ chiên"
    if "đậu hũ trắng" in n:
        return "đậu hũ trắng"
    if "nấm mèo" in n:
        return "nấm mèo"
    if "miến" in n:
        return "miến"
    if "đậu xanh" in n:
        return "đậu xanh"
    if "bào ngư" in n:
        return "nấm bào ngư"
    if "bông cải" in n:
        return "bông cải"
    if "vịt" in n:
        return "vịt"
    if "khoai môn" in n:
        return "khoai môn"
    if "chả cá" in n:
        return "chả cá"
    if "xương heo" in n or "xương hầm" in n:
        return "xương hầm"
    if "kim chi" in n:
        return "kim chi"
    if "hoành thánh" in n:
        return "hoành thánh"
    if "bò gàu" in n:
        return "bò gàu"
    if "chả lụa" in n:
        return "chả lụa"
    if "chả bò" in n:
        return "chả bò"
    if "cà chua" in n:
        return "cà chua"
    if "mực nang" in n or "mực" in n:
        return "mực"
    if "bò tái" in n:
        return "bò tái"
    if "bò nạm" in n:
        return "bò nạm"
    if "bò viên" in n:
        return "bò viên"
        
    return SHORT_NAMES.get(n, n)

def format_single_ingredient_quota(amount, unit, short_name):
    amt_str = format_number_clean(amount)
    u = str(unit).strip().lower() if unit else ""
    if not u:
        return f"{amt_str} {short_name}"
    
    if u in ["gr", "g", "kg", "ml", "l", "c"]:
        return f"{amt_str}{u} {short_name}"
    else:
        return f"{amt_str} {u} {short_name}"

def format_single_ingredient_total(value, unit, short_name):
    val_str = format_number_clean(value)
    u = str(unit).strip().lower() if unit else ""
    if not u:
        return f"{val_str} {short_name}"
        
    if u in ["quả", "qua", "cái", "cai", "con", "miếng", "mieng", "viên", "vien", "cây", "cay", "khay"]:
        # Count units: omit unit to match user's requirement (e.g. 340 trứng gà)
        return f"{val_str} {short_name}"
    elif u in ["gr", "g", "kg", "ml", "l", "c"]:
        return f"{val_str}{u} {short_name}"
    else:
        return f"{val_str} {u} {short_name}"

def clean_site_name(site_text):
    s = str(site_text).strip().lower()
    s_no_space = s.replace(" ", "")
    if "fab1" in s_no_space:
        return "fab1"
    if "fab2" in s_no_space:
        return "fab2"
    if "vpc" in s_no_space:
        return "vpc"
    if "til" in s_no_space:
        return "til"
    
    # Remove common prefix like scavi or sofa
    for prefix in ["scavi - ", "scavi-", "sofa - ", "sofa-"]:
        if s.startswith(prefix):
            s = s[len(prefix):]
    return s

def get_site_sort_key(site_name):
    s = str(site_name).lower()
    if "vpc" in s:
        return (0, s)
    if "fab1" in s:
        return (1, s)
    if "fab2" in s:
        return (2, s)
    if "til" in s:
        return (3, s)
    return (4, s)

def map_columns(headers):
    col_map = {}
    for idx, col in enumerate(headers, start=1):
        if col is None:
            continue
        c_clean = str(col).lower().strip()
        
        if c_clean in ["ngày", "ngay", "date"]:
            col_map["ngay"] = idx
        elif c_clean in ["ca", "shift"]:
            col_map["ca"] = idx
        elif c_clean in ["khách hàng", "khach hang", "mã kh", "ma kh", "khach_hang", "mã khách hàng"] and "site" not in c_clean:
            col_map["ma_kh"] = idx
        elif c_clean in ["bếp", "bep", "kitchen"]:
            col_map["bep"] = idx
        elif c_clean in ["site ăn", "site an", "site"]:
            col_map["site_an"] = idx
        elif c_clean in ["cơ cấu suất ăn", "co cau suat an", "suất ăn", "suat an", "cơ cấu menu", "co cau menu"]:
            col_map["co_cau"] = idx
        elif c_clean in ["loại món ăn", "loai mon an", "nhóm món", "nhom mon", "cơ cấu món ăn", "co cau mon an"]:
            col_map["loai_mon"] = idx
        elif c_clean in ["nguyên vật liệu", "nguyen vat lieu", "nvl", "tên nguyên vật liệu", "ten nguyen vat lieu"] and "đầy đủ" not in c_clean and "day du" not in c_clean:
            col_map["nvl"] = idx
        elif c_clean in ["món ăn", "mon an", "tên món", "ten mon"] and "cơ cấu" not in c_clean and "loại" not in c_clean and "loai" not in c_clean:
            col_map["mon"] = idx
        elif c_clean in ["số lượng", "so luong", "s.lượng", "qty", "quantity"] and "dự báo" not in c_clean and "du bao" not in c_clean:
            col_map["qty"] = idx
        elif c_clean in ["định mức", "dinh muc", "định lượng", "dinh luong"] and "cam kết" not in c_clean and "cam ket" not in c_clean:
            col_map["dinh_muc"] = idx
        elif c_clean in ["đvt", "đơn vị tính", "don vi tinh", "đơn vị", "don vi", "dvt"] and "mua hàng" not in c_clean and "mua hang" not in c_clean:
            col_map["dvt"] = idx
        elif any(x in c_clean for x in ["kl yêu cầu", "khoi luong yeu cau", "kl yeu cau", "quy đổi", "quy doi", "kl thực tế", "kl thuc te"]):
            col_map["kl_mua"] = idx
        elif any(x in c_clean for x in ["đơn vị tính mua hàng", "don vi tinh mua hang", "dvt mua", "dvt_mua"]):
            col_map["dvt_mua"] = idx
            
    return col_map

def process_all_files_to_split():
    # Read files in Excel
    excel_files = sorted(input_dir.glob("*.xlsx"))
    excel_files += sorted(input_dir.glob("*.xlsm"))
    excel_files += sorted(input_dir.glob("*.csv"))
    excel_files = [f for f in excel_files if not f.name.startswith("~$")]
    
    if not excel_files:
        print("Không tìm thấy file nào trong thư mục Excel.")
        return
        
    all_parsed_rows = []
    
    for file_path in excel_files:
        print(f"Đang đọc file: {file_path.name}")
        if file_path.suffix.lower() == ".csv":
            import csv
            from io import StringIO
            wb = Workbook()
            ws = wb.active
            text = file_path.read_bytes().decode("utf-8-sig", errors="ignore")
            reader = csv.reader(StringIO(text))
            for r in reader:
                ws.append(r)
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = map_columns(headers)
        
        # Read rows
        raw_rows = []
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if any(x is not None for x in row_vals):
                raw_rows.append(row_vals)
                
        # Forward fill
        prev_vals = {}
        fill_keys = ["ngay", "ma_kh", "bep", "ca", "co_cau", "loai_mon", "mon", "qty", "site_an"]
        for row in raw_rows:
            row_vals = list(row)
            for key in fill_keys:
                idx = col_map.get(key)
                if idx is not None and (idx - 1) < len(row_vals):
                    val = row_vals[idx - 1]
                    if val is None or str(val).strip() == "":
                        val = prev_vals.get(key, "")
                    else:
                        val = str(val).strip()
                    row_vals[idx - 1] = val
                    prev_vals[key] = val
            
            # Pack as dictionary for easy access
            row_dict = {}
            for key, col_idx in col_map.items():
                if col_idx is not None and (col_idx - 1) < len(row_vals):
                    row_dict[key] = row_vals[col_idx - 1]
                else:
                    row_dict[key] = ""
            all_parsed_rows.append(row_dict)
            
    # Now group by Customer (Khách hàng)
    cust_groups = defaultdict(list)
    for r in all_parsed_rows:
        cust = r.get("ma_kh", "").strip()
        if cust:
            cust_groups[cust].append(r)
            
    # For each customer, generate a sheet
    output_dir.mkdir(parents=True, exist_ok=True)
    
    for cust, rows in cust_groups.items():
        # Get dates and kitchens
        dates = sorted(list(set(str(r.get("ngay", "")).strip() for r in rows if r.get("ngay"))))
        kitchens = sorted(list(set(str(r.get("bep", "")).strip() for r in rows if r.get("bep"))))
        
        date_str = ", ".join(dates) if dates else "Chưa xác định"
        kitchen_str = ", ".join(kitchens) if kitchens else "Chưa xác định"
        
        # Identify if this customer has multiple sites
        sites = set(clean_site_name(r.get("site_an", "")) for r in rows if r.get("site_an"))
        sites.discard("")
        is_multi_site = len(sites) > 1
        
        # Group rows by (ca, co_cau, loai_mon, mon)
        dish_groups = defaultdict(list)
        for r in rows:
            ca = str(r.get("ca") or "").strip()
            co_cau = str(r.get("co_cau") or "").strip()
            loai_mon = str(r.get("loai_mon") or "").strip()
            mon = str(r.get("mon") or "").strip()
            group_key = (ca, co_cau, loai_mon, mon)
            dish_groups[group_key].append(r)
            
        # Construct processed rows
        processed_data = []
        for (ca, co_cau, loai_mon, mon), r_list in dish_groups.items():
            # Gather unique sites in this group
            site_qty_map = defaultdict(float)
            site_nvl_totals = defaultdict(lambda: defaultdict(float)) # site -> nvl -> total_kl
            nvl_dinh_muc = {} # nvl -> (dm_val, dm_unit, dvt_mua)
            
            for r in r_list:
                site = clean_site_name(r.get("site_an", ""))
                qty = to_numeric(r.get("qty")) or 0.0
                site_qty_map[site] += qty
                
                nvl = str(r.get("nvl") or "").strip()
                if nvl:
                    dm_val = to_numeric(r.get("dinh_muc")) or 0.0
                    dm_unit = str(r.get("dvt") or "").strip()
                    dvt_mua = str(r.get("dvt_mua") or "").strip()
                    kl_mua = to_numeric(r.get("kl_mua")) or 0.0
                    
                    nvl_dinh_muc[nvl] = (dm_val, dm_unit, dvt_mua)
                    site_nvl_totals[site][nvl] += kl_mua
            
            # Sort sites
            sorted_sites = sorted(list(site_qty_map.keys()), key=get_site_sort_key)
            
            # 1. Format ODSV "loại món ăn" column logic:
            # "công ty ODSV thì cột hiển thị dựa vào dữ liệu suất ăn, các công ty khác hiển thị theo món ăn"
            if cust == "ODSV":
                loai_mon_val = co_cau
            else:
                loai_mon_val = loai_mon
                
            # 2. Format Quantity (số lượng)
            if is_multi_site:
                qty_val = "\n".join(f"{s} - {format_number_clean(site_qty_map[s])}" for s in sorted_sites)
            else:
                qty_val = format_number_clean(sum(site_qty_map.values()))
                
            # 3. Format Quota (định mức của món ăn đó)
            quota_parts = []
            for nvl, (dm_val, dm_unit, _) in nvl_dinh_muc.items():
                short_nvl = get_shopping_short_name(nvl)
                quota_parts.append(format_single_ingredient_quota(dm_val, dm_unit, short_nvl))
            quota_str = f"{mon.lower()} " + " + ".join(quota_parts)
            
            # 4. Format Total Shopping Weight (khối lượng đi chợ tổng)
            if is_multi_site:
                site_shopping_lines = []
                for s in sorted_sites:
                    # Construct shopping total for this site
                    weight_parts = []
                    for nvl in nvl_dinh_muc:
                        kl = site_nvl_totals[s].get(nvl, 0.0)
                        _, _, dvt_mua = nvl_dinh_muc[nvl]
                        short_nvl = get_shopping_short_name(nvl)
                        weight_parts.append(format_single_ingredient_total(kl, dvt_mua, short_nvl))
                    weight_str = " + ".join(weight_parts)
                    site_shopping_lines.append(f"{s} - {quota_str} | {weight_str}")
                shopping_val = "\n".join(site_shopping_lines)
            else:
                weight_parts = []
                for nvl in nvl_dinh_muc:
                    # Sum weight across all rows in this group
                    total_kl = sum(site_nvl_totals[s].get(nvl, 0.0) for s in site_qty_map)
                    _, _, dvt_mua = nvl_dinh_muc[nvl]
                    short_nvl = get_shopping_short_name(nvl)
                    weight_parts.append(format_single_ingredient_total(total_kl, dvt_mua, short_nvl))
                weight_str = " + ".join(weight_parts)
                shopping_val = f"{quota_str} | {weight_str}"
                
            processed_data.append({
                "cust": cust,
                "ca": ca,
                "co_cau": co_cau,
                "loai_mon": loai_mon_val,
                "qty": qty_val,
                "mon": mon,
                "quota": quota_str,
                "shopping": shopping_val
            })
            
        # Write Excel File
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = cust
        
        # Enable grid lines explicitly
        ws_out.sheet_view.showGridLines = True
        
        # Styles
        font_title = Font(name="Calibri", size=16, bold=True)
        font_meta = Font(name="Calibri", size=11, bold=True)
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        font_qty = Font(name="Calibri", size=11, bold=True)
        
        fill_header = PatternFill(start_color="145C52", end_color="145C52", fill_type="solid") # Dark Teal
        
        thin_border = Border(
            left=Side(style='thin', color='D8DEE6'),
            right=Side(style='thin', color='D8DEE6'),
            top=Side(style='thin', color='D8DEE6'),
            bottom=Side(style='thin', color='D8DEE6')
        )
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        
        # Write Title and Metadata
        ws_out.merge_cells("A2:H2")
        ws_out["A2"] = f"BẢNG ĐỊNH MỨC VÀ ĐI CHỢ - {cust.upper()}"
        ws_out["A2"].font = font_title
        ws_out["A2"].alignment = align_center
        
        ws_out["A3"] = f"Bếp: {kitchen_str}"
        ws_out["A3"].font = font_meta
        
        ws_out["A4"] = f"Ngày: {date_str}"
        ws_out["A4"].font = font_meta
        
        ws_out.row_dimensions[2].height = 25
        ws_out.row_dimensions[3].height = 18
        ws_out.row_dimensions[4].height = 18
        
        # Table Headers
        headers_list = [
            "Khách hàng",
            "Ca",
            "Cơ cấu suất ăn",
            "Loại món ăn",
            "Số lượng",
            "Tên món ăn",
            "Định mức món ăn",
            "Khối lượng đi chợ tổng"
        ]
        
        ws_out.row_dimensions[6].height = 28
        for col_idx, h in enumerate(headers_list, start=1):
            cell = ws_out.cell(6, col_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
        # Write Data
        row_idx = 7
        for item in processed_data:
            # Set values
            ws_out.cell(row_idx, 1, item["cust"])
            ws_out.cell(row_idx, 2, item["ca"])
            ws_out.cell(row_idx, 3, item["co_cau"])
            ws_out.cell(row_idx, 4, item["loai_mon"])
            ws_out.cell(row_idx, 5, item["qty"])
            ws_out.cell(row_idx, 6, item["mon"])
            ws_out.cell(row_idx, 7, item["quota"])
            ws_out.cell(row_idx, 8, item["shopping"])
            
            # Dynamic Row Height based on number of newlines (e.g. for multi-site)
            lines_qty = str(item["qty"]).count("\n") + 1
            lines_shop = str(item["shopping"]).count("\n") + 1
            max_lines = max(lines_qty, lines_shop)
            ws_out.row_dimensions[row_idx].height = max_lines * 18 + 5
            
            # Apply styles
            for col_idx in range(1, 9):
                cell = ws_out.cell(row_idx, col_idx)
                cell.font = font_qty if col_idx == 5 else font_data
                cell.border = thin_border
                
                # Alignments
                if col_idx in [1, 2, 3, 4, 5]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
            row_idx += 1
            
        # Auto-fit Column Widths
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for r in range(6, row_idx):
                val = ws_out.cell(r, col_idx).value
                if val:
                    # Find maximum line length for multi-line cells
                    lines = str(val).split("\n")
                    max_len = max(max_len, max(len(line) for line in lines))
            
            ws_out.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Save File
        out_name = f"{cust} - format 2.xlsx"
        out_path = output_dir / out_name
        save_workbook(wb_out, out_path)
        print(f"Đã lưu file khách hàng: {out_path.name}")
        
    print("Hoàn tất xử lý định dạng và tách file.")
