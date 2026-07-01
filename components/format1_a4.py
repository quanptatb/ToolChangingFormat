from copy import copy
from datetime import date, datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.properties import PageSetupProperties

from .common import (
    source_start_row, title_row, header_row, body_start_row,
    title_font_size, header_font_size, body_font_size,
    col_ngay, col_ma_kh, col_ca, col_so_luong, col_ten_mon, col_khoi_luong, col_khoi_luong_di_cho,
    parse_date_value, format_date, collect_date_text, kitchen_name_from_file,
    unmerge_group_columns, get_lixil_site_order, text_key, get_co_cau_order,
    meal_type_label, is_generic_meal_label, normalized_value,
    merged_value_lookup, effective_cell_value, to_numeric, clean_ingredient_name,
    normalize_shift, get_shift_order, save_workbook, parse_print_number, a4_row_empty
)

A4_OUTPUT_HEADERS = [
    "Khách hàng",
    "Ca",
    "Site",
    "Loại món ăn",
    "Món ăn",
    "Số lượng",
    "Định mức\ncam kết",
    "Nguyên liệu",
    "Định\nmức",
    "Đơn vị\ntính",
    "KL yêu cầu\nsản xuất",
    "Đơn vị tính\nmua hàng",
    "KL duyệt đi chợ",
]

A4_LAST_COL = len(A4_OUTPUT_HEADERS)
A4_HEADER_ROW = 2
A4_DATA_START_ROW = 3


def sort_rows_by_groups(ws, max_row, max_col, col_ngay_idx, col_ca_idx, col_ma_kh_idx, col_site_an_idx, col_co_cau_idx, col_ten_mon_idx, col_so_luong_idx):
    prev_ngay = None
    prev_ca = None
    prev_company = None
    prev_site = None
    prev_co_cau = None
    prev_dish = None
    prev_quantity = None
    
    rows = []
    for row_idx in range(source_start_row, max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
        
        if col_ngay_idx is not None and col_ngay_idx <= len(values):
            prev_ngay = normalized_value(values[col_ngay_idx - 1], prev_ngay)
            values[col_ngay_idx - 1] = prev_ngay
            
        if col_ca_idx is not None and col_ca_idx <= len(values):
            prev_ca = normalized_value(values[col_ca_idx - 1], prev_ca)
            values[col_ca_idx - 1] = prev_ca
            
        if col_ma_kh_idx is not None and col_ma_kh_idx <= len(values):
            prev_company = normalized_value(values[col_ma_kh_idx - 1], prev_company)
            values[col_ma_kh_idx - 1] = prev_company
            
        if col_site_an_idx is not None and col_site_an_idx <= len(values):
            prev_site = normalized_value(values[col_site_an_idx - 1], prev_site)
            values[col_site_an_idx - 1] = prev_site
            
        if col_co_cau_idx is not None and col_co_cau_idx <= len(values):
            prev_co_cau = normalized_value(values[col_co_cau_idx - 1], prev_co_cau)
            values[col_co_cau_idx - 1] = prev_co_cau
            
        if col_ten_mon_idx is not None and col_ten_mon_idx <= len(values):
            prev_dish = normalized_value(values[col_ten_mon_idx - 1], prev_dish)
            values[col_ten_mon_idx - 1] = prev_dish
            
        if col_so_luong_idx is not None and col_so_luong_idx <= len(values):
            prev_quantity = normalized_value(values[col_so_luong_idx - 1], prev_quantity)
            values[col_so_luong_idx - 1] = prev_quantity

        rows.append((row_idx, values))

    def get_row_sort_key(item):
        row_idx, values = item
        
        ngay_val = values[col_ngay_idx - 1] if col_ngay_idx else ""
        dt_val = parse_date_value(ngay_val)
        if dt_val:
            date_key = (0, dt_val)
        else:
            date_key = (1, str(ngay_val))
        
        ca_val = values[col_ca_idx - 1] if col_ca_idx else ""
        ca_key = get_shift_order(normalize_shift(ca_val))
        
        kh_val = str(values[col_ma_kh_idx - 1] or "").strip().upper() if col_ma_kh_idx else ""
        
        site_val = str(values[col_site_an_idx - 1] or "").strip() if col_site_an_idx else ""
        if kh_val == "LIXIL":
            site_key = (get_lixil_site_order(site_val), site_val.casefold())
        else:
            site_key = (0, site_val.casefold())
            
        if col_co_cau_idx:
            co_cau_raw = str(values[col_co_cau_idx - 1] or "").strip()
            co_cau_key = (get_co_cau_order(co_cau_raw, kh_val), co_cau_raw.lower())
        else:
            co_cau_key = (99, "")
            
        dish_val = str(values[col_ten_mon_idx - 1] or "").strip().lower() if col_ten_mon_idx else ""
        
        return (date_key, ca_key, kh_val, site_key, co_cau_key, dish_val, row_idx)

    rows.sort(key=get_row_sort_key)

    for new_row, (orig_row_idx, values) in enumerate(rows, start=source_start_row):
        for col, value in enumerate(values, start=1):
            ws.cell(new_row, col).value = value
        ws.cell(new_row, 1).value = new_row - source_start_row + 1


def set_center(ws, row, col):
    ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")


def merge_group_columns(ws, max_row, column_indices, col_ngay_idx, col_ca_idx, col_ma_kh_idx, col_site_an_idx, col_ten_mon_idx, col_so_luong_idx, page_break_rows=None):
    valid_cols = [c for c in column_indices if c is not None]
    if not valid_cols:
        return
    page_break_rows = set(page_break_rows or ())

    row_keys = {}
    for row in range(body_start_row, max_row + 1):
        key = []
        for c in valid_cols:
            val = ws.cell(row, c).value
            key.append(val)
        row_keys[row] = key

    for col_idx_in_hierarchy, col in enumerate(valid_cols):
        start_row = body_start_row
        current_prefix = [row_keys[body_start_row][i] for i in range(col_idx_in_hierarchy + 1)]
        
        for row in range(body_start_row + 1, max_row + 2):
            prefix = [row_keys[row][i] for i in range(col_idx_in_hierarchy + 1)] if row <= max_row else None
            
            if row in page_break_rows or prefix != current_prefix:
                end_row = row - 1
                if start_row < end_row:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col,
                        end_row=end_row,
                        end_column=col,
                    )
                # Apply alignment to merged range start cell
                if col in (col_ngay_idx, col_ca_idx, col_ma_kh_idx, col_site_an_idx, col_so_luong_idx):
                    ws.cell(start_row, col).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                else:
                    ws.cell(start_row, col).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                current_prefix = prefix
                start_row = row


def add_group_page_breaks(ws, max_row, column_indices, max_body_rows=22):
    """Keep printed groups readable and prevent merged cells crossing pages."""
    valid_cols = [c for c in column_indices if c is not None]
    ws.row_breaks.brk = []
    if not valid_cols or max_row < body_start_row + max_body_rows:
        return set()

    def key_at(row, depth):
        return tuple(ws.cell(row, col).value for col in valid_cols[:depth])

    # Prefer a new page at a Site boundary; fall back to a dish boundary when
    # one site contains more rows than a page can comfortably show.
    major_depth = min(4, len(valid_cols))
    major_boundaries = {
        row for row in range(body_start_row + 1, max_row + 1)
        if key_at(row, major_depth) != key_at(row - 1, major_depth)
    }
    detail_boundaries = {
        row for row in range(body_start_row + 1, max_row + 1)
        if key_at(row, len(valid_cols)) != key_at(row - 1, len(valid_cols))
    }

    break_rows = set()
    page_start = body_start_row
    while page_start + max_body_rows <= max_row:
        latest_start = page_start + max_body_rows
        earliest_start = page_start + max(10, max_body_rows // 2)
        preferred = [
            row for row in major_boundaries
            if earliest_start <= row <= latest_start
        ]
        fallback = [
            row for row in detail_boundaries
            if earliest_start <= row <= latest_start
        ]
        next_page = max(preferred or fallback or [latest_start])
        if next_page <= page_start:
            next_page = latest_start
        break_rows.add(next_page)
        ws.row_breaks.append(Break(id=next_page - 1))
        page_start = next_page

    return break_rows


def visible_last_column(ws, max_col):
    for col in range(max_col, 0, -1):
        value = ws.cell(header_row, col).value
        if isinstance(value, str) and value.strip().casefold() == "sourceid":
            continue
        return col
    return max_col


def apply_print_layout(ws, max_row, max_col, title_text, col_ma_kh_idx=None, col_ca_idx=None, col_so_luong_idx=None, col_ten_mon_idx=None, col_khoi_luong_idx=None, col_khoi_luong_di_cho_idx=None, col_ngay_idx=None, col_site_an_idx=None, col_co_cau_idx=None):
    if col_ma_kh_idx is None: col_ma_kh_idx = col_ma_kh
    if col_ca_idx is None: col_ca_idx = col_ca
    if col_so_luong_idx is None: col_so_luong_idx = col_so_luong
    if col_ten_mon_idx is None: col_ten_mon_idx = col_ten_mon
    if col_khoi_luong_idx is None: col_khoi_luong_idx = col_khoi_luong
    if col_khoi_luong_di_cho_idx is None: col_khoi_luong_di_cho_idx = col_khoi_luong_di_cho

    last_visible_col = visible_last_column(ws, max_col)

    # Keep the requested grouping columns visible. Only the technical ID column
    # and an optional SourceId column are excluded from the printed report.
    hidden_cols_indices = [1]
    col_12_val = ws.cell(header_row, 12).value
    if col_12_val and str(col_12_val).strip().lower() == "sourceid":
        hidden_cols_indices.append(12)

    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].hidden = c in hidden_cols_indices

    first_visible_col = min(c for c in range(1, max_col + 1) if c not in hidden_cols_indices)
    first_visible_letter = get_column_letter(first_visible_col)
    last_visible_letter = get_column_letter(last_visible_col)

    ws.merge_cells(
        start_row=title_row,
        start_column=first_visible_col,
        end_row=title_row,
        end_column=last_visible_col,
    )
    title_cell = ws.cell(title_row, first_visible_col)
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", bold=True, size=title_font_size, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="404040")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 28

    # Set column widths dynamically based on role mapping (scaled for size 13 font)
    role_widths = {
        col_ngay_idx: 14,
        col_ma_kh_idx: 18,
        col_ca_idx: 8,
        col_so_luong_idx: 10,
        col_ten_mon_idx: 24,
        col_khoi_luong_idx: 10,
        col_khoi_luong_di_cho_idx: 14,
    }
    if col_site_an_idx:
        role_widths[col_site_an_idx] = 14
    if col_co_cau_idx:
        role_widths[col_co_cau_idx] = 14
    
    col_nvl_idx = None
    col_dvt_idx = None
    col_dvt_dc_idx = None
    
    for c in range(1, max_col + 1):
        val = ws.cell(header_row, c).value
        if not val:
            continue
        v_str = str(val).lower()
        if "nguyên vật liệu" in v_str or "nguyen vat lieu" in v_str or "nvl" in v_str:
            col_nvl_idx = c
        elif "đơn vị tính" in v_str or "don vi tinh" in v_str or "dvt" in v_str:
            if "đi chợ" in v_str or "di cho" in v_str or "mua hàng" in v_str or "mua hang" in v_str:
                col_dvt_dc_idx = c
            else:
                col_dvt_idx = c
                
    if col_nvl_idx:
        role_widths[col_nvl_idx] = 30
    if col_dvt_idx:
        role_widths[col_dvt_idx] = 12
    if col_dvt_dc_idx:
        role_widths[col_dvt_dc_idx] = 12

    for c, width in role_widths.items():
        if c is not None and c <= max_col:
            ws.column_dimensions[get_column_letter(c)].width = width

    thin = Side(style="thin", color="D9D9D9")
    medium = Side(style="medium", color="B7B7B7")
    header_fill = PatternFill("solid", fgColor="C0C0C0")
    group_fill = PatternFill("solid", fgColor="F2F2F2")

    for col in range(1, last_visible_col + 1):
        cell = ws.cell(header_row, col)
        cell.font = Font(name="Calibri", bold=True, size=header_font_size, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    ws.row_dimensions[header_row].height = 64

    previous_group_key = None
    for row in range(body_start_row, max_row + 1):
        group_key = (
            ws.cell(row, col_ngay_idx).value if col_ngay_idx else None,
            ws.cell(row, col_ca_idx).value if col_ca_idx else None,
            ws.cell(row, col_ma_kh_idx).value,
            ws.cell(row, col_site_an_idx).value if col_site_an_idx else None,
            ws.cell(row, col_co_cau_idx).value if col_co_cau_idx else None,
            ws.cell(row, col_ten_mon_idx).value,
        )
        is_group_start = group_key != previous_group_key
        previous_group_key = group_key

        for col in range(1, last_visible_col + 1):
            cell = ws.cell(row, col)
            updated_font = copy(cell.font)
            updated_font.name = "Calibri"
            updated_font.size = body_font_size
            cell.font = updated_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=thin,
                right=thin,
                top=medium if is_group_start else thin,
                bottom=thin,
            )
            if is_group_start and col in (
                col_ngay_idx,
                col_ma_kh_idx,
                col_ca_idx,
                col_site_an_idx,
                col_co_cau_idx,
                col_ten_mon_idx,
            ):
                cell.fill = group_fill

        ws.row_dimensions[row].height = 24

    for row in range(body_start_row, max_row + 1):
        for col in (col_ngay_idx, col_ma_kh_idx, col_ca_idx, col_site_an_idx, col_co_cau_idx):
            if col is not None and col <= max_col:
                ws.cell(row, col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
        ws.cell(row, col_ten_mon_idx).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

    for row in range(body_start_row, max_row + 1):
        ws.cell(row, col_so_luong_idx).number_format = "#,##0.###"
        if col_khoi_luong_idx is not None and col_khoi_luong_idx <= max_col:
            ws.cell(row, col_khoi_luong_idx).number_format = "#,##0.###"
        if col_khoi_luong_di_cho_idx is not None and col_khoi_luong_di_cho_idx <= max_col:
            ws.cell(row, col_khoi_luong_di_cho_idx).number_format = "#,##0.###"
        ws.cell(row, col_so_luong_idx).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        for col in (col_khoi_luong_idx, col_khoi_luong_di_cho_idx):
            if col is not None and col <= max_col:
                ws.cell(row, col).alignment = Alignment(
                    horizontal="right", vertical="top", wrap_text=True
                )

    ws.freeze_panes = f"{get_column_letter(first_visible_col)}3"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"{first_visible_letter}{header_row}:{last_visible_letter}{max_row}"
    ws.print_area = f"{first_visible_letter}{title_row}:{last_visible_letter}{max_row}"
    ws.print_title_rows = "1:2"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.blackAndWhite = True
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.pageOrder = "downThenOver"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.oddHeader.center.text = ""
    ws.oddFooter.center.text = ""
    ws.sheet_view.zoomScale = 85


def process_sheet(ws, file_path, date_mode="auto", selected_kitchen=None):
    max_row = ws.max_row
    if max_row < source_start_row:
        return
    max_col = ws.max_column

    # Build col_map dynamically from row 1 (header row)
    col_map = {
        "ngay": col_ngay,
        "ma_kh": col_ma_kh,
        "ca": col_ca,
        "so_luong": col_so_luong,
        "ten_mon": col_ten_mon,
        "khoi_luong": col_khoi_luong,
        "khoi_luong_di_cho": col_khoi_luong_di_cho,
        "site_an": None,
        "co_cau": None
    }
    
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    has_any_header = any(h is not None for h in header)
    if has_any_header:
        for idx, col in enumerate(header, start=1):
            if col is None:
                continue
            c_lower = str(col).lower().strip()
            if any(x in c_lower for x in ["ngày", "ngay", "date"]):
                col_map["ngay"] = idx
            elif any(x in c_lower for x in ["khách hàng", "khach hang", "mã kh", "ma kh", "khach_hang", "mã khách hàng"]) and "site" not in c_lower:
                col_map["ma_kh"] = idx
            elif any(x in c_lower for x in ["ca", "shift"]):
                col_map["ca"] = idx
            elif any(x in c_lower for x in ["site ăn", "site an", "site"]):
                col_map["site_an"] = idx
            elif any(x in c_lower for x in ["cơ cấu món ăn", "co cau mon an", "cơ cấu", "co cau", "nhóm món", "nhom mon"]):
                col_map["co_cau"] = idx
            elif any(x in c_lower for x in ["món ăn", "mon an", "tên món", "ten mon"]) and "cơ cấu" not in c_lower and "co cau" not in c_lower:
                col_map["ten_mon"] = idx
            elif any(x in c_lower for x in ["số lượng", "so luong", "s.lượng", "qty", "quantity"]):
                col_map["so_luong"] = idx
            elif any(x in c_lower for x in ["định mức", "dinh muc", "khối lượng", "định lượng"]) and "đi chợ" not in c_lower and "di cho" not in c_lower and "yêu cầu" not in c_lower and "yeu cau" not in c_lower and "mua hàng" not in c_lower and "mua hang" not in c_lower:
                col_map["khoi_luong"] = idx
            elif any(x in c_lower for x in ["khối lượng đi chợ", "đi chợ", "di cho", "định lương đi chợ", "yêu cầu sản xuất", "yeu cau san xuat", "kl yêu cầu", "kl yeu cau", "mua hàng", "mua hang"]) and "đơn vị tính" not in c_lower and "don vi tinh" not in c_lower and "dvt" not in c_lower and "đvt" not in c_lower:
                col_map["khoi_luong_di_cho"] = idx

    col_ngay_idx = col_map["ngay"]
    col_ma_kh_idx = col_map["ma_kh"]
    col_ca_idx = col_map["ca"]
    col_so_luong_idx = col_map["so_luong"]
    col_ten_mon_idx = col_map["ten_mon"]
    col_khoi_luong_idx = col_map["khoi_luong"]
    col_khoi_luong_di_cho_idx = col_map["khoi_luong_di_cho"]
    col_site_an_idx = col_map["site_an"]
    col_co_cau_idx = col_map["co_cau"]

    # CSV imports arrive as text. Store dates and quantities as real Excel
    # values so printouts are clean and users can continue sorting/calculating.
    for row in range(source_start_row, max_row + 1):
        if col_ngay_idx:
            parsed_date = parse_date_value(ws.cell(row, col_ngay_idx).value)
            if parsed_date:
                ws.cell(row, col_ngay_idx).value = parsed_date
                ws.cell(row, col_ngay_idx).number_format = "dd/mm/yyyy"
        for numeric_col in (
            col_so_luong_idx,
            col_khoi_luong_idx,
            col_khoi_luong_di_cho_idx,
        ):
            if numeric_col:
                ws.cell(row, numeric_col).value = to_numeric(
                    ws.cell(row, numeric_col).value
                )

    if date_mode == "none":
        date_text = ""
    else:
        date_text = collect_date_text(ws, source_start_row, max_row, col_ngay_idx)
        
    if selected_kitchen:
        kitchen_name = selected_kitchen
    else:
        kitchen_name = kitchen_name_from_file(file_path)
        
    title_text = f"Bếp: {kitchen_name}"
    if date_text:
        title_text = f"{title_text} | Ngày: {date_text}"

    unmerge_group_columns(ws, source_start_row)
    sort_rows_by_groups(
        ws, max_row, max_col, 
        col_ngay_idx, col_ca_idx, col_ma_kh_idx, col_site_an_idx, col_co_cau_idx, col_ten_mon_idx, col_so_luong_idx
    )

    ws.insert_rows(title_row)
    max_row = ws.max_row

    column_indices = [col_ngay_idx, col_ca_idx, col_ma_kh_idx, col_site_an_idx, col_co_cau_idx, col_ten_mon_idx]
    apply_print_layout(
        ws, max_row, max_col, title_text, 
        col_ma_kh_idx, col_ca_idx, col_so_luong_idx, col_ten_mon_idx, col_khoi_luong_idx, col_khoi_luong_di_cho_idx, col_ngay_idx,
        col_site_an_idx, col_co_cau_idx
    )
    page_break_rows = add_group_page_breaks(ws, max_row, column_indices)
    merge_group_columns(
        ws, max_row, column_indices, col_ngay_idx, col_ca_idx,
        col_ma_kh_idx, col_site_an_idx, col_ten_mon_idx, col_so_luong_idx,
        page_break_rows=page_break_rows,
    )


def choose_a4_dish_column(ws, header_row_idx, candidates):
    lookup = merged_value_lookup(ws)
    best = None
    max_sample_row = min(ws.max_row, header_row_idx + 80)

    for col in candidates:
        header_text = clean_header(ws.cell(header_row_idx, col).value)
        nonempty = 0
        generic = 0
        distinct = set()
        for row in range(header_row_idx + 1, max_sample_row + 1):
            value = effective_cell_value(ws, lookup, row, col)
            value_key = text_key(value)
            if not value_key:
                continue
            nonempty += 1
            distinct.add(value_key)
            if is_generic_meal_label(value):
                generic += 1
        header_bonus = 8 if "ten mon" in header_text else 0
        score = nonempty + len(distinct) + header_bonus - generic * 4
        option = (score, col)
        if best is None or option > best:
            best = option

    return best[1] if best else None


def clean_header(value):
    return text_key(value).replace("\n", " ")


def is_numeric_like(value):
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    return bool(re.fullmatch(r"-?\d+(?:[,.]\d+)?", str(value).strip()))


def choose_a4_ingredient_column(ws, header_row_idx, candidates):
    lookup = merged_value_lookup(ws)
    best = None
    max_sample_row = min(ws.max_row, header_row_idx + 80)
    expanded = {}

    for source_col in candidates:
        for col in (source_col, source_col + 1):
            if 1 <= col <= ws.max_column:
                distance = abs(col - source_col)
                expanded[col] = min(distance, expanded.get(col, distance))

    for col, distance in sorted(expanded.items()):
        header_text = clean_header(ws.cell(header_row_idx, col).value)
        nonempty = 0
        textish = 0
        numeric = 0
        short_values = 0
        distinct = set()

        for row in range(header_row_idx + 1, max_sample_row + 1):
            value = effective_cell_value(ws, lookup, row, col)
            value_key = text_key(value)
            if not value_key:
                continue
            nonempty += 1
            distinct.add(value_key)
            if is_numeric_like(value):
                numeric += 1
            elif any(ch.isalpha() for ch in value_key):
                textish += 1
                if len(value_key) <= 2:
                    short_values += 1

        header_bonus = 0
        if "ten nguyen" in header_text or "ten nvl" in header_text:
            header_bonus += 24
        elif "nguyen lieu" in header_text or "nguyen vat lieu" in header_text or "nvl" in header_text:
            header_bonus += 8

        header_penalty = 0
        if any(pattern in header_text for pattern in ("ma nvl", "ma nguyen", "stt", "so thu tu", "no.", "id")):
            header_penalty += 30
        if any(pattern in header_text for pattern in ("dinh muc", "dinh luong", "don vi", "dvt", "so luong", "kl ")):
            header_penalty += 30

        score = (
            header_bonus
            + nonempty
            + textish * 5
            + len(distinct) * 2
            - numeric * 8
            - short_values * 2
            - header_penalty
            - distance
        )
        option = (score, textish, len(distinct), -distance, -col)
        if best is None or option > best[0]:
            best = (option, col)

    return best[1] if best else None


def find_a4_source_columns(ws):
    def matches(header_text, patterns, excluded=()):
        return any(pattern in header_text for pattern in patterns) and not any(
            pattern in header_text for pattern in excluded
        )

    required_keys = {
        "date",
        "customer",
        "shift",
        "site",
        "structure",
        "dish",
        "quantity",
        "ingredient",
        "norm",
        "unit",
        "required",
        "purchase_unit",
    }
    optional_keys = {"approved", "committed_norm"}

    for row in range(1, min(ws.max_row, 10) + 1):
        columns = {}
        dish_candidates = []
        ingredient_candidates = []
        for col in range(1, ws.max_column + 1):
            header_text = clean_header(ws.cell(row, col).value)
            if not header_text:
                continue
            if matches(header_text, ("ngay", "date")):
                columns.setdefault("date", col)
            elif matches(header_text, ("khach hang", "ma kh", "ma khach hang"), ("site",)):
                columns.setdefault("customer", col)
            elif matches(header_text, ("co cau", "nhom mon")):
                columns.setdefault("structure", col)
            elif header_text in {"ca", "shift"} or header_text.startswith("ca "):
                columns.setdefault("shift", col)
            elif matches(header_text, ("site an", "site")):
                columns.setdefault("site", col)
            elif matches(header_text, ("mon an", "ten mon"), ("co cau", "nhom mon", "loai mon")):
                dish_candidates.append(col)
            elif matches(header_text, ("so luong", "qty", "quantity")):
                columns.setdefault("quantity", col)
            elif matches(header_text, ("nguyen vat lieu", "nguyen lieu", "nvl")):
                ingredient_candidates.append(col)
            elif matches(
                header_text,
                ("dinh muc cam ket", "dm cam ket", "cam ket"),
                ("yeu cau", "di cho", "mua hang"),
            ):
                columns.setdefault("committed_norm", col)
            elif matches(
                header_text,
                ("dinh muc", "dinh luong"),
                ("cam ket", "yeu cau", "di cho", "mua hang"),
            ):
                columns.setdefault("norm", col)
            elif matches(
                header_text,
                ("don vi tinh", "dvt", "dvt"),
                ("mua hang", "di cho", "yeu cau"),
            ):
                columns.setdefault("unit", col)
            elif matches(header_text, ("yeu cau san xuat", "kl yeu cau", "khoi luong di cho")):
                columns.setdefault("required", col)
            elif matches(header_text, ("don vi tinh mua hang", "dvt mua hang", "don vi mua hang")):
                columns.setdefault("purchase_unit", col)
            elif matches(header_text, ("duyet di cho", "kl duyet")):
                columns.setdefault("approved", col)

        if dish_candidates:
            columns["dish"] = choose_a4_dish_column(ws, row, dish_candidates)
        if ingredient_candidates:
            columns["ingredient"] = choose_a4_ingredient_column(ws, row, ingredient_candidates)

        if required_keys.issubset(columns):
            for key in optional_keys:
                columns.setdefault(key, None)
            return row, columns

    missing = ", ".join(sorted(required_keys))
    raise ValueError(f"Không tìm thấy đủ cột nguồn để xuất Format 1 A4: {missing}.")




def collect_a4_records(ws):
    header_row_idx, columns = find_a4_source_columns(ws)
    lookup = merged_value_lookup(ws)
    records = []
    previous = {key: None for key in ("date", "customer", "shift", "site", "structure", "dish", "quantity")}

    for source_row in range(header_row_idx + 1, ws.max_row + 1):
        raw = {
            key: effective_cell_value(ws, lookup, source_row, col)
            for key, col in columns.items()
        }
        for key in previous:
            raw[key] = normalized_value(raw.get(key), previous[key])
            previous[key] = raw[key]

        values = [
            raw["date"],
            raw["customer"],
            raw["shift"],
            raw["site"],
            meal_type_label(raw.get("structure")),
            raw["dish"],
            parse_print_number(raw["quantity"]),
            parse_print_number(raw.get("committed_norm")),
            clean_ingredient_name(raw["ingredient"]),
            parse_print_number(raw["norm"]),
            raw["unit"],
            parse_print_number(raw["required"]),
            raw["purchase_unit"],
            parse_print_number(raw.get("approved")),
        ]
        if a4_row_empty(values):
            continue
        records.append(
            {
                "values": values,
                "meal_order": get_co_cau_order(raw.get("structure"), raw.get("customer")),
                "source_index": len(records),
            }
        )

    return sort_a4_records(records)


def sort_a4_records(records):
    groups = []
    start = 0
    while start < len(records):
        group_key = tuple(text_key(records[start]["values"][idx]) for idx in range(4))
        end = start + 1
        while end < len(records):
            next_key = tuple(text_key(records[end]["values"][idx]) for idx in range(4))
            if next_key != group_key:
                break
            end += 1
        group = records[start:end]
        group.sort(key=lambda record: (record["meal_order"], record["source_index"]))
        first_values = group[0]["values"]
        groups.append(
            {
                "date": first_values[0],
                "customer": first_values[1],
                "shift": first_values[2],
                "site": first_values[3],
                "source_index": group[0]["source_index"],
                "rows": group,
            }
        )
        start = end

    sorted_records = []
    start = 0
    while start < len(groups):
        outer_key = tuple(text_key(groups[start][key]) for key in ("date", "customer", "shift"))
        end = start + 1
        while end < len(groups):
            next_outer_key = tuple(text_key(groups[end][key]) for key in ("date", "customer", "shift"))
            if next_outer_key != outer_key:
                break
            end += 1

        outer_groups = groups[start:end]
        if "lixil" in text_key(outer_groups[0]["customer"]):
            outer_groups.sort(
                key=lambda group: (
                    get_lixil_site_order(group["site"]),
                    text_key(group["site"]),
                    group["source_index"],
                )
            )
        for group in outer_groups:
            sorted_records.extend(group["rows"])
        start = end
    return [record["values"] for record in sorted_records]


def a4_date_title(rows):
    dates = []
    seen = set()
    for row in rows:
        date_value = row[0] if row else None
        dt_val = parse_date_value(date_value)
        text = format_date(dt_val) if dt_val else str(date_value or "").strip()
        if text and text not in seen:
            seen.add(text)
            dates.append((dt_val or date_value, text))

    if not dates:
        return "Ngày:"
    if len(dates) == 1:
        return f"Ngày: {dates[0][1]}"

    sortable_dates = [item for item in dates if isinstance(item[0], (datetime, date))]
    if len(sortable_dates) == len(dates):
        sortable_dates.sort(key=lambda item: item[0])
        return f"Ngày: {sortable_dates[0][1]} - {sortable_dates[-1][1]}"
    return "Ngày: " + ", ".join(text for _, text in dates[:3])


def committed_norm_fill(value):
    key = text_key(value)
    if "chua co dinh muc" in key:
        return PatternFill("solid", fgColor="FFC7CE")
    if "chua duyet" in key:
        return PatternFill("solid", fgColor="FFF2CC")
    if "da duyet" in key:
        return PatternFill("solid", fgColor="C6EFCE")
    return None


def apply_a4_base_format(ws, last_row):
    thin_gray = Side(style="thin", color="808080")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    widths = [12, 6.5, 9, 9.5, 18, 7, 9, 22, 7, 7.5, 10.5, 10, 10.5]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for index in range(A4_LAST_COL + 1, 27):
        ws.column_dimensions[get_column_letter(index)].hidden = True

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=A4_LAST_COL)
    date_cell = ws.cell(1, 1)
    date_cell.font = Font(name="Arial", size=9, bold=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    date_cell.border = border
    ws.row_dimensions[1].height = 22

    for cell in ws[A4_HEADER_ROW]:
        cell.fill = header_fill
        cell.font = Font(name="Arial", size=8, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[A4_HEADER_ROW].height = 32

    center_cols = {1, 2, 3, 4, 7, 10, 12}
    right_cols = {6, 9, 11, 13}
    for row in range(A4_DATA_START_ROW, last_row + 1):
        longest = max(len(str(ws.cell(row, col).value or "")) for col in (5, 8))
        if longest > 85:
            ws.row_dimensions[row].height = 42
        elif longest > 55:
            ws.row_dimensions[row].height = 32
        elif longest > 35:
            ws.row_dimensions[row].height = 24

        for col in range(1, A4_LAST_COL + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Arial", size=8)
            cell.border = border
            if col in right_cols:
                horizontal = "right"
            elif col in center_cols:
                horizontal = "center"
            else:
                horizontal = "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)

        ws.cell(row, 6).number_format = "0"
        for col in (9, 11, 13):
            value = ws.cell(row, col).value
            ws.cell(row, col).number_format = "0.###" if isinstance(value, float) and not value.is_integer() else "0"

        status_fill = committed_norm_fill(ws.cell(row, 7).value)
        if status_fill:
            ws.cell(row, 7).fill = status_fill


def same_a4_merge_key(rows, row_a, row_b, key_col, parent_cols):
    values_a = rows[row_a - A4_DATA_START_ROW]
    values_b = rows[row_b - A4_DATA_START_ROW]
    if text_key(values_a[key_col - 1]) == "":
        return False
    if text_key(values_a[key_col - 1]) != text_key(values_b[key_col - 1]):
        return False
    return all(text_key(values_a[parent - 1]) == text_key(values_b[parent - 1]) for parent in parent_cols)


def merge_a4_runs(ws, rows, sheet_col, key_col, parent_cols=()):
    if not rows:
        return
    start = A4_DATA_START_ROW
    last = len(rows) + A4_DATA_START_ROW - 1
    for row in range(A4_DATA_START_ROW + 1, last + 2):
        if row <= last and same_a4_merge_key(rows, row - 1, row, key_col, parent_cols):
            continue
        if row - 1 > start:
            ws.merge_cells(start_row=start, start_column=sheet_col, end_row=row - 1, end_column=sheet_col)
            ws.cell(start, sheet_col).alignment = Alignment(
                horizontal="left" if sheet_col == 5 else "center",
                vertical="center",
                wrap_text=True,
            )
        start = row


def configure_a4_print(ws, last_row):
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:2"
    ws.print_area = f"A1:{get_column_letter(A4_LAST_COL)}{last_row}"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def process_sheet_format1_a4(ws_source):
    rows = collect_a4_records(ws_source)
    if not rows:
        raise ValueError("File nguồn không có dữ liệu để xuất Format 1.")

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Mau A4"
    ws.append([a4_date_title(rows)])
    ws.append(A4_OUTPUT_HEADERS)
    for row in rows:
        ws.append(row[1:])

    last_row = len(rows) + A4_DATA_START_ROW - 1
    apply_a4_base_format(ws, last_row)
    merge_a4_runs(ws, rows, 1, 2, (1,))
    merge_a4_runs(ws, rows, 2, 3, (1, 2))
    merge_a4_runs(ws, rows, 3, 4, (1, 2, 3))
    merge_a4_runs(ws, rows, 4, 5, (1, 2, 3, 4))
    merge_a4_runs(ws, rows, 5, 6, (1, 2, 3, 4, 5))
    configure_a4_print(ws, last_row)
    return workbook
