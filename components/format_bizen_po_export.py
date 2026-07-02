import os
from copy import copy
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter

from .common import merged_value_lookup, effective_cell_value, input_dir, output_dir

BIZEN_EXPORT_IDENTIFIER = "BIZEN - CATERING_Xuất PO ra ex"
BIZEN_EXPORT_FORMAT_MODE = "bizen_po_export"

# Headers required for detection
_REQUIRED_HEADERS_ALT_1 = {
    "Ngày",
    "Mã PO",
    "Nhà cung cấp",
    "Mã nguyên vật liệu",
    "Tên nguyên vật liệu",
    "Đơn vị tính",
    "Tổng Khối lượng đặt hàng",
    "Tổng Đơn giá",
    "Tổng thành tiền",
    "Khách hàng",
    "Ca",
    "Nơi giao hàng (Khách hàng)",
    "Giờ giao hàng"
}

_REQUIRED_HEADERS_ALT_2 = {
    "Ngày",
    "Mã PO",
    "Nhà cung cấp",
    "Mã nguyên vật liệu",
    "Tên nguyên vật liệu",
    "Đơn vị tính",
    "Sum: Khối lượng đặt hàng",
    "Sum: Đơn giá",
    "Sum: Thành tiền",
    "Khách hàng",
    "Ca",
    "Nơi giao hàng (Khách hàng)",
    "Giờ giao hàng"
}

def can_process_bizen_po_export(ws):
    headers = set()
    for c in range(1, min(ws.max_column + 1, 50)):
        val = ws.cell(1, c).value
        if val:
            headers.add(str(val).strip())
    return _REQUIRED_HEADERS_ALT_1.issubset(headers) or _REQUIRED_HEADERS_ALT_2.issubset(headers)

def _parse_currency(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("₫", "").replace("đ", "").replace("VND", "").replace("vnd", "").strip()
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) == 2 and len(parts[1]) == 3:
            text = text.replace(".", "")
        elif len(parts) > 2:
            text = text.replace(".", "")
    try:
        num = float(text)
        return int(num) if num == int(num) else num
    except (ValueError, OverflowError):
        return value

def _parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        num = float(text)
        return int(num) if num == int(num) else num
    except ValueError:
        return value

def _detect_source_columns(ws):
    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            header_map[str(val).strip()] = c

    mapping = {
        "ngay": header_map.get("Ngày"),
        "ma_po": header_map.get("Mã PO"),
        "nha_cung_cap": header_map.get("Nhà cung cấp"),
        "ma_hang": header_map.get("Mã nguyên vật liệu"),
        "dien_giai": header_map.get("Tên nguyên vật liệu"),
        "khoi_luong": header_map.get("Tổng Khối lượng đặt hàng") or header_map.get("Sum: Khối lượng đặt hàng"),
        "don_vi": header_map.get("Đơn vị tính"),
        "don_gia": header_map.get("Tổng Đơn giá") or header_map.get("Sum: Đơn giá"),
        "thanh_tien": header_map.get("Tổng thành tiền") or header_map.get("Sum: Thành tiền"),
        "khach_hang": header_map.get("Khách hàng"),
        "ca_an": header_map.get("Ca"),
        "noi_giao": header_map.get("Nơi giao hàng (Khách hàng)"),
        "gio_giao": header_map.get("Giờ giao hàng"),
        "thue_vat": header_map.get("Sum: Thuế VAT") or header_map.get("Thuế VAT") or header_map.get("Tổng Thuế VAT"),
        "thanh_tien_sau_thue": header_map.get("Sum: Thành tiền sau thuế") or header_map.get("Thành tiền sau thuế") or header_map.get("Tổng tiền Sau thuế"),
    }
    
    required = ["ma_po", "ma_hang", "dien_giai", "don_vi", "khoi_luong", "don_gia", "thanh_tien"]
    missing = [k for k in required if mapping[k] is None]
    if missing:
        raise ValueError("Không tìm thấy các cột bắt buộc: " + ", ".join(missing))
    return mapping

def process_sheet_bizen_po_export(ws):
    col_map = _detect_source_columns(ws)
    lookup = merged_value_lookup(ws)

    # Đọc dữ liệu và gom nhóm theo Mã PO
    grouped_data = defaultdict(list)
    
    for r in range(2, ws.max_row + 1):
        ma_po = effective_cell_value(ws, lookup, r, col_map["ma_po"])
        if not ma_po:
            continue
            
        ngay = effective_cell_value(ws, lookup, r, col_map["ngay"])
        nha_cung_cap = effective_cell_value(ws, lookup, r, col_map["nha_cung_cap"])
        ma_hang = effective_cell_value(ws, lookup, r, col_map["ma_hang"])
        dien_giai = effective_cell_value(ws, lookup, r, col_map["dien_giai"])
        khoi_luong_raw = effective_cell_value(ws, lookup, r, col_map["khoi_luong"])
        don_vi = effective_cell_value(ws, lookup, r, col_map["don_vi"])
        don_gia_raw = effective_cell_value(ws, lookup, r, col_map["don_gia"])
        thanh_tien_raw = effective_cell_value(ws, lookup, r, col_map["thanh_tien"])
        khach_hang = effective_cell_value(ws, lookup, r, col_map["khach_hang"])
        ca_an = effective_cell_value(ws, lookup, r, col_map["ca_an"])
        noi_giao = effective_cell_value(ws, lookup, r, col_map["noi_giao"])
        gio_giao = effective_cell_value(ws, lookup, r, col_map["gio_giao"])
        
        # Đọc Thuế VAT và Thành tiền sau thuế
        thue_vat_raw = effective_cell_value(ws, lookup, r, col_map["thue_vat"]) if col_map.get("thue_vat") else 0
        thanh_tien_sau_thue_raw = effective_cell_value(ws, lookup, r, col_map["thanh_tien_sau_thue"]) if col_map.get("thanh_tien_sau_thue") else None
        
        khoi_luong = _parse_number(khoi_luong_raw)
        don_gia = _parse_currency(don_gia_raw)
        thanh_tien = _parse_currency(thanh_tien_raw)
        thue_vat = _parse_number(thue_vat_raw)
        thanh_tien_sau_thue = thanh_tien if thanh_tien_sau_thue_raw is None else _parse_currency(thanh_tien_sau_thue_raw)
        
        grouped_data[ma_po].append({
            "ngay": ngay,
            "ma_po": ma_po,
            "nha_cung_cap": nha_cung_cap,
            "ma_hang": ma_hang,
            "dien_giai": dien_giai,
            "khoi_luong": khoi_luong,
            "don_vi": don_vi,
            "don_gia": don_gia,
            "thanh_tien": thanh_tien,
            "khach_hang": khach_hang,
            "ca_an": ca_an,
            "noi_giao": noi_giao,
            "gio_giao": gio_giao,
            "thue_vat": thue_vat,
            "thanh_tien_sau_thue": thanh_tien_sau_thue
        })
        
    if not grouped_data:
        raise ValueError("Không có dữ liệu PO.")

    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "file định dạng fotmat chuyển đổi chính.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Không tìm thấy file mẫu tại {template_path}")

    tpl_wb = load_workbook(template_path)
    
    # Output workbook
    out_wb = Workbook()
    
    # 1. Summary sheet as the first sheet
    summary_ws = out_wb.active
    summary_ws.title = "Tổng hợp PO trong ngày"
    summary_ws.views.sheetView[0].showGridLines = True
    
    tpl_summary_ws = tpl_wb['Tổng hợp PO trong ngày']
    
    # Helper to copy cell styles
    def _copy_cell_style(src, dst, copy_fill=True, copy_border=True):
        dst.font = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        if copy_fill:
            dst.fill = copy(src.fill)
        if copy_border:
            dst.border = copy(src.border)

    # Helper to copy sheet layout basics (widths, row heights, values, styles)
    def _copy_sheet_basics(src_ws, dst_ws, copy_rows_count=8, copy_fill_border=True):
        for col_idx in range(1, src_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in src_ws.column_dimensions:
                dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
                
        for r in range(1, copy_rows_count + 1):
            if r in src_ws.row_dimensions:
                dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height
                
            for c in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(r, c)
                dst_cell = dst_ws.cell(r, c, value=src_cell.value)
                if src_cell.has_style:
                    _copy_cell_style(src_cell, dst_cell, copy_fill=copy_fill_border, copy_border=copy_fill_border)
                    
        for merged_range in src_ws.merged_cells.ranges:
            if merged_range.min_row <= copy_rows_count:
                dst_ws.merge_cells(str(merged_range))

    # Copy template summary header rows (1-4)
    _copy_sheet_basics(tpl_summary_ws, summary_ws, copy_rows_count=4)

    # Border and Fill style definitions for TableStyle-like appearance
    thin_side = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    row_fill_light = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    number_format_thousands = '#,##0'
    number_format_decimal = '#,##0.00'
    
    import datetime
    
    # Apply header formatting to summary sheet headers (Row 4)
    for c in range(1, 7):
        cell = summary_ws.cell(row=4, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    # 2. Process each PO into separate sheets
    summary_rows_info = []
    unique_pos = sorted(grouped_data.keys())
    
    for idx, ma_po in enumerate(unique_pos, start=1):
        sheet_name = f"PO{idx}"
        po_ws = out_wb.create_sheet(title=sheet_name)
        po_ws.views.sheetView[0].showGridLines = True
        
        tpl_pot_ws = tpl_wb['POT1']
        _copy_sheet_basics(tpl_pot_ws, po_ws, copy_rows_count=8)
        
        # Apply header formatting to detailed PO sheet headers (Row 8)
        for c in range(1, 11):
            cell = po_ws.cell(row=8, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border
            
        rows = grouped_data[ma_po]
        
        # Sort rows: Khách hàng (A-Z) and then Ca (Ca 1 -> Ca 2 -> Ca 3 -> ...)
        from components.common import normalize_comp, normalize_shift, get_shift_order
        rows.sort(key=lambda r: (
            normalize_comp(r["khach_hang"]), 
            get_shift_order(normalize_shift(r["ca_an"])),
            str(r["ma_hang"] or "")
        ))
        
        first_row = rows[0]
        
        # Write metadata in row 4, 5, 6 under column E
        # Ngày
        ngay_val = first_row["ngay"]
        cell_ngay = po_ws.cell(row=4, column=5, value=ngay_val)
        cell_ngay.font = Font(name="Calibri", size=12, bold=True)
        cell_ngay.alignment = Alignment(horizontal="left", vertical="center")
        if isinstance(ngay_val, datetime.datetime):
            cell_ngay.number_format = 'DD/MM/YYYY'
            
        # Mã PO
        cell_po = po_ws.cell(row=5, column=5, value=first_row["ma_po"])
        cell_po.font = Font(name="Calibri", size=12, bold=True)
        cell_po.alignment = Alignment(horizontal="left", vertical="center")
        
        # Nhà cung cấp
        cell_ncc = po_ws.cell(row=6, column=5, value=first_row["nha_cung_cap"])
        cell_ncc.font = Font(name="Calibri", size=12, bold=True)
        cell_ncc.alignment = Alignment(horizontal="left", vertical="center")
        
        # Merge value cells for metadata
        po_ws.merge_cells("E4:J4")
        po_ws.merge_cells("E5:J5")
        po_ws.merge_cells("E6:J6")
        
        # Write PO detail items
        start_row = 9
        for i, r_data in enumerate(rows):
            r_idx = start_row + i
            
            qty = r_data["khoi_luong"] or 0
            raw_amt = r_data["thanh_tien"] or 0
            # Correct summed unit price from pivot table
            if qty > 0:
                price = raw_amt / qty
            else:
                price = r_data["don_gia"] or 0
                
            values = [
                r_data["ma_hang"],
                r_data["dien_giai"],
                qty,
                r_data["don_vi"],
                price,
                f"=C{r_idx}*E{r_idx}",  # Thành tiền formula
                r_data["khach_hang"],
                r_data["ca_an"],
                r_data["noi_giao"],
                r_data["gio_giao"]
            ]
            
            # Alternating background fill for rows
            current_fill = row_fill_light if i % 2 == 1 else row_fill_white
            
            for col_idx, val in enumerate(values, start=1):
                cell = po_ws.cell(row=r_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=11)
                cell.border = thin_border
                cell.fill = current_fill
                
                # Column alignments and formatting
                if col_idx in {1, 4, 7, 8, 9, 10}:
                    cell.alignment = align_center
                elif col_idx == 2:
                    cell.alignment = align_left
                elif col_idx in {3, 5, 6}:
                    cell.alignment = align_right
                    if col_idx == 3:
                        if isinstance(qty, float) and not qty.is_integer():
                            cell.number_format = number_format_decimal
                        else:
                            cell.number_format = number_format_thousands
                    elif col_idx in {5, 6}:
                        cell.number_format = number_format_thousands
                        
        N = 8 + len(rows)
        
        # Calculate PO VAT rate based on the actual lines
        rates = []
        for r_data in rows:
            before = r_data["thanh_tien"] or 0
            net = r_data["thanh_tien_sau_thue"] or before
            if before > 0:
                rates.append((net - before) / before)
        po_vat_rate = max(rates) if rates else 0.0
        po_vat_rate = round(po_vat_rate, 2)
        
        # Helper to write total rows with styling
        def write_po_total_row(ws, r_idx, label, val_or_form, is_pct=False):
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=9)
            cell_lbl = ws.cell(row=r_idx, column=1, value=label)
            cell_lbl.font = Font(name="Calibri", size=11, bold=True)
            cell_lbl.alignment = align_center
            
            for c in range(1, 10):
                c_cell = ws.cell(row=r_idx, column=c)
                c_cell.border = thin_border
                c_cell.fill = total_fill
                
            cell_val = ws.cell(row=r_idx, column=10, value=val_or_form)
            cell_val.font = Font(name="Calibri", size=11, bold=True)
            cell_val.alignment = align_right
            cell_val.border = thin_border
            cell_val.fill = total_fill
            if is_pct:
                cell_val.number_format = '0%'
            else:
                cell_val.number_format = number_format_thousands

        # Write PO totals at the bottom of the table
        write_po_total_row(po_ws, N + 1, "Tổng tiền", f"=SUM(F9:F{N})")
        write_po_total_row(po_ws, N + 2, "Thuế suất thuế GTGT", po_vat_rate, is_pct=True)
        write_po_total_row(po_ws, N + 3, "Tiền thuế GTGT", f"=J{N+1}*J{N+2}")
        write_po_total_row(po_ws, N + 4, "Tổng tiền thanh toán", f"=J{N+1}+J{N+3}")
        
        # Add Excel Table (Ctrl+T) for detailed sheet table (A8:J{N})
        from openpyxl.worksheet.table import Table, TableStyleInfo
        table_name = f"POTable_{idx}"
        tab = Table(displayName=table_name, ref=f"A8:J{N}")
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        po_ws.add_table(tab)
        
        # Auto-fit column widths
        for col_idx in range(1, 11):
            max_len = 0
            for row_idx in range(1, N + 1):
                cell_val = po_ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    if isinstance(cell_val, (int, float)):
                        display_len = len("{:,.0f}".format(cell_val))
                    else:
                        display_len = len(str(cell_val))
                    if display_len > max_len:
                        max_len = display_len
            adjusted_width = max_len + 4
            if col_idx in [2, 9]:
                adjusted_width = max(20, min(adjusted_width, 60))
            else:
                adjusted_width = max(10, min(adjusted_width, 40))
            po_ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
        # Freeze panes & Auto-filter on detailed table
        po_ws.freeze_panes = "A9"
        po_ws.auto_filter.ref = f"A8:J{N}"
        
        summary_rows_info.append({
            "sheet_name": sheet_name,
            "total_row": N + 1,
            "vat_row": N + 3,
            "grand_total_row": N + 4
        })

    # 3. Populate summary sheet with formulas linking to each POT sheet
    for idx, info in enumerate(summary_rows_info, start=1):
        r_idx = 4 + idx
        s_name = info["sheet_name"]
        
        current_fill = row_fill_light if idx % 2 == 1 else row_fill_white
        
        # Col A: Ngày
        cell_ngay = summary_ws.cell(row=r_idx, column=1, value=f"='{s_name}'!E4")
        cell_ngay.font = Font(name="Calibri", size=11)
        cell_ngay.alignment = align_center
        cell_ngay.border = thin_border
        cell_ngay.fill = current_fill
        cell_ngay.number_format = 'DD/MM/YYYY'
        
        # Col B: Mã Po
        cell_po = summary_ws.cell(row=r_idx, column=2, value=f"='{s_name}'!E5")
        cell_po.font = Font(name="Calibri", size=11)
        cell_po.alignment = align_center
        cell_po.border = thin_border
        cell_po.fill = current_fill
        
        # Col C: Nhà cung cấp
        cell_ncc = summary_ws.cell(row=r_idx, column=3, value=f"='{s_name}'!E6")
        cell_ncc.font = Font(name="Calibri", size=11)
        cell_ncc.alignment = align_left
        cell_ncc.border = thin_border
        cell_ncc.fill = current_fill
        
        # Col D: Thành tiền
        cell_amt = summary_ws.cell(row=r_idx, column=4, value=f"='{s_name}'!J{info['total_row']}")
        cell_amt.font = Font(name="Calibri", size=11)
        cell_amt.alignment = align_right
        cell_amt.border = thin_border
        cell_amt.fill = current_fill
        cell_amt.number_format = number_format_thousands
        
        # Col E: Thuế VAT
        cell_vat = summary_ws.cell(row=r_idx, column=5, value=f"='{s_name}'!J{info['vat_row']}")
        cell_vat.font = Font(name="Calibri", size=11)
        cell_vat.alignment = align_right
        cell_vat.border = thin_border
        cell_vat.fill = current_fill
        cell_vat.number_format = number_format_thousands
        
        # Col F: Thành tiền sau thuế 
        cell_net = summary_ws.cell(row=r_idx, column=6, value=f"='{s_name}'!J{info['grand_total_row']}")
        cell_net.font = Font(name="Calibri", size=11)
        cell_net.alignment = align_right
        cell_net.border = thin_border
        cell_net.fill = current_fill
        cell_net.number_format = number_format_thousands

    # Summary Total Row
    M = 4 + len(summary_rows_info)
    total_row_idx = M + 1
    
    # Add Excel Table (Ctrl+T) for summary sheet table (A4:F{M})
    from openpyxl.worksheet.table import Table, TableStyleInfo
    sum_tab = Table(displayName="SummaryTable", ref=f"A4:F{M}")
    sum_style = TableStyleInfo(
        name="TableStyleMedium2", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    sum_tab.tableStyleInfo = sum_style
    summary_ws.add_table(sum_tab)

    for col_idx in range(1, 7):
        cell = summary_ws.cell(row=total_row_idx, column=col_idx)
        cell.border = thin_border
        cell.fill = total_fill
        
    cell_lbl = summary_ws.cell(row=total_row_idx, column=1, value="Tổng")
    cell_lbl.font = Font(name="Calibri", size=11, bold=True)
    cell_lbl.alignment = align_center
    
    cell_count_po = summary_ws.cell(row=total_row_idx, column=2, value=f"=COUNTA(B5:B{M})")
    cell_count_po.font = Font(name="Calibri", size=11, bold=True)
    cell_count_po.alignment = align_center
    
    cell_amt_tot = summary_ws.cell(row=total_row_idx, column=4, value=f"=SUM(D5:D{M})")
    cell_amt_tot.font = Font(name="Calibri", size=11, bold=True)
    cell_amt_tot.alignment = align_right
    cell_amt_tot.number_format = number_format_thousands
    
    cell_vat_tot = summary_ws.cell(row=total_row_idx, column=5, value=f"=SUM(E5:E{M})")
    cell_vat_tot.font = Font(name="Calibri", size=11, bold=True)
    cell_vat_tot.alignment = align_right
    cell_vat_tot.number_format = number_format_thousands
    
    cell_net_tot = summary_ws.cell(row=total_row_idx, column=6, value=f"=SUM(F5:F{M})")
    cell_net_tot.font = Font(name="Calibri", size=11, bold=True)
    cell_net_tot.alignment = align_right
    cell_net_tot.number_format = number_format_thousands
    
    # Auto-fit columns of summary sheet
    for col_idx in range(1, 7):
        max_len = 0
        for row_idx in range(1, total_row_idx + 1):
            cell_val = summary_ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)):
                    display_len = len("{:,.0f}".format(cell_val))
                else:
                    display_len = len(str(cell_val))
                if display_len > max_len:
                    max_len = display_len
        adjusted_width = max_len + 4
        adjusted_width = max(12, min(adjusted_width, 40))
        summary_ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
    tpl_wb.close()
    return out_wb
