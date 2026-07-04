"""
Format Bizen Catering – chuyển đổi file PO từ BIZEN - CATERING sang bảng
dữ liệu gọn gàng với 12 cột:

  Mã PO | Ngày | Mã hàng | Tên nhà cung cấp | Diễn giải | Số lượng |
  Đơn giá | Đơn vị | Thành tiền | Khách hàng | Ca | Nơi giao hàng

Dữ liệu được nhóm theo Mã PO, cột Ngày chỉ hiển thị ở dòng đầu của
mỗi nhóm Mã PO.

Nhận diện file theo 2 cách:
1. Tên file chứa "BIZEN - CATERING_Đặt hàng (PO)"
2. Bất kỳ file nào có đủ các cột bắt buộc (Mã PO, Mã nguyên vật liệu,
   Tên nguyên vật liệu, Đơn vị tính, Nhà cung cấp, Đơn giá, Thành tiền)
"""

import re
from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .common import merged_value_lookup, effective_cell_value


# ---------------------------------------------------------------------------
# Tên file nhận diện (chỉ cần chứa phần trước "_Lưới" / "_quan" / …)
# ---------------------------------------------------------------------------
BIZEN_IDENTIFIER = "BIZEN - CATERING_Đặt hàng (PO)"
BIZEN_FORMAT_MODE = "bizen_po"

# Các header bắt buộc để nhận diện file theo nội dung
_REQUIRED_HEADERS = {
    "Mã PO",
    "Mã nguyên vật liệu",
    "Tên nguyên vật liệu",
    "Đơn vị tính",
}

def can_process_bizen_po(ws):
    """Kiểm tra worksheet có đủ các cột bắt buộc để xử lý format BIZEN PO."""
    headers = set()
    for c in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(1, c).value
        if val:
            headers.add(str(val).strip())
    
    has_base = _REQUIRED_HEADERS.issubset(headers)
    has_don_gia = "Đơn giá" in headers or "Sum: Đơn giá" in headers or "Tổng Đơn giá" in headers
    has_thanh_tien = "Thành tiền" in headers or "Sum: Thành tiền" in headers or "Tổng thành tiền" in headers
    return has_base and has_don_gia and has_thanh_tien

# ---------------------------------------------------------------------------
# Header đầu ra (12 cột) – đúng thứ tự yêu cầu
# ---------------------------------------------------------------------------
OUTPUT_HEADERS = [
    "Mã PO",              # 1
    "Ngày",                # 2
    "Mã hàng",             # 3
    "Tên nhà cung cấp",   # 4
    "Diễn giải",           # 5
    "Số lượng",            # 6
    "Đơn giá",             # 7
    "Đơn vị",              # 8
    "Thành tiền",          # 9
    "Khách hàng",          # 10
    "Ca",                  # 11
    "Nơi giao hàng",      # 12
]

# Column indices trong output (1-indexed)
_OUT_MA_PO = 1
_OUT_NGAY = 2
_OUT_MA_HANG = 3
_OUT_NCC = 4
_OUT_DIEN_GIAI = 5
_OUT_SO_LUONG = 6
_OUT_DON_GIA = 7
_OUT_DON_VI = 8
_OUT_THANH_TIEN = 9
_OUT_KHACH_HANG = 10
_OUT_CA = 11
_OUT_NOI_GIAO = 12

# Các cột căn giữa
_CENTER_COLS = {_OUT_MA_PO, _OUT_NGAY, _OUT_MA_HANG, _OUT_DON_VI, _OUT_KHACH_HANG, _OUT_CA}
# Các cột căn phải (số)
_RIGHT_COLS = {_OUT_SO_LUONG, _OUT_DON_GIA, _OUT_THANH_TIEN}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_currency(value):
    """Loại bỏ ký hiệu tiền tệ (₫, đ, VND …) và chuyển sang số.

    Ví dụ:
        '₫65,000'   → 65000
        'đ65.000'   → 65000
        '65.000đ'   → 65000
        '₫1,100'    → 1100
        65000        → 65000   (giữ nguyên nếu đã là số)
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None

    # Loại bỏ các ký hiệu tiền tệ
    text = text.replace("₫", "").replace("đ", "").replace("VND", "").replace("vnd", "").strip()

    # Xử lý dấu phân cách
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) == 2 and len(parts[1]) == 3:
            text = text.replace(".", "")
        elif len(parts) > 2:
            text = text.replace(".", "")

    try:
        num = float(text)
        return int(num) if num == int(num) else num
    except (ValueError, OverflowError):
        return value


def _parse_number(value):
    """Parse giá trị số (Số lượng, Đơn giá dạng number thuần)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        num = float(text)
        return int(num) if num == int(num) else num
    except ValueError:
        return value


def _detect_source_columns(ws):
    """Tự động phát hiện cột dựa theo header row 1.

    Trả về dict mapping tên logic → col index (1-indexed).
    Nếu không tìm thấy đủ cột bắt buộc thì raise ValueError.
    """
    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            header_map[str(val).strip()] = c

    mapping = {
        "ma_po": header_map.get("Mã PO"),                       # → Mã PO
        "ngay": header_map.get("Ngày"),                          # → Ngày
        "ma_nvl": header_map.get("Mã nguyên vật liệu"),         # → Mã hàng
        "ten_nvl": header_map.get("Tên nguyên vật liệu"),       # → Diễn giải
        "so_luong": header_map.get("Khối lượng đặt hàng")
                    or header_map.get("Số lượng")
                    or header_map.get("Sum: Khối lượng đặt hàng")
                    or header_map.get("Tổng Khối lượng đặt hàng"),
        "don_vi": header_map.get("Đơn vị tính"),                 # → Đơn vị
        "nha_cung_cap": header_map.get("Nhà cung cấp"),         # → Tên nhà cung cấp
        "don_gia": header_map.get("Đơn giá")
                   or header_map.get("Sum: Đơn giá")
                   or header_map.get("Tổng Đơn giá"),            # → Đơn giá
        "thanh_tien": header_map.get("Thành tiền")
                      or header_map.get("Sum: Thành tiền")
                      or header_map.get("Tổng thành tiền"),      # → Thành tiền
        "khach_hang": header_map.get("Khách hàng"),              # → Khách hàng
        "ca": header_map.get("Ca")                               # → Ca
              or header_map.get("Ca ăn"),
        "noi_giao": header_map.get("Nơi giao hàng (Khách hàng)"),  # → Nơi giao hàng
    }

    required = ["ma_po", "ma_nvl", "ten_nvl", "don_vi", "nha_cung_cap", "don_gia", "thanh_tien"]
    missing = [k for k in required if mapping[k] is None]
    if missing:
        found = list(header_map.keys())
        raise ValueError(
            "Không tìm thấy các cột bắt buộc trong file nguồn: "
            + ", ".join(missing)
            + f". Các cột đã đọc được ở dòng 1 là: {found}"
        )
    return mapping


# ---------------------------------------------------------------------------
# Main processor
# ---------------------------------------------------------------------------

def process_sheet_bizen_po(ws):
    """Đọc worksheet nguồn (BIZEN PO Lưới) và trả về Workbook mới đã định dạng."""

    col_map = _detect_source_columns(ws)
    lookup = merged_value_lookup(ws)

    # Thu thập dữ liệu
    rows_data = []
    for r in range(2, ws.max_row + 1):
        ma_po = effective_cell_value(ws, lookup, r, col_map["ma_po"])
        ma_hang = effective_cell_value(ws, lookup, r, col_map["ma_nvl"])
        ten_nvl = effective_cell_value(ws, lookup, r, col_map["ten_nvl"])
        don_vi = effective_cell_value(ws, lookup, r, col_map["don_vi"])
        nha_cung_cap = effective_cell_value(ws, lookup, r, col_map["nha_cung_cap"])
        don_gia_raw = effective_cell_value(ws, lookup, r, col_map["don_gia"])
        thanh_tien_raw = effective_cell_value(ws, lookup, r, col_map["thanh_tien"])
        so_luong_raw = effective_cell_value(ws, lookup, r, col_map["so_luong"])

        # Cột Ngày
        ngay_col = col_map.get("ngay")
        ngay = effective_cell_value(ws, lookup, r, ngay_col) if ngay_col else None

        # Các cột optional
        khach_hang_col = col_map.get("khach_hang")
        khach_hang = effective_cell_value(ws, lookup, r, khach_hang_col) if khach_hang_col else None
        if khach_hang:
            kh_str = str(khach_hang).strip()
            if kh_str.lower() in ("osv", "odsv"):
                khach_hang = "ODSV"


        ca_col = col_map.get("ca")
        ca = effective_cell_value(ws, lookup, r, ca_col) if ca_col else None

        noi_giao_col = col_map.get("noi_giao")
        noi_giao = effective_cell_value(ws, lookup, r, noi_giao_col) if noi_giao_col else None

        # Bỏ qua dòng hoàn toàn trống
        if ma_po is None and ma_hang is None and ten_nvl is None and don_gia_raw is None:
            continue

        so_luong = _parse_number(so_luong_raw)
        don_gia = _parse_currency(don_gia_raw)
        thanh_tien = _parse_currency(thanh_tien_raw)

        rows_data.append({
            "ma_po": ma_po,
            "ngay": ngay,
            "ma_hang": ma_hang,
            "nha_cung_cap": nha_cung_cap,
            "dien_giai": ten_nvl,
            "don_vi": don_vi,
            "so_luong": so_luong,
            "don_gia": don_gia,
            "thanh_tien": thanh_tien,
            "khach_hang": khach_hang,
            "ca": ca,
            "noi_giao": noi_giao,
        })

    if not rows_data:
        raise ValueError("Không tìm thấy dữ liệu trong file nguồn.")

    # -----------------------------------------------------------------------
    # Sắp xếp dữ liệu theo Mã PO để nhóm lại
    # -----------------------------------------------------------------------
    rows_data.sort(key=lambda row: str(row["ma_po"] or ""))

    # -----------------------------------------------------------------------
    # Xử lý cột Ngày: chỉ hiển thị ở dòng đầu tiên của mỗi nhóm Mã PO
    # -----------------------------------------------------------------------
    prev_po = None
    for row in rows_data:
        current_po = row["ma_po"]
        if current_po == prev_po:
            row["ngay_display"] = None  # để trống cho các dòng sau
        else:
            row["ngay_display"] = row["ngay"]
        prev_po = current_po

    # -----------------------------------------------------------------------
    # Tạo workbook đầu ra
    # -----------------------------------------------------------------------
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Đặt hàng"

    # ---- Styles ----------------------------------------------------------
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_side = Side(style="thin", color="B0B0B0")
    header_border = Border(
        left=Side(style="thin", color="305496"),
        right=Side(style="thin", color="305496"),
        top=Side(style="thin", color="305496"),
        bottom=Side(style="thin", color="305496"),
    )
    body_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    number_format_thousands = '#,##0'
    number_format_decimal = '#,##0.00'

    # ---- Header row ------------------------------------------------------
    for col_idx, header_text in enumerate(OUTPUT_HEADERS, start=1):
        cell = out_ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align

    # ---- Data rows -------------------------------------------------------
    for row_idx, data in enumerate(rows_data, start=2):
        values = [
            data["ma_po"],
            data["ngay_display"],
            data["ma_hang"],
            data["nha_cung_cap"],
            data["dien_giai"],
            data["so_luong"],
            data["don_gia"],
            data["don_vi"],
            data["thanh_tien"],
            data["khach_hang"],
            data["ca"],
            data["noi_giao"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = out_ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = body_border

            # Alignment theo yêu cầu
            if col_idx in _CENTER_COLS:
                cell.alignment = align_center
            elif col_idx in _RIGHT_COLS:
                cell.alignment = align_right
                # Number format
                if isinstance(val, (int, float)):
                    if isinstance(val, float) and not val.is_integer():
                        cell.number_format = number_format_decimal
                    else:
                        cell.number_format = number_format_thousands
            else:
                cell.alignment = align_left

    # ---- Merge cells cho cột Mã PO và Ngày theo nhóm Mã PO ---------------
    last_row = len(rows_data) + 1
    if last_row > 2:
        group_start = 2  # first data row
        for row_idx in range(3, last_row + 2):  # +2 to handle last group
            current_po = out_ws.cell(row=row_idx, column=_OUT_MA_PO).value if row_idx <= last_row else None
            prev_po_val = out_ws.cell(row=row_idx - 1, column=_OUT_MA_PO).value
            if current_po != prev_po_val or row_idx > last_row:
                group_end = row_idx - 1
                if group_end > group_start:
                    # Merge cột Mã PO
                    out_ws.merge_cells(
                        start_row=group_start,
                        start_column=_OUT_MA_PO,
                        end_row=group_end,
                        end_column=_OUT_MA_PO,
                    )
                    out_ws.cell(row=group_start, column=_OUT_MA_PO).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    # Merge cột Ngày
                    out_ws.merge_cells(
                        start_row=group_start,
                        start_column=_OUT_NGAY,
                        end_row=group_end,
                        end_column=_OUT_NGAY,
                    )
                    out_ws.cell(row=group_start, column=_OUT_NGAY).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                group_start = row_idx

    # ---- Auto-fit column widths ------------------------------------------
    for col_idx in range(1, len(OUTPUT_HEADERS) + 1):
        max_len = len(OUTPUT_HEADERS[col_idx - 1])
        for row_idx in range(2, last_row + 1):
            cell_val = out_ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)):
                    display_len = len("{:,.0f}".format(cell_val))
                else:
                    display_len = len(str(cell_val))
                if display_len > max_len:
                    max_len = display_len
        adjusted_width = max_len + 4
        adjusted_width = max(10, min(adjusted_width, 45))
        out_ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # ---- Auto filter trên header row ------------------------------------
    last_col_letter = get_column_letter(len(OUTPUT_HEADERS))
    out_ws.auto_filter.ref = "A1:{}{}".format(last_col_letter, last_row)

    # ---- Freeze header row -----------------------------------------------
    out_ws.freeze_panes = "A2"

    # ---- Alternating row colors (subtle) ---------------------------------
    even_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for row_idx in range(2, last_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(OUTPUT_HEADERS) + 1):
                cell = out_ws.cell(row=row_idx, column=col_idx)
                # Không ghi đè fill cho merged cells (chỉ top-left cell có value)
                if cell.value is not None or col_idx != _OUT_NGAY:
                    cell.fill = even_fill

    return out_wb
