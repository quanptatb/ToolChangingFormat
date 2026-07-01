from copy import copy
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote, unquote
import re
import csv
import unicodedata
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===== THƯ MỤC INPUT/OUTPUT =====
input_dir = Path("Excel")
output_dir = Path("Đã gộp")
max_upload_bytes = 50 * 1024 * 1024

# ===== CỘT DỮ LIỆU =====
col_ngay = 2
col_ma_kh = 3
col_ca = 4
col_so_luong = 5
col_ten_mon = 6
col_khoi_luong = 8
col_khoi_luong_di_cho = 10

source_start_row = 2
title_row = 1
header_row = 2
body_start_row = 3

title_font_size = 18
header_font_size = 14
body_font_size = 13

visible_column_widths = {
    "C": 10,
    "D": 5.5,
    "E": 6,
    "F": 15,
    "G": 18,
    "H": 7,
    "I": 8,
    "J": 9,
    "K": 5.5,
}

hidden_print_columns = ("A", "B", "L")

APPROVAL_FORMAT_MODE = "duyet_dinh_muc"

SHORT_NAMES = {
    "thịt heo nạc vai": "nạc",
    "thịt heo vai": "nạc",
    "thịt heo nạc dăm": "nạc dăm",
    "thịt xay": "xay",
    "thịt bằm": "xay",
    "trứng gà": "trứng",
    "trứng gà 300 quả/cây": "trứng",
    "cá biển chưa xác định": "cá biển",
    "thịt vai heo chưa xác định": "nạc",
    "mọc chưa xác định": "mọc",
    "cá sống chưa xác định": "cá sống",
    "tôm chưa xác định": "tôm",
    "cá basa chưa xác định": "cá basa",
    "đậu hũ chiên chưa xác định": "đậu",
    "đậu hũ trắng": "đậu",
    "thịt vịt chưa xác định": "vịt",
    "trứng cút": "cút",
    "chả cây": "chả",
    "chả lụa": "chả",
    "chả chay": "chả chay",
    "đậu hũ chiên": "đậu",
    "đậu hũ": "đậu",
    "đậu hũ da": "đậu da",
    "đậu cove": "ve",
    "cà rốt": "rốt",
    "nạm bò": "nạm",
    "bò nạm": "nạm",
    "giò heo": "giò",
    "bông cải": "bông cải",
    "sườn heo": "sườn heo",
    "dưa chua": "dưa chua",
    "mọc": "mọc",
    "tôm size 100": "tôm",
    "tôm size 50": "tôm",
    "tôm thẻ size 100 con": "tôm",
    "tôm thẻ size 60 con": "tôm",
    "tôm khô": "tôm khô",
    "măng chua": "măng",
    "trứng": "trứng",
    "giá": "giá",
    "rau dền": "rau dền",
    "rau ngót": "rau ngót",
    "cà tím": "cà tím",
    "hành tây": "hành tây",
    "cá thu": "cá thu",
    "má gà": "gà",
    "má gà tươi": "gà",
    "gà ta": "gà ta",
}

CUSTOMER_SHORT_NAMES = {
    "mon an sang om": "Ăn sáng OM",
    "scavi com khach 50k": "SCAVI 50K",
    "scavi com khach 70k": "SCAVI 70K",
    "sofa com quan ly 40k": "SOFA QL",
}


def ensure_visible_worksheet(workbook, preferred=None):
    """Make an output workbook valid even when its source sheets were hidden."""
    worksheets = workbook.worksheets
    if not worksheets:
        raise ValueError("File Excel không có trang tính dữ liệu để xử lý.")

    visible_sheets = [ws for ws in worksheets if ws.sheet_state == "visible"]
    if visible_sheets:
        active_sheet = preferred if preferred in visible_sheets else visible_sheets[0]
    else:
        active_sheet = preferred if preferred in worksheets else worksheets[0]
        active_sheet.sheet_state = "visible"

    workbook.active = active_sheet
    return active_sheet


def normalized_value(raw_value, previous_value=None):
    if raw_value is None:
        return previous_value
    if isinstance(raw_value, str) and raw_value.strip() == "":
        return previous_value
    return raw_value


def key_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return " ".join(value.split()).casefold()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def format_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    if value is None:
        return ""
    return str(value)


def extract_date_from_filename(file_path, year=None):
    if year is None:
        year = datetime.now().year
    stem = file_path.stem
    m_full = re.search(r"(\d{4})[._-](\d{1,2})[._-](\d{1,2})", stem)
    if m_full:
        return f"{m_full.group(3).zfill(2)}/{m_full.group(2).zfill(2)}/{m_full.group(1)}"
    m_full_rev = re.search(r"(\d{1,2})[._-](\d{1,2})[._-](\d{4})", stem)
    if m_full_rev:
        return f"{m_full_rev.group(1).zfill(2)}/{m_full_rev.group(2).zfill(2)}/{m_full_rev.group(3)}"
    m_short = re.search(r"[_\s](\d{1,2})[._-](\d{1,2})(?:$|[_\s])", stem)
    if not m_short:
        m_short = re.search(r"(\d{1,2})[._-](\d{1,2})", stem)
    if m_short:
        day = int(m_short.group(1))
        month = int(m_short.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return f"{str(day).zfill(2)}/{str(month).zfill(2)}/{year}"
    return None


def parse_date_value(val):
    if isinstance(val, (datetime, date)):
        return val
    if not val:
        return None
    val_str = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None


def collect_date_text(ws, start_row, max_row, col_ngay_idx=None):
    if col_ngay_idx is None:
        col_ngay_idx = col_ngay
    if col_ngay_idx is None:
        return ""
    dates = []
    seen = set()

    for row in range(start_row, max_row + 1):
        value = ws.cell(row, col_ngay_idx).value
        dt_val = parse_date_value(value)
        text = format_date(dt_val) if dt_val else str(value).strip()
        if text and text not in seen:
            seen.add(text)
            dates.append((dt_val or value, text))

    if not dates:
        return ""
    if len(dates) == 1:
        return dates[0][1]

    sortable_dates = [item for item in dates if isinstance(item[0], (datetime, date))]
    if len(sortable_dates) == len(dates):
        sortable_dates.sort(key=lambda item: item[0])
        return f"{sortable_dates[0][1]} - {sortable_dates[-1][1]}"

    return ", ".join(text for _, text in dates[:3])


def kitchen_name_from_file(file_path):
    name = file_path.stem.replace("_", " ").replace("-", " ")
    name = " ".join(name.split())
    if name.casefold().startswith("bom "):
        name = name[4:].strip()
    if name:
        name = name[:1].upper() + name[1:]
    return name or file_path.stem


def unmerge_group_columns(ws, start_row, col_ma_kh_idx=None, col_ca_idx=None, col_so_luong_idx=None, col_ten_mon_idx=None):
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row:
            try:
                ws.unmerge_cells(str(merged_range))
            except Exception:
                pass


def get_lixil_site_order(site_name):
    if not site_name:
        return 99
    s = str(site_name).lower().replace(" ", "").replace("-", "")
    if "fab1" in s:
        return 1
    if "fab2" in s:
        return 2
    if "vpc" in s:
        return 3
    return 99


def text_key(value):
    text = "" if value is None else str(value)
    text = text.replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFD", text.casefold())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return " ".join(text.split())


def get_co_cau_order(co_cau_val, customer_val=None):
    if not co_cau_val:
        return 99
    s = text_key(co_cau_val)
    order = [
        (("mon man", "suat man", "suat com", "com", "cai thien", "cai tihen", "an sang"), 1),
        (("mon nuoc", "suat nuoc", "nuoc", "chao"), 2),
        (("mon chay", "suat chay", "chay"), 3),
    ]
    for keywords, index in order:
        if any(keyword in s for keyword in keywords):
            return index
    return 99


def meal_type_label(co_cau_val):
    order = get_co_cau_order(co_cau_val)
    if order == 1:
        return "Mặn"
    if order == 2:
        return "Nước"
    if order == 3:
        return "Chay"
    return co_cau_val or ""


def is_generic_meal_label(value):
    key = text_key(value)
    return key in {
        "man",
        "nuoc",
        "chay",
        "mon man",
        "mon nuoc",
        "mon chay",
        "suat man",
        "suat nuoc",
        "suat chay",
    }


def save_workbook(wb, output_file):
    ensure_visible_worksheet(wb)
    try:
        wb.save(output_file)
        return output_file
    except PermissionError:
        for index in range(1, 100):
            suffix = "mới" if index == 1 else f"mới {index}"
            fallback_file = output_file.with_name(
                f"{output_file.stem} - {suffix}{output_file.suffix}"
            )
            if not fallback_file.exists():
                wb.save(fallback_file)
                return fallback_file
        raise


def safe_filename(filename):
    filename = (filename or "file.xlsx").replace("\\", "/").split("/")[-1].strip()
    filename = filename.replace("\r", "").replace("\n", "")
    return filename or "file.xlsx"


def output_filename(filename):
    input_name = safe_filename(filename)
    suffix = Path(input_name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        suffix = ".xlsx"
    return f"{Path(input_name).stem} - da dinh dang{suffix}"


def output_filename_for_format(filename, format_mode):
    input_name = safe_filename(filename)
    stem = Path(input_name).stem
    if format_mode == "format2":
        return f"{stem} - giay di cho.xlsx"
    if format_mode == APPROVAL_FORMAT_MODE:
        return f"{stem} - duyet dinh muc.xlsx"
    if format_mode == "bizen_po":
        return f"{stem} - bizen po.xlsx"
    return output_filename(filename)


def content_disposition_filename(filename):
    filename = safe_filename(filename)
    ascii_fallback = filename.encode("ascii", "ignore").decode("ascii") or "bom.xlsx"
    ascii_fallback = ascii_fallback.replace('"', "")
    return f'attachment; filename="{ascii_fallback}"; filename*=UTF-8\'\'{quote(filename)}'


def merged_value_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = value
    return lookup


def effective_cell_value(ws, lookup, row, col):
    if col is None:
        return None
    return lookup.get((row, col), ws.cell(row, col).value)


def to_numeric(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str:
        return None
    try:
        if "." in val_str:
            return float(val_str)
        return int(val_str)
    except ValueError:
        try:
            return float(val_str.replace(",", "."))
        except ValueError:
            return val_str


def format_number_clean(value):
    value = to_numeric(value)
    if isinstance(value, bool) or value is None:
        return "" if value is None else str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value).strip()


def clean_ingredient_name(value):
    text = "" if value is None else str(value).strip()
    if not text:
        return ""

    result = []
    depth = 0
    for char in text:
        if char in "(（":
            depth += 1
            if result and not result[-1].isspace():
                result.append(" ")
            continue
        if char in ")）":
            if depth > 0:
                depth -= 1
            continue
        if depth == 0:
            result.append(char)

    return " ".join("".join(result).split())


def is_blank_cell_value(value):
    return value is None or str(value).strip() == ""


def short_customer_name(name):
    return CUSTOMER_SHORT_NAMES.get(text_key(name), name)


def concise_customer_label(names, all_names):
    missing = [name for name in all_names if name not in names]
    short_names = [short_customer_name(name) for name in names]
    short_missing = [short_customer_name(name) for name in missing]
    if len(names) == len(all_names):
        return "Tất cả KH"
    if len(names) >= 8:
        if missing and len(missing) <= 5:
            return f"{len(names)} KH (trừ {', '.join(short_missing)})"
        return f"{len(names)} KH ({', '.join(short_names[:5])}, ...)"
    return ", ".join(short_names)


def strip_accents(text):
    if not text:
        return ""
    text = text.replace("đ", "d").replace("Đ", "D")
    normalized = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in normalized if not unicodedata.combining(c))


def normalize_comp(name):
    if not name:
        return ""
    s = str(name).lower()
    s = re.sub(r'[^a-z0-9]', '', s)
    if 'omdigital' in s or 'odsv' in s or 'osv' in s:
        return 'osv'
    if 'lixil' in s:
        return 'lixil'
    if 'smc' in s:
        return 'smc'
    if 'zamil' in s:
        return 'zamil'
    if 'wacoal' in s:
        return 'wacoal'
    if 'quadrille' in s:
        return 'quadrille'
    if 'watabe' in s:
        return 'watabe'
    if 'vikyno' in s:
        return 'vikyno'
    if 'shine' in s:
        return 'shine'
    if 'catthai' in s or 'cat-thai' in s or 'cat_thai' in s:
        return 'catthai'
    if 'maspro' in s:
        return 'maspro'
    if 'medic' in s:
        return 'medic'
    if 'scavi' in s:
        if '50k' in s or '50' in s:
            return 'scavicomkhach50k'
        if '70k' in s or '70' in s:
            return 'scavicomkhach70k'
        return 'scavi'
    if 'sofa' in s:
        if 'ql' in s or 'quanly' in s or 'comql' in s or '40k' in s:
            return 'sofacomql40k'
        return 'sofa'
    if 'dona' in s:
        return 'donaquebang'
    if 'saitexd13' in s:
        return 'saitexd13'
    if 'saitex4' in s:
        return 'saitex4'
    if 'saitex6' in s:
        return 'saitex6icd'
    if 'figla' in s:
        return 'figla'
    if 'shiseido' in s or 'shisedo' in s:
        return 'shisedo7a'
    if 'jfe' in s:
        return 'jfe'
    if 'briskheat' in s or 'brishheart' in s:
        return 'brishheart'
    if 'artwell' in s:
        return 'artwell'
    return s


def normalize_shift(shift_name):
    if not shift_name:
        return ""
    s = str(shift_name).lower()
    if 'ca 1' in s or 'ca1' in s or 'canteen 1' in s or 'c1' in s:
        return 'ca1'
    if 'ca 2' in s or 'ca2' in s or 'canteen 2' in s or 'c2' in s:
        return 'ca2'
    if 'ca 3' in s or 'ca3' in s or 'c3' in s:
        return 'ca3'
    if 'ăn sáng' in s or 'ansang' in s:
        if '2' in s:
            return 'ansang2'
        return 'ansang'
    return s


def get_dish_category(dish_name, nhom_mon=None):
    d_lower = str(dish_name).lower()
    nm_lower = str(nhom_mon).lower() if nhom_mon else ""
    if 'chay' in nm_lower or 'chay' in d_lower:
        return 'chay'
    if any(x in nm_lower for x in ['tráng miệng', 'trang mieng', 'trái cây', 'trai cay']) or \
       any(x in d_lower for x in ['trái cây', 'trai cay', 'chuối', 'nhãn', 'dưa hấu', 'sữa chua', 'bánh da lợn']):
        return 'trangmieng'
    if any(x in nm_lower for x in ['nước', 'nuoc', 'cháo', 'chao']):
        return 'nuoc'
    if any(d_lower.startswith(x) for x in ['bún ', 'mì ', 'hủ tíu', 'canh bún', 'bánh đa', 'nui ', 'phở ', 'cháo ']):
        return 'nuoc'
    return 'main'


def get_vietnamese_weekday(date_str):
    if " - " in date_str:
        date_str = date_str.split(" - ")[0]
    try:
        dt = datetime.strptime(date_str, "%d/%m/%Y")
        weekday = dt.weekday()
        mapping = {
            0: "Thứ 2",
            1: "Thứ 3",
            2: "Thứ 4",
            3: "Thứ 5",
            4: "Thứ 6",
            5: "Thứ 7",
            6: "Chủ nhật"
        }
        return f"{mapping.get(weekday, '')} - "
    except Exception:
        return "Thứ … - "


def get_shift_order(sh_n):
    if sh_n == 'ca1': return 1
    if sh_n == 'ca2': return 2
    if sh_n == 'ca3': return 3
    if sh_n == 'ansang': return 4
    if sh_n == 'ansang2': return 5
    return 99


def get_cat_order(cat_n):
    if cat_n == 'main': return 1
    if cat_n == 'nuoc': return 2
    if cat_n == 'chay': return 3
    if cat_n == 'trangmieng': return 4
    return 99


def parse_print_number(value):
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not re.fullmatch(r"-?\d+(?:[,.]\d+)?", text):
        return value
    number = float(text.replace(",", "."))
    return int(number) if number.is_integer() else number


def a4_row_empty(values):
    return all(value is None or str(value).strip() == "" for value in values)

