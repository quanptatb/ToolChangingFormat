from datetime import datetime, timedelta
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .common import normalize_quantity_mode, QUANTITY_MODE_FORECAST
from .format2_shopping import parse_source_sheet, populate_template_workbook_fill_only


FORMAT2_UPDATE_MODE = "format2_update"


def _norm(value):
    return " ".join(str(value or "").strip().casefold().split())


def _headers(ws):
    return {_norm(cell.value): cell.column for cell in ws[1] if cell.value is not None}


def _date_key(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    if text.isdigit():
        return (datetime(1899, 12, 30) + timedelta(days=int(text))).strftime("%Y-%m-%d")
    parts = text.split("/")
    if len(parts) == 3 and all(part.isdigit() for part in parts):
        return f"{int(parts[2]):04d}-{int(parts[1]):02d}-{int(parts[0]):02d}"
    return _norm(text)


def _require(headers, names, source_label):
    missing = [name for name in names if _norm(name) not in headers]
    if missing:
        raise ValueError(f"{source_label} thiếu cột: {', '.join(missing)}")


def _key(ws, row, headers, names):
    values = []
    for index, name in enumerate(names):
        value = ws.cell(row, headers[_norm(name)]).value
        values.append(_date_key(value) if index == 0 else _norm(value))
    return tuple(values)


def _site_key(value):
    return re.sub(r"[^a-z0-9]", "", _norm(value))


def _saitex_company_for_site(customer, site):
    """Return the Format 2 company label implied by a Saitex site."""
    customer_key = _site_key(customer)
    site_text = unicodedata.normalize("NFD", _norm(site))
    site_key = re.sub(
        r"[^a-z0-9]", "", "".join(ch for ch in site_text if unicodedata.category(ch) != "Mn")
    )
    if "saitex" not in customer_key and "saitex" not in site_key:
        return None
    if "may" in site_key or "cat" in site_key:
        return "Saitex D13"
    if re.fullmatch(r"(?:saitex)?(?:site)?(?:4|5)", site_key):
        return "Saitex 4"
    if re.fullmatch(r"(?:saitex)?(?:site)?6", site_key):
        return "Saitex 6"
    return None


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _merge_vertical(ws, start_row, end_row, column):
    if end_row <= start_row:
        return
    for merged in ws.merged_cells.ranges:
        if merged.min_col == column and merged.max_col == column and not (
            merged.max_row < start_row or merged.min_row > end_row
        ):
            return
    ws.merge_cells(start_row=start_row, start_column=column, end_row=end_row, end_column=column)
    ws.cell(start_row, column).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )


def _merge_standard_vegetarian_groups(ws, sheet_name):
    panels = [(2, 3, 5)]
    panels.append((7, 7 if sheet_name == "Bếp trung tâm 2" else 8, 9 if sheet_name == "Bếp trung tâm 2" else 10))
    for shift_col, qty_col, weight_col in panels:
        for merged in list(ws.merged_cells.ranges):
            if merged.min_col != shift_col or merged.max_col != shift_col:
                continue
            label = _norm(ws.cell(merged.min_row, shift_col).value)
            if "chay" not in label:
                continue
            start_row, end_row = merged.min_row, merged.max_row
            quantities = [ws.cell(row, qty_col).value for row in range(start_row, end_row + 1)]
            nonblank = [value for value in quantities if value not in (None, "")]
            if nonblank and len(nonblank) == len(quantities) and len({_norm(value) for value in nonblank}) == 1:
                _merge_vertical(ws, start_row, end_row, qty_col)
            weights = [ws.cell(row, weight_col).value for row in range(start_row, end_row + 1)]
            if all(value in (None, "") for value in weights):
                _merge_vertical(ws, start_row, end_row, weight_col)


def _merge_central_vegetarian_table(ws):
    title_row = next(
        (
            row
            for row in range(1, ws.max_row + 1)
            if "chay" in _norm(ws.cell(row, 3).value)
        ),
        None,
    )
    if not title_row:
        return
    row = title_row + 1
    while row <= ws.max_row:
        shift = ws.cell(row, 2).value
        company = ws.cell(row, 3).value
        if shift in (None, "") and company in (None, ""):
            row += 1
            continue
        end_row = row
        while end_row + 1 <= ws.max_row:
            if _norm(ws.cell(end_row + 1, 2).value) != _norm(shift):
                break
            if _norm(ws.cell(end_row + 1, 3).value) != _norm(company):
                break
            end_row += 1
        if end_row > row:
            for column in (2, 3, 4):
                values = [ws.cell(r, column).value for r in range(row, end_row + 1)]
                if len({_norm(value) for value in values}) == 1:
                    _merge_vertical(ws, row, end_row, column)
            if all(ws.cell(r, 6).value in (None, "") for r in range(row, end_row + 1)):
                _merge_vertical(ws, row, end_row, 6)
        row = end_row + 1


def _fit_update_rows(ws, sheet_name):
    max_col = 9 if sheet_name == "Bếp trung tâm 2" else 10
    for row in range(3, ws.max_row + 1):
        max_lines = 1
        has_value = False
        for column in range(1, max_col + 1):
            cell = ws.cell(row, column)
            if isinstance(cell, MergedCell) or cell.value in (None, ""):
                continue
            has_value = True
            width = ws.column_dimensions[get_column_letter(column)].width or 10
            font_size = float(cell.font.sz or 10)
            chars_per_line = max(6, int(width * 10.5 / font_size))
            lines = sum(
                max(1, (len(part) + chars_per_line - 1) // chars_per_line)
                for part in str(cell.value).split("\n")
            )
            max_lines = max(max_lines, lines)
        if not has_value:
            continue
        target = max(18, max_lines * 11 + 5)
        limit = 78 if sheet_name == "Bếp tại chỗ" else 64
        ws.row_dimensions[row].height = min(limit, target)


def _polish_format2_update(wb):
    for ws in wb.worksheets:
        if ws.title not in {"Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"}:
            continue
        _merge_standard_vegetarian_groups(ws, ws.title)
        if ws.title == "Bếp trung tâm 2":
            _merge_central_vegetarian_table(ws)
        _fit_update_rows(ws, ws.title)
    return wb


def process_sheet_format2_update(ws_menu, ws_quantity, quantity_mode=None):
    """Dùng mẫu Format 2, menu từ LSX và loại số lượng đã chọn từ bảng tổng hợp."""
    menu_headers = _headers(ws_menu)
    quantity_headers = _headers(ws_quantity)
    base_keys = ["Ngày", "Khách hàng", "Site ăn", "Ca", "Cơ cấu suất ăn"]
    quantity_keys = base_keys + ["Cơ cấu menu"]
    use_forecast = normalize_quantity_mode(quantity_mode) == QUANTITY_MODE_FORECAST
    quantity_column = "Số lượng dự báo" if use_forecast else "Số lượng cô Nga duyệt đặt hàng"
    alternate_quantity_column = "Số lượng cô Nga duyệt đặt hàng" if use_forecast else "Số lượng dự báo"
    menu_quantity_column = "Số lượng dự báo" if use_forecast else "Số lượng cô Nga duyệt"
    alternate_menu_quantity_column = "Số lượng cô Nga duyệt" if use_forecast else "Số lượng dự báo"
    _require(menu_headers, base_keys + [menu_quantity_column], "Bảng lệnh sản xuất")
    _require(
        quantity_headers,
        quantity_keys + [quantity_column, alternate_quantity_column],
        "Bảng tổng hợp số lượng",
    )

    quantities = {}
    site_labels = {}
    for row in range(2, ws_quantity.max_row + 1):
        raw_key = _key(ws_quantity, row, quantity_headers, quantity_keys)
        key = (raw_key[0], raw_key[1], _site_key(raw_key[2]), *raw_key[3:])
        site_labels[key[2]] = ws_quantity.cell(row, quantity_headers[_norm("Site ăn")]).value
        value = ws_quantity.cell(row, quantity_headers[_norm(quantity_column)]).value
        if value in (None, ""):
            value = ws_quantity.cell(row, quantity_headers[_norm(alternate_quantity_column)]).value
        if key not in quantities or (quantities[key] in (None, "") and value not in (None, "")):
            quantities[key] = value

    expanded = Workbook()
    expanded_ws = expanded.active
    expanded_ws.title = ws_menu.title
    expanded_ws.append([cell.value for cell in ws_menu[1]])
    site_col = menu_headers[_norm("Site ăn")]
    customer_col = menu_headers[_norm("Khách hàng")]
    qty_col = menu_headers[_norm(menu_quantity_column)]
    alternate_qty_col = menu_headers.get(_norm(alternate_menu_quantity_column))
    structure_col = menu_headers.get(_norm("Cơ cấu menu"))
    purchase_weight_cols = [
        column
        for header, column in menu_headers.items()
        if any(marker in header for marker in ("kl yêu cầu", "kl yeu cau", "quy đổi", "quy doi", "kl thực tế", "kl thuc te"))
    ]
    matched = 0
    for row in range(2, ws_menu.max_row + 1):
        values = [cell.value for cell in ws_menu[row]]
        if values[qty_col - 1] in (None, "") and alternate_qty_col:
            alternate_value = values[alternate_qty_col - 1]
            if alternate_value not in (None, ""):
                values[qty_col - 1] = alternate_value
        base = _key(ws_menu, row, menu_headers, base_keys)
        raw_sites = str(ws_menu.cell(row, site_col).value or "").split(",")
        sites = [_site_key(site) for site in raw_sites if _site_key(site)]
        structures = {
            key[-1]
            for key in quantities
            if key[0] == base[0] and key[1] == base[1] and key[3] == base[3]
            and key[4] == base[4] and key[2] in sites
        }
        if structure_col:
            structures = {_norm(ws_menu.cell(row, structure_col).value)}
        current_qty = _number(values[qty_col - 1])
        candidates = []
        for structure in structures:
            site_values = [quantities.get((base[0], base[1], site, base[3], base[4], structure)) for site in sites]
            numeric = [_number(value) for value in site_values]
            missing_indexes = [index for index, value in enumerate(numeric) if value is None]
            if not missing_indexes:
                total = sum(numeric)
                if current_qty is None or abs(total - current_qty) < 0.001:
                    candidates.append((structure, site_values))
            elif len(missing_indexes) == 1 and current_qty is not None:
                residual = current_qty - sum(value for value in numeric if value is not None)
                if residual >= -0.001:
                    residual = max(0.0, residual)
                    completed_values = list(site_values)
                    completed_values[missing_indexes[0]] = residual
                    candidates.append((structure, completed_values))
        if len(candidates) == 1:
            _, site_values = candidates[0]
            for site, value in zip(sites, site_values):
                new_values = list(values)
                new_values[site_col - 1] = site_labels.get(site, new_values[site_col - 1])
                company = _saitex_company_for_site(
                    new_values[customer_col - 1], new_values[site_col - 1]
                )
                if company:
                    new_values[customer_col - 1] = company
                new_values[qty_col - 1] = value
                for column in purchase_weight_cols:
                    new_values[column - 1] = None
                expanded_ws.append(new_values)
                matched += 1
        else:
            if len(sites) == 1:
                company = _saitex_company_for_site(values[customer_col - 1], values[site_col - 1])
                if company:
                    values[customer_col - 1] = company
            expanded_ws.append(values)

    parsed_rows = parse_source_sheet(expanded_ws, None, quantity_mode)
    for parsed_row in parsed_rows:
        parsed_row["_recalculate_weight"] = True
        parsed_row["_site_split"] = True
    return _polish_format2_update(populate_template_workbook_fill_only(parsed_rows, None))
