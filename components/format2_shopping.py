import re
from pathlib import Path
from collections import defaultdict
from copy import copy
import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from .common import (
    clean_ingredient_name,
    parse_date_value,
    format_date,
    to_numeric,
    format_number_clean,
    strip_accents,
    merged_value_lookup,
    effective_cell_value,
    get_vietnamese_weekday,
    save_workbook,
    normalize_comp,
    normalize_quantity_mode,
    QUANTITY_MODE_FORECAST
)
from .rules import get_shopping_short_name, clean_site_name, get_rule_for_customer

def normalize_template_shift_or_category(val):
    if not val:
        return ""
    s = str(val).lower().strip()
    s = strip_accents(s).replace(" ", "").replace(",", "").replace(".", "").replace("-", "")

    # 1. Categories first (to prevent substrings like 'c1' in 'nuocc1' matching shifts)
    if "chay" in s:
        return "chay"
    if "trangmieng" in s or "tmieng" in s:
        return "trangmieng"
    if "chao" in s:
        return "chao"
    if "nuoc" in s:
        return "nuoc"
    if "caithien" in s:
        return "cai_thien"

    # 2. Shifts second
    if "ca1" in s or "c1" in s:
        return "ca1"
    if "ca2" in s or "c2" in s:
        return "ca2"
    if "ca3" in s or "c3" in s:
        return "ca3"
    if "ansang" in s:
        return "ansang"

    return s

def get_row_company(ws, lookup, r, panel_start_col):
    val = effective_cell_value(ws, lookup, r, panel_start_col)
    if val:
        return val

    for row in range(r - 1, 2, -1):
        v = effective_cell_value(ws, lookup, row, panel_start_col)
        if v:
            return v

    # Scan downwards if we are at the top and found nothing
    for row in range(r + 1, ws.max_row + 1):
        v = effective_cell_value(ws, lookup, row, panel_start_col)
        if v:
            return v

    return None

def get_row_active_main_shift(ws, lookup, r, shift_col):
    for row in range(r, 2, -1):
        v = effective_cell_value(ws, lookup, row, shift_col)
        if v:
            v_clean = strip_accents(str(v).lower().strip()).replace(" ", "")
            if v_clean in ["ca1", "ca2", "ca3"]:
                return v_clean
    return "ca1"

def get_row_shift(ws, lookup, r, shift_col, sheet_name=None, panel_idx=None):
    val = effective_cell_value(ws, lookup, r, shift_col)
    if val:
        return val

    for row in range(r - 1, 2, -1):
        v = effective_cell_value(ws, lookup, row, shift_col)
        if v:
            return v

    return None

def infer_template_category_from_dish_hint(value):
    text = strip_accents(str(value or "").lower())
    if not text:
        return ""
    compact = text.replace(" ", "").replace(",", "").replace(".", "").replace("-", "")

    if "monchao" in compact or "suatchao" in compact:
        return "chao"
    if "monnuoc" in compact or "suatnuoc" in compact:
        return "nuoc"
    if "monchay" in compact or "suatchay" in compact:
        return "chay"
    if "trangmieng" in compact or "tmieng" in compact:
        return "trangmieng"
    if "caithien" in compact:
        return "cai_thien"
    return ""

def find_format2_fallback_data_rows(grouped_data, sheet_name, comp_norm, key_norm):
    data_rows = grouped_data.get((sheet_name, comp_norm, key_norm), [])
    if data_rows:
        return data_rows

    if key_norm.startswith("chay_"):
        for fallback_key in ("chay_ca1", "chay_ca2", "chay_ca3", "chay"):
            if fallback_key == key_norm:
                continue
            data_rows = grouped_data.get((sheet_name, comp_norm, fallback_key), [])
            if data_rows:
                return data_rows

    return []

def format2_cell_has_border(cell):
    border = cell.border
    return any(
        side is not None and side.style
        for side in (border.left, border.right, border.top, border.bottom)
    )

def row_has_format2_panel_frame(ws, row, cols):
    for col in cols:
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        if format2_cell_has_border(cell):
            return True
    return False

def compact_format2_dishes_for_slots(dishes, slot_count):
    if slot_count <= 0:
        return []
    if len(dishes) <= slot_count:
        return dishes

    def join_values(items, key):
        values = [
            str(item.get(key) or "").strip()
            for item in items
            if str(item.get(key) or "").strip()
        ]
        return "\n".join(values)

    compacted = [dict(dish) for dish in dishes[:slot_count - 1]]
    overflow = dishes[slot_count - 1:]
    compacted.append({
        "qty": join_values(overflow, "qty"),
        "mon": join_values(overflow, "mon"),
        "recipe": join_values(overflow, "recipe"),
    })
    return compacted

def format_single_ingredient_quota(amount, unit, short_name):
    amt_str = format_number_clean(amount)
    u = str(unit).strip().lower() if unit else ""
    if not u:
        return f"{amt_str} {short_name}"

    if u in ["gr", "g", "kg", "ml", "l", "c"]:
        return f"{amt_str}{u} {short_name}"
    else:
        return f"{amt_str} {u} {short_name}"

def format_single_ingredient_total(value, unit, short_name):
    val_str = format_number_clean(value)
    u = str(unit).strip().lower() if unit else ""
    if not u:
        return f"{val_str} {short_name}"

    if u in ["quả", "qua", "cái", "cai", "con", "miếng", "mieng", "viên", "vien", "cây", "cay", "khay"]:
        return f"{val_str} {short_name}"
    elif u in ["gr", "g", "kg", "ml", "l", "c"]:
        return f"{val_str}{u} {short_name}"
    else:
        return f"{val_str} {u} {short_name}"

def normalize_unit(unit):
    text = strip_accents(str(unit or "").strip().lower())
    return text.replace(" ", "")

def convert_amount_between_units(value, from_unit, to_unit):
    amount = to_numeric(value)
    if amount in [None, ""]:
        return None

    source = normalize_unit(from_unit)
    target = normalize_unit(to_unit)
    if not source or not target:
        return None
    if source == target:
        return float(amount)

    weight_to_gram = {
        "kg": 1000.0,
        "g": 1.0,
        "gr": 1.0,
        "gram": 1.0,
    }
    volume_to_ml = {
        "l": 1000.0,
        "lit": 1000.0,
        "ml": 1.0,
    }
    count_units = {
        "cai", "qua", "con", "mieng", "vien", "cay", "khay",
        "goi", "hop", "chai", "bo", "phan",
    }

    if source in weight_to_gram and target in weight_to_gram:
        return float(amount) * weight_to_gram[source] / weight_to_gram[target]
    if source in volume_to_ml and target in volume_to_ml:
        return float(amount) * volume_to_ml[source] / volume_to_ml[target]
    if source in count_units and target in count_units:
        return float(amount)
    return None

def infer_qty_from_weight(kl_mua, dvt_mua, dinh_muc, dvt):
    total_in_quota_unit = convert_amount_between_units(kl_mua, dvt_mua, dvt)
    quota = to_numeric(dinh_muc)
    if total_in_quota_unit in [None, ""] or quota in [None, "", 0]:
        return None
    try:
        inferred = float(total_in_quota_unit) / float(quota)
    except (TypeError, ValueError, ZeroDivisionError):
        return None
    if inferred <= 0:
        return None
    return inferred

def get_site_sort_key(site_name, cust=None):
    s = str(site_name).lower()
    comp_key = normalize_comp(cust) if cust else ""

    if comp_key == "lixil":
        if "fab1" in s:
            return (0, s)
        if "fab2" in s:
            return (1, s)
        if "vpc" in s:
            return (2, s)
        if "til" in s:
            return (3, s)
    else:
        if "vpc" in s:
            return (0, s)
        if "fab1" in s:
            return (1, s)
        if "fab2" in s:
            return (2, s)
        if "til" in s:
            return (3, s)
    return (4, s)

def map_columns(headers, quantity_mode=None):
    quantity_mode = normalize_quantity_mode(quantity_mode)
    col_map = {}
    for idx, col in enumerate(headers, start=1):
        if col is None:
            continue
        c_clean = str(col).lower().strip()
        c_no_accent = strip_accents(c_clean)

        if c_clean in ["ngày", "ngay", "date"]:
            col_map["ngay"] = idx
        elif c_clean in ["ca", "shift"]:
            col_map["ca"] = idx
        elif c_clean in ["khách hàng", "khach hang", "mã kh", "ma kh", "khach_hang", "mã khách hàng"] and "site" not in c_clean:
            col_map["ma_kh"] = idx
        elif c_clean in ["bếp", "bep", "kitchen"]:
            col_map["bep"] = idx
        elif c_clean in ["site ăn", "site an", "site"]:
            col_map["site_an"] = idx
        elif c_clean in ["cơ cấu suất ăn", "co cau suat an", "suất ăn", "suat an", "cơ cấu menu", "co cau menu"]:
            col_map["co_cau"] = idx
        elif c_clean in ["loại món ăn", "loai mon an", "nhóm món", "nhom mon", "cơ cấu món ăn", "co cau mon an"]:
            col_map["loai_mon"] = idx
        elif c_clean in ["nguyên vật liệu", "nguyen vat lieu", "nvl", "tên nguyên vật liệu", "ten nguyen vat lieu"] and "đầy đủ" not in c_clean and "day du" not in c_clean:
            col_map["nvl"] = idx
        elif c_clean in ["món ăn", "mon an", "tên món", "ten mon"] and "cơ cấu" not in c_clean and "loại" not in c_clean and "loai" not in c_clean:
            col_map["mon"] = idx
        elif (
            ("so luong" in c_no_accent or "s.luong" in c_no_accent)
            and ("co nga duyet" in c_no_accent or "nga duyet" in c_no_accent)
        ):
            col_map["qty_approved"] = idx
        elif ("so luong" in c_no_accent or "s.luong" in c_no_accent) and "du bao" in c_no_accent:
            col_map["qty_forecast"] = idx
        elif (
            c_clean in ["số lượng", "so luong", "s.lượng", "qty", "quantity"]
            or "so luong" in c_no_accent
            or "s.luong" in c_no_accent
        ) and "du bao" not in c_no_accent:
            col_map.setdefault("qty", idx)
        elif c_clean in ["định mức", "dinh muc", "định lượng", "dinh luong"] and "cam kết" not in c_clean and "cam ket" not in c_clean:
            col_map["dinh_muc"] = idx
        elif c_clean in ["đvt", "đơn vị tính", "don vi tinh", "đơn vị", "don vi", "dvt"] and "mua hàng" not in c_clean and "mua hang" not in c_clean:
            col_map["dvt"] = idx
        elif any(x in c_clean for x in ["kl yêu cầu", "khoi luong yeu cau", "kl yeu cau", "quy đổi", "quy doi", "kl thực tế", "kl thuc te"]):
            col_map["kl_mua"] = idx
        elif any(x in c_clean for x in ["đơn vị tính mua hàng", "don vi tinh mua hang", "dvt mua", "dvt_mua"]):
            col_map["dvt_mua"] = idx

    if quantity_mode == QUANTITY_MODE_FORECAST:
        quantity_candidates = ("qty_forecast", "qty", "qty_approved")
    else:
        quantity_candidates = ("qty_approved", "qty", "qty_forecast")

    for key in quantity_candidates:
        if key in col_map:
            col_map["qty"] = col_map[key]
            break

    return col_map

def parse_source_sheet(ws_source, file_path, quantity_mode=None):
    headers = [ws_source.cell(1, c).value for c in range(1, ws_source.max_column + 1)]
    col_map = map_columns(headers, quantity_mode)
    normalized_mode = normalize_quantity_mode(quantity_mode)
    selected_qty_key = "qty_forecast" if normalized_mode == QUANTITY_MODE_FORECAST else "qty_approved"

    # Read rows
    raw_rows = []
    for r in range(2, ws_source.max_row + 1):
        row_vals = [ws_source.cell(r, c).value for c in range(1, ws_source.max_column + 1)]
        if any(x is not None for x in row_vals):
            raw_rows.append(row_vals)

    # Một số file có sẵn cả hai cột số lượng nhưng cột được chọn hoàn toàn
    # trống. Trước đây trường hợp này vẫn khóa vào cột trống, khiến Format 2
    # xuất ra mẫu không có khối lượng. Chỉ ưu tiên cột đã chọn khi nó thực sự
    # có dữ liệu; nếu không thì dùng cột số lượng còn lại.
    if normalized_mode == QUANTITY_MODE_FORECAST:
        quantity_candidates = ("qty_forecast", "qty", "qty_approved")
    else:
        quantity_candidates = ("qty_approved", "qty", "qty_forecast")

    selected_qty_key_in_use = None
    for candidate in quantity_candidates:
        col_idx = col_map.get(candidate)
        if col_idx is None:
            continue
        if any(
            col_idx <= len(row) and row[col_idx - 1] is not None
            and str(row[col_idx - 1]).strip() != ""
            for row in raw_rows
        ):
            col_map["qty"] = col_idx
            selected_qty_key_in_use = candidate
            break

    selected_qty_column_present = selected_qty_key_in_use == selected_qty_key

    # Forward fill
    prev_vals = {}
    fill_keys = ["ngay", "ma_kh", "bep", "ca", "co_cau", "loai_mon", "mon", "site_an"]
    qty_group_keys = ["ngay", "ma_kh", "bep", "ca", "co_cau", "loai_mon", "mon", "site_an"]
    prev_qty_group = None
    prev_qty_value = None
    parsed_rows = []
    for row in raw_rows:
        row_vals = list(row)
        for key in fill_keys:
            idx = col_map.get(key)
            if idx is not None and (idx - 1) < len(row_vals):
                val = row_vals[idx - 1]
                if val is None or str(val).strip() == "":
                    val = prev_vals.get(key, "")
                else:
                    val = str(val).strip()
                row_vals[idx - 1] = val
                prev_vals[key] = val

        qty_idx = col_map.get("qty")
        if qty_idx is not None and (qty_idx - 1) < len(row_vals):
            group_key = tuple(
                str(row_vals[col_map[key] - 1]).strip()
                for key in qty_group_keys
                if col_map.get(key) is not None and (col_map[key] - 1) < len(row_vals)
            )
            qty_val = row_vals[qty_idx - 1]
            if qty_val is None or str(qty_val).strip() == "":
                if prev_qty_group == group_key:
                    row_vals[qty_idx - 1] = prev_qty_value
            else:
                qty_val = str(qty_val).strip()
                row_vals[qty_idx - 1] = qty_val
                prev_qty_group = group_key
                prev_qty_value = qty_val

        # Convert to dict
        row_dict = {}
        for key, col_idx in col_map.items():
            if col_idx is not None and (col_idx - 1) < len(row_vals):
                row_dict[key] = row_vals[col_idx - 1]
            else:
                row_dict[key] = ""
        row_dict["_selected_qty_blank"] = (
            selected_qty_column_present
            and (row_dict.get("qty") is None or str(row_dict.get("qty")).strip() == "")
        )
        parsed_rows.append(row_dict)

    return parsed_rows

def format_group_dishes(cust, r_list):
    # Group r_list by dish name (mon)
    dish_rows = defaultdict(list)
    for r in r_list:
        mon = str(r.get("mon") or "").strip()
        dish_rows[mon].append(r)

    cust_norm = normalize_comp(cust)
    is_osv = (cust_norm in ["osv", "odsv"])

    # Check if this customer has multiple sites in the dataset
    sites = set(clean_site_name(r.get("site_an", ""), cust) for r in r_list if r.get("site_an"))
    sites.discard("")
    is_multi_site = len(sites) > 1

    # Sort sites
    sorted_sites = sorted(list(sites), key=lambda x: get_site_sort_key(x, cust))

    dishes_data = []
    for mon, rows_mon in dish_rows.items():
        # Get ingredients and totals
        site_qty_map = defaultdict(float)
        site_nvl_totals = defaultdict(lambda: defaultdict(float))
        nvl_dinh_muc = {}
        nvl_has_weight = defaultdict(bool)

        for r in rows_mon:
            site = clean_site_name(r.get("site_an", ""), cust)
            qty_raw = r.get("qty")
            qty_blank = qty_raw is None or str(qty_raw).strip() == ""
            qty = to_numeric(qty_raw) or 0.0
            selected_qty_blank = bool(r.get("_selected_qty_blank"))

            nvl = str(r.get("nvl") or "").strip()
            if nvl:
                dm_val = to_numeric(r.get("dinh_muc")) or 0.0
                dm_unit = str(r.get("dvt") or "").strip()
                dvt_mua = str(r.get("dvt_mua") or "").strip()
                kl_raw = to_numeric(r.get("kl_mua"))
                has_kl_mua = kl_raw not in [None, ""]
                kl_mua = kl_raw if has_kl_mua else None

                if not has_kl_mua and r.get("_recalculate_weight") and qty and dm_val:
                    target_unit = dvt_mua or dm_unit
                    kl_mua = convert_amount_between_units(qty * dm_val, dm_unit, target_unit)
                    has_kl_mua = kl_mua not in [None, ""]

                if qty_blank and not selected_qty_blank:
                    qty = infer_qty_from_weight(kl_mua, dvt_mua, dm_val, dm_unit) or 0.0

                nvl_dinh_muc[nvl] = (dm_val, dm_unit, dvt_mua or dm_unit)
                if has_kl_mua:
                    site_nvl_totals[site][nvl] += kl_mua
                    nvl_has_weight[nvl] = True

            if qty or r.get("_site_split"):
                site_qty_map[site] = max(site_qty_map[site], qty)

        # 1. Format Quantity
        if is_multi_site:
            qty_val = "\n".join(
                f"{s} - {format_number_clean(site_qty_map[s])}"
                for s in sorted_sites
                if s in site_qty_map
            )
        else:
            total_qty = sum(site_qty_map.values())
            qty_val = format_number_clean(total_qty) if total_qty > 0 else ""

        # 2. Format Quota
        quota_parts = []
        for nvl, (dm_val, dm_unit, _) in nvl_dinh_muc.items():
            short_nvl = get_shopping_short_name(nvl, cust)
            quota_parts.append(format_single_ingredient_quota(dm_val, dm_unit, short_nvl))

        display_mon = mon
        ca_display = None
        if is_osv:
            co_cau_label = str(rows_mon[0].get("co_cau") or "").strip()
            co_cau_clean = strip_accents(co_cau_label.lower())
            if "ansang" in co_cau_clean or "sang" in co_cau_clean:
                ca_display = co_cau_label

        if display_mon:
            quota_str = f"{display_mon.lower()} " + " + ".join(quota_parts)
        else:
            quota_str = " + ".join(quota_parts)

        # 3. Format Total Shopping Weight
        if is_multi_site:
            site_shopping_lines = []
            for s in sorted_sites:
                has_weight = any(nvl in site_nvl_totals[s] for nvl in nvl_dinh_muc)
                if site_qty_map[s] == 0 and not has_weight:
                    continue
                weight_parts = []
                for nvl in nvl_dinh_muc:
                    if nvl not in site_nvl_totals[s]:
                        continue
                    kl = site_nvl_totals[s].get(nvl, 0.0)
                    _, _, dvt_mua = nvl_dinh_muc[nvl]
                    short_nvl = get_shopping_short_name(nvl, cust)
                    weight_parts.append(format_single_ingredient_total(kl, dvt_mua, short_nvl))
                weight_str = " + ".join(weight_parts)
                if weight_str:
                    site_shopping_lines.append(f"{s} - {weight_str}")
            shopping_val = "\n".join(site_shopping_lines)
        else:
            weight_parts = []
            total_sites = set(site_nvl_totals.keys()) or set(site_qty_map.keys())
            for nvl in nvl_dinh_muc:
                if not nvl_has_weight[nvl]:
                    continue
                total_kl = sum(site_nvl_totals[s].get(nvl, 0.0) for s in total_sites)
                _, _, dvt_mua = nvl_dinh_muc[nvl]
                short_nvl = get_shopping_short_name(nvl, cust)
                weight_parts.append(format_single_ingredient_total(total_kl, dvt_mua, short_nvl))
            weight_str = " + ".join(weight_parts)
            shopping_val = weight_str

        dishes_data.append({
            "qty": qty_val,
            "mon": quota_str,
            "recipe": shopping_val,
            "ca_display": ca_display
        })

    return dishes_data

def shift_merged_cells_cols(ws, insert_idx, amount=1):
    new_ranges = []
    for r in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = r.min_col, r.min_row, r.max_col, r.max_row
        if min_col >= insert_idx:
            min_col += amount
        if max_col >= insert_idx:
            max_col += amount
        new_range = CellRange(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        new_ranges.append(new_range)
    ws.merged_cells.ranges = new_ranges

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def write_value(ws, row, col, value):
    cell = ws.cell(row, col)
    if type(cell).__name__ != "MergedCell":
        cell.value = value
        return cell

    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            anchor = ws.cell(rng.min_row, rng.min_col)
            anchor.value = value
            return anchor
    return cell

def transform_central_kitchen_sheet(ws):
    # Insert columns to make them match Bếp tại chỗ 10-column layout
    ws.insert_cols(5, 1)
    shift_merged_cells_cols(ws, 5, 1)
    ws.insert_cols(10, 1)
    shift_merged_cells_cols(ws, 10, 1)

    # Configure widths and visibility
    ws.column_dimensions['C'].hidden = False
    ws.column_dimensions['G'].hidden = False
    ws.column_dimensions['H'].hidden = False
    ws.column_dimensions['C'].width = 6.43
    ws.column_dimensions['G'].width = 7.14
    ws.column_dimensions['H'].width = 6.43
    ws.column_dimensions['E'].width = 37.43
    ws.column_dimensions['J'].width = 37.43

    # Set headers
    ws.cell(2, 3).value = "SL"
    ws.cell(2, 5).value = "Khối lượng đi chợ"
    ws.cell(2, 8).value = "SL"
    ws.cell(2, 10).value = "Khối lượng đi chợ"

    # Copy header styles
    for r in [1, 2]:
        copy_cell_style(ws.cell(r, 2), ws.cell(r, 3))
        copy_cell_style(ws.cell(r, 4), ws.cell(r, 5))
        copy_cell_style(ws.cell(r, 7), ws.cell(r, 8))
        copy_cell_style(ws.cell(r, 9), ws.cell(r, 10))

    try:
        ws.merge_cells("I1:J1")
    except Exception:
        pass

def is_vegetarian_row(r):
    co_cau = str(r.get("co_cau") or "").lower()
    loai_mon = str(r.get("loai_mon") or "").lower()
    return "chay" in strip_accents(co_cau) or "chay" in strip_accents(loai_mon)

def get_chay_group_key(cty_val):
    s = normalize_comp(cty_val)
    if s in ["212", "213"] or any(x in s for x in ["scavi", "sofa", "artwell", "shiseido", "shisedo"]):
        return "scavi_sofa_chay"
    if s in ["204", "205"] or "briskheat" in s:
        return "briskheat_chay"
    if s in ["194", "195"] or "figla" in s:
        return "figla_chay"
    if s in ["227", "228"] or "saitex" in s:
        return "saitex_figla_ca2_chay"
    return s

def get_chay_template_group_key(cty_val, shift_val, r):
    c = normalize_comp(cty_val)
    if 21 <= r <= 24:
        return "scavi_sofa_chay"
    if 25 <= r <= 27:
        return "briskheat_chay"
    if 36 <= r <= 38:
        return "figla_chay"
    if 41 <= r <= 45:
        return "saitex_figla_ca2_chay"
    if 48 <= r <= 49:
        return "ca3_chay"
    return c

def get_chay_record_group_key(cust_norm, ca_norm):
    if any(x in cust_norm for x in ["scavi", "sofa", "artwell", "shiseido", "shisedo", "7a", "osv", "odsv"]):
        return "scavi_sofa_chay"
    if "briskheat" in cust_norm:
        return "briskheat_chay"
    if "figla" in cust_norm:
        if "ca2" in ca_norm or "c2" in ca_norm:
            return "saitex_figla_ca2_chay"
        return "figla_chay"
    if "saitex" in cust_norm:
        return "saitex_figla_ca2_chay"
    return cust_norm

def get_comp_chay_qty_display(cust_norm, veggie_recs):
    recs = []
    for r in veggie_recs:
        c = normalize_comp(r.get("ma_kh"))
        if c == cust_norm:
            recs.append(r)
        elif cust_norm == "saitexd13" and "d13" in c and "saitex" in c:
            recs.append(r)
        elif cust_norm == "dona" and "dona" in c:
            recs.append(r)

    seen = set()
    unique_recs = []
    for r in recs:
        h = (r.get("ca"), r.get("mon"), r.get("qty"))
        if h not in seen:
            seen.add(h)
            unique_recs.append(r)

    if not unique_recs:
        return None

    shift_qtys = {}
    for r in unique_recs:
        ca = normalize_template_shift_or_category(r.get("ca")) or "ca1"
        qty = to_numeric(r.get("qty")) or 0
        if ca not in shift_qtys or qty > shift_qtys[ca]:
            shift_qtys[ca] = qty

    if not shift_qtys:
        return None
    if len(shift_qtys) == 1:
        return str(list(shift_qtys.values())[0])
    else:
        parts = []
        for ca in sorted(shift_qtys.keys()):
            parts.append(f"{ca.upper().replace('CA', 'Ca ')} - {shift_qtys[ca]}")
        return "\n".join(parts)

def populate_bt2_veggie_table(ws, grouped_data, all_parsed_rows):
    veggie_recs = [r for r in all_parsed_rows if is_vegetarian_row(r) and str(r.get("bep") or "").strip() in ["Bếp trung tâm", "Bếp trung tâm 2"]]

    # 1. Main veggie groups population
    main_groups = {
        "scavi_sofa_chay": [23, 24],
        "briskheat_chay": [26, 27],
        "figla_chay": [36, 37, 38, 39],
        "saitex_figla_ca2_chay": [44, 45],
        "ca3_chay": [48, 49]
    }

    def safe_write_cell(row, col, val, align=None):
        c = ws.cell(row, col)
        if type(c).__name__ != "MergedCell":
            c.value = val
            if align:
                c.alignment = align

    def bt2_dish_col(row):
        # Most BT2 veggie rows use column I for "Món ăn"; a few sample rows
        # intentionally merge G:I, so their writable cell is the merge origin.
        if type(ws.cell(row, 9)).__name__ != "MergedCell":
            return 9
        return 7

    for group_name, row_list in main_groups.items():
        recs = grouped_data.get(("Bếp trung tâm 2", group_name, ""))
        dishes = format_group_dishes(recs[0].get("ma_kh") if recs else None, recs) if recs else []

        for idx, r_idx in enumerate(row_list):
            dish_col = bt2_dish_col(r_idx)
            if idx < len(dishes):
                dish = dishes[idx]
                safe_write_cell(r_idx, dish_col, dish["mon"], Alignment(horizontal="left", vertical="center", wrap_text=True))
                safe_write_cell(r_idx, 8, dish["qty"], Alignment(horizontal="center", vertical="center", wrap_text=True))
                safe_write_cell(r_idx, 10, dish["recipe"], Alignment(horizontal="left", vertical="center", wrap_text=True))
            else:
                safe_write_cell(r_idx, 7, None)
                safe_write_cell(r_idx, 9, None)
                safe_write_cell(r_idx, 8, None)
                safe_write_cell(r_idx, 10, None)

    # 2. Company portion list population
    company_rows = [
        # Left side: (row, col, cust_norm, default_display)
        (29, 6, "watabe", "Watabe"),
        (30, 6, "catthai", "Cát Thái"),
        (31, 6, "maspro", "Maspro"),
        (32, 6, "medic", "Medic"),
        (33, 6, "vikyno", "Vikyno"),
        (34, 6, "dona", "Dona QB"),
        # Right side: (row, col, cust_norm, default_display)
        (29, 9, "saitexd13", "Saitex D13"),
        (30, 9, "saitex4", "Saitex 4"),
        (31, 9, "saitex6", "Saitex 6"),
        (32, 9, "figla", "Figla"),
        (33, 9, "zamil", "Zamil"),
        (34, 9, "osv", "OSV")
    ]

    for row_idx, col_idx, cust_norm, default_display in company_rows:
        qty_val = get_comp_chay_qty_display(cust_norm, veggie_recs)
        qty_col = 8 if col_idx == 6 else 10

        if qty_val:
            safe_write_cell(row_idx, col_idx, default_display, Alignment(horizontal="center", vertical="center", wrap_text=True))
            safe_write_cell(row_idx, qty_col, qty_val, Alignment(horizontal="center", vertical="center", wrap_text=True))
        else:
            safe_write_cell(row_idx, col_idx, default_display, Alignment(horizontal="center", vertical="center", wrap_text=True))
            safe_write_cell(row_idx, qty_col, None)

def clear_bt2_veggie_template(ws):
    start_row = None
    for row in range(3, ws.max_row + 1):
        row_text = "".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1))
        if "monchay" in strip_accents(row_text.lower()).replace(" ", ""):
            start_row = row
            break

    if not start_row:
        return

    for rng in list(ws.merged_cells.ranges):
        if rng.max_row >= start_row:
            ws.unmerge_cells(
                start_row=rng.min_row,
                start_column=rng.min_col,
                end_row=rng.max_row,
                end_column=rng.max_col,
            )
    ws.delete_rows(start_row, ws.max_row - start_row + 1)

def shift_sort_key(shift_value):
    normalized = normalize_template_shift_or_category(shift_value)
    if normalized == "ca1":
        return (1, "Ca 1")
    if normalized == "ca2":
        return (2, "Ca 2")
    if normalized == "ca3":
        return (3, "Ca 3")
    if normalized == "ansang":
        return (4, "Ăn sáng")
    return (99, str(shift_value or ""))

def append_central_veggie_table(ws, all_parsed_rows):
    veggie_rows = [
        row
        for row in all_parsed_rows
        if is_vegetarian_row(row)
        and str(row.get("bep") or "").strip() in ["Bếp trung tâm", "Bếp trung tâm 2"]
    ]
    if not veggie_rows:
        return

    grouped = defaultdict(list)
    order = []
    for row in veggie_rows:
        shift_label = str(row.get("ca") or "").strip() or "Ca 1"
        customer = str(row.get("ma_kh") or "").strip()
        key = (shift_label, customer)
        if key not in grouped:
            order.append(key)
        grouped[key].append(row)

    order.sort(key=lambda item: (shift_sort_key(item[0]), normalize_comp(item[1]), item[1]))

    start = ws.max_row + 2
    max_col = 9
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = copy(ws.cell(2, 1).fill)

    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=max_col)
    title_cell = ws.cell(start, 1)
    title_cell.value = "MÓN CHAY BẾP TRUNG TÂM"
    title_cell.font = Font(name="Calibri", size=18, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.border = thin_border
    ws.row_dimensions[start].height = 24

    header_row = start + 1
    headers = ["Ca", "Cty", "SL", "Món ăn", "Khối lượng", None, None, None, None]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(header_row, idx)
        cell.value = value
        cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF" if value else "000000")
        if value:
            cell.fill = copy(header_fill)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 22

    row_idx = header_row + 1
    for shift_label, customer in order:
        dishes = format_group_dishes(customer, grouped[(shift_label, customer)])
        if not dishes:
            continue
        first_row = row_idx
        for dish in dishes:
            values = [shift_label, customer, dish["qty"], dish["mon"], dish["recipe"]]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx)
                cell.value = value
                cell.border = thin_border
                if col_idx in [1, 2, 3]:
                    cell.font = Font(name="Calibri", size=11, bold=col_idx in [1, 2, 3])
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif col_idx == 4:
                    cell.font = Font(name="Calibri", size=13)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.font = Font(name="Calibri", size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            for col_idx in range(6, max_col + 1):
                ws.cell(row_idx, col_idx).border = thin_border

            max_lines = max(
                str(dish["qty"]).count("\n") + 1,
                str(dish["mon"]).count("\n") + 1,
                str(dish["recipe"]).count("\n") + 1,
            )
            ws.row_dimensions[row_idx].height = max(34.2, max_lines * 15 + 5)
            row_idx += 1

        if row_idx - first_row > 1:
            ws.merge_cells(start_row=first_row, start_column=1, end_row=row_idx - 1, end_column=1)
            ws.merge_cells(start_row=first_row, start_column=2, end_row=row_idx - 1, end_column=2)
            ws.cell(first_row, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(first_row, 2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def delete_row_and_fix_merges(ws, row_idx):
    old_ranges = list(ws.merged_cells.ranges)
    ws.merged_cells.ranges = []
    ws.delete_rows(row_idx, 1)
    for rng in old_ranges:
        min_col, min_row, max_col, max_row = rng.min_col, rng.min_row, rng.max_col, rng.max_row
        if min_row <= row_idx <= max_row:
            if min_row == max_row:
                continue
            else:
                max_row -= 1
        elif min_row > row_idx:
            min_row -= 1
            max_row -= 1

        if (max_row > min_row) or (max_col > min_col):
            new_rng = CellRange(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            ws.merged_cells.add(new_rng)

def delete_unused_rows(ws, sheet_name):
    # 1. Unmerge and fill the Cty and Ca columns for both panels starting from row 3
    def unmerge_and_fill_column(ws, col_idx):
        ranges_to_unmerge = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= 3 and rng.min_col == col_idx and rng.max_col == col_idx:
                ranges_to_unmerge.append(rng)

        for rng in ranges_to_unmerge:
            val = ws.cell(rng.min_row, rng.min_col).value
            ws.unmerge_cells(start_row=rng.min_row, start_column=rng.min_col, end_row=rng.max_row, end_column=rng.max_col)
            for r in range(rng.min_row, rng.max_row + 1):
                key = (r, col_idx)
                if key in ws._cells:
                    if type(ws._cells[key]).__name__ == "MergedCell":
                        del ws._cells[key]
                ws.cell(r, col_idx).value = val

    unmerge_and_fill_column(ws, 1) # Panel 1 Cty
    unmerge_and_fill_column(ws, 2) # Panel 1 Ca
    unmerge_and_fill_column(ws, 6) # Panel 2 Cty
    unmerge_and_fill_column(ws, 7) # Panel 2 Ca

    is_header = [False] * (ws.max_row + 1)
    for r in range(1, ws.max_row + 1):
        row_txt = "".join(str(ws.cell(r, c).value or "").lower() for c in range(1, ws.max_column + 1))
        if r in [1, 2]:
            is_header[r] = True
            continue
        if "monchay" in strip_accents(row_txt) or "ca1:" in row_txt or "ca2:" in row_txt or "ca3:" in row_txt:
            is_header[r] = True

    row_active = [False] * (ws.max_row + 1)

    def has_values_in_cols(row_idx, data_cols):
        for col_idx in data_cols:
            val = ws.cell(row_idx, col_idx).value
            if val in [None, "", 0, "0"]:
                continue
            if isinstance(val, str) and val.startswith("="):
                continue
            return True
        return False

    def has_visible_output_data(row_idx):
        return has_values_in_cols(row_idx, [3, 4, 5, 8, 9, 10])

    def panel_has_visible_output(row_idx, panel_idx):
        data_cols = [3, 4, 5] if panel_idx == 1 else [8, 9, 10]
        return has_values_in_cols(row_idx, data_cols)

    # For Bếp trung tâm 2, find the MÓN CHAY marker row to preserve the veggie section
    veggie_start_row = None
    if sheet_name == "Bếp trung tâm 2":
        for r in range(3, ws.max_row + 1):
            row_txt = "".join(str(ws.cell(r, c).value or "").lower() for c in range(1, ws.max_column + 1))
            if "monchay" in strip_accents(row_txt).replace(" ", ""):
                veggie_start_row = r
                break

    for r in range(1, ws.max_row + 1):
        if is_header[r]:
            continue
        if has_visible_output_data(r):
            row_active[r] = True

    for r in range(3, ws.max_row + 1):
        if is_header[r]:
            has_active_below = False
            for next_r in range(r + 1, ws.max_row + 1):
                if is_header[next_r]:
                    break
                if row_active[next_r]:
                    has_active_below = True
                    break
            if has_active_below:
                row_active[r] = True

    rows_to_delete = []
    for r in range(3, ws.max_row + 1):
        if not row_active[r]:
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        delete_row_and_fix_merges(ws, r)

    # 4. Remerge the columns
    def remerge_column(ws, col_idx):
        current_is_header = [False] * (ws.max_row + 1)
        panel_cols = range(1, 6) if col_idx <= 2 else range(6, 11)
        lookup = merged_value_lookup(ws)
        for r in range(1, ws.max_row + 1):
            row_txt = "".join(str(ws.cell(r, c).value or "").lower() for c in panel_cols)
            if r in [1, 2]:
                current_is_header[r] = True
                continue
            if "monchay" in strip_accents(row_txt) or "ca1:" in row_txt or "ca2:" in row_txt or "ca3:" in row_txt:
                current_is_header[r] = True

        r = 3
        while r <= ws.max_row:
            if current_is_header[r]:
                r += 1
                continue
            panel_idx = 1 if col_idx <= 2 else 2
            if not panel_has_visible_output(r, panel_idx):
                r += 1
                continue
            val = ws.cell(r, col_idx).value
            if val is None or str(val).strip() == "":
                r += 1
                continue

            start_row = r
            end_row = r
            comp_col = 1 if col_idx <= 4 else 6
            start_comp = effective_cell_value(ws, lookup, start_row, comp_col)
            for next_r in range(r + 1, ws.max_row + 1):
                if current_is_header[next_r]:
                    break
                if not panel_has_visible_output(next_r, panel_idx):
                    break
                # Ensure we do not merge Ca cells across different companies
                next_comp = effective_cell_value(ws, lookup, next_r, comp_col)
                if next_comp != start_comp:
                    break
                if ws.cell(next_r, col_idx).value == val:
                    end_row = next_r
                else:
                    break

            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                c = ws.cell(start_row, col_idx)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                r = end_row + 1
            else:
                r += 1

    remerge_column(ws, 1) # Panel 1 Cty
    remerge_column(ws, 2) # Panel 1 Ca
    remerge_column(ws, 6) # Panel 2 Cty
    remerge_column(ws, 7) # Panel 2 Ca

    # Standard thin gray border for all data cells
    thin_side = Side(border_style="thin", color="A0A0A0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r in range(3, ws.max_row + 1):
        for c in range(1, 11):
            # Check if this cell's panel has header keywords in this row
            if c <= 5:
                row_txt = "".join(str(ws.cell(r, col).value or "").lower() for col in range(1, 6))
            else:
                row_txt = "".join(str(ws.cell(r, col).value or "").lower() for col in range(6, 11))

            is_cell_hdr = "monchay" in strip_accents(row_txt) or "ca1:" in row_txt or "ca2:" in row_txt or "ca3:" in row_txt

            if not is_cell_hdr:
                cell = ws.cell(r, c)

                # Column-specific typography
                bold = c in [1, 3, 6, 8]
                size = 10 if c in [5, 10] else 11

                orig_font = cell.font
                color = orig_font.color if orig_font else None
                is_bold = bold or (orig_font.bold if orig_font else False)
                is_italic = orig_font.italic if orig_font else False

                cell.font = Font(name="Aptos Narrow", size=size, bold=is_bold, italic=is_italic, color=color)
                cell.border = thin_border

                # Column-specific alignments
                if c in [2, 3, 7, 8]:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

def compact_bt2_right_panel(ws):
    has_staged_values = any(
        ws.cell(row, 10).value not in (None, "")
        for row in range(3, ws.max_row + 1)
    )
    if ws.cell(2, 7).value == "SL" and ws.cell(2, 8).value == "Món ăn" and not has_staged_values:
        return

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= 2 and rng.min_col >= 7 and rng.max_col <= 10:
            ws.unmerge_cells(
                start_row=rng.min_row,
                start_column=rng.min_col,
                end_row=rng.max_row,
                end_column=rng.max_col,
            )

    for row in range(2, ws.max_row + 1):
        g_val = ws.cell(row, 7).value
        h_val = ws.cell(row, 8).value
        i_val = ws.cell(row, 9).value
        j_val = ws.cell(row, 10).value

        if h_val is not None or i_val is not None or j_val is not None:
            if g_val is not None and h_val is None and i_val is None and j_val is not None:
                ws.cell(row, 7).value = None
                ws.cell(row, 8).value = g_val
                ws.cell(row, 9).value = j_val
            else:
                ws.cell(row, 7).value = h_val
                ws.cell(row, 8).value = i_val
                ws.cell(row, 9).value = j_val
            ws.cell(row, 10).value = None

        for col in range(7, 10):
            val = ws.cell(row, col).value
            if isinstance(val, str) and val.startswith("="):
                ws.cell(row, col).value = None

    for rng in list(ws.merged_cells.ranges):
        if str(rng) == "I1:J1":
            ws.unmerge_cells("I1:J1")
            break
    try:
        ws.merge_cells("H1:I1")
    except Exception:
        pass

    ws.cell(2, 7).value = "SL"
    ws.cell(2, 8).value = "Món ăn"
    ws.cell(2, 9).value = "Khối lượng"
    ws.cell(2, 10).value = None
    ws.column_dimensions["J"].hidden = True
    ws.column_dimensions["J"].width = 0

def clear_blank_panel_boxes(ws, is_bt2):
    blank_fill = PatternFill(fill_type=None)
    blank_border = Border()

    def has_content(row_idx, cols):
        for col_idx in cols:
            val = ws.cell(row_idx, col_idx).value
            if val in [None, "", 0, "0"]:
                continue
            if isinstance(val, str) and val.startswith("="):
                continue
            return True
        return False

    def clear_panel(row_idx, cols):
        for col_idx in cols:
            cell = ws.cell(row_idx, col_idx)
            if type(cell).__name__ == "MergedCell":
                continue
            cell.value = None
            cell.border = blank_border
            cell.fill = blank_fill

    right_panel_cols = range(6, 10) if is_bt2 else range(6, 11)
    right_content_cols = [7, 8, 9] if is_bt2 else [8, 9, 10]

    for row in range(3, ws.max_row + 1):
        left_has_content = has_content(row, [3, 4, 5])
        right_has_content = has_content(row, right_content_cols)

        if left_has_content and not right_has_content:
            clear_panel(row, right_panel_cols)
        elif right_has_content and not left_has_content:
            clear_panel(row, range(1, 6))

def apply_reference_menu_style(ws, sheet_name):
    is_bt2 = sheet_name == "Bếp trung tâm 2"
    if is_bt2:
        compact_bt2_right_panel(ws)

    title_map = {
        "Bếp tại chỗ": "BẾP TẠI CHỖ",
        "Bếp trung tâm": "BẾP TRUNG TÂM",
        "Bếp trung tâm 2": "BẾP TRUNG TÂM",
    }
    if sheet_name in title_map:
        title_col = 8 if is_bt2 else 9
        write_value(ws, 1, title_col, title_map[sheet_name])

    ws.page_margins.left = 0
    ws.page_margins.right = 0
    ws.page_margins.top = 0
    ws.page_margins.bottom = 0
    ws.row_dimensions[1].height = 21
    ws.row_dimensions[2].height = 23.4

    for col_letter, width in {
        "A": 8.33, "B": 7.11, "C": 8.4, "D": 55,
        "E": 36, "F": 8.33, "G": 7.11, "H": 8.4,
        "I": 55, "J": 36,
    }.items():
        ws.column_dimensions[col_letter].width = width
        ws.column_dimensions[col_letter].hidden = False
    if is_bt2:
        ws.column_dimensions["G"].width = 8.4
        ws.column_dimensions["H"].width = 55
        ws.column_dimensions["I"].width = 36
        ws.column_dimensions["J"].hidden = True
        ws.column_dimensions["J"].width = 0

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    max_visible_col = 9 if is_bt2 else 10
    for col in range(1, max_visible_col + 1):
        header = ws.cell(2, col)
        header.font = Font(name="Calibri", size=18, bold=False, color="FFFFFF")
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.border = thin_border
    if is_bt2:
        ws.cell(2, 7).value = "SL"
        ws.cell(2, 8).value = "Món ăn"
        ws.cell(2, 9).value = "Khối lượng"
        ws.cell(2, 10).value = None

    title_cols = [4, 5, 6, 8] if is_bt2 else [4, 5, 6, 9]
    for col in title_cols:
        cell = ws.cell(1, col)
        cell.font = Font(name="Calibri", size=16, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(3, ws.max_row + 1):
        current_height = ws.row_dimensions[row].height or 0
        if current_height < 34.2:
            ws.row_dimensions[row].height = 34.2

        row_txt = strip_accents("".join(str(ws.cell(row, col).value or "").lower() for col in range(1, max_visible_col + 1)))
        row_key = row_txt.replace(" ", "")
        is_chay_header = "monchay" in row_key or "ca1:" in row_key or "ca2:" in row_key or "ca3:" in row_key

        for col in range(1, max_visible_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ == "MergedCell":
                continue
            cell.border = thin_border

            if is_chay_header:
                cell.font = Font(name="Calibri", size=13, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "").strip().upper() == "MÓN CHAY":
                    cell.font = Font(name="Calibri", size=22, bold=True)
                continue

            if col in [1, 6]:
                cell.font = Font(name="Calibri", size=13, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col == 2 or (col == 7 and not is_bt2):
                cell.font = Font(name="Calibri", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col == 3 or (col == 8 and not is_bt2) or (col == 7 and is_bt2):
                cell.font = Font(name="Calibri", size=13, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col == 4 or (col == 9 and not is_bt2) or (col == 8 and is_bt2):
                cell.font = Font(name="Calibri", size=13, bold=False)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=11, bold=False)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    clear_blank_panel_boxes(ws, is_bt2)

def fix_central_kitchen_formulas(ws):
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if val and isinstance(val, str) and val.startswith("="):
                val = re.sub(r'\bG(\d+)\b', r'H\1', val)
                val = re.sub(r'\bF(\d+)\b', r'G\1', val)
                val = re.sub(r'\bE(\d+)\b', r'F\1', val)
                cell.value = val

def insert_template_rows(ws, insert_at, n_inserted, company_col=1):
    ws.insert_rows(insert_at, n_inserted)
    ref_row = insert_at - 1
    val_cols = [1, 2] if company_col == 1 else [6, 7]
    for r_idx in range(insert_at, insert_at + n_inserted):
        for col in range(1, ws.max_column + 1):
            ref_cell = ws.cell(ref_row, col)
            new_cell = ws.cell(r_idx, col)
            if ref_cell.font:
                new_cell.font = Font(
                    name=ref_cell.font.name,
                    size=ref_cell.font.size,
                    bold=ref_cell.font.bold,
                    italic=ref_cell.font.italic,
                    color=ref_cell.font.color
                )
            if ref_cell.fill:
                new_cell.fill = PatternFill(
                    fill_type=ref_cell.fill.fill_type,
                    start_color=ref_cell.fill.start_color,
                    end_color=ref_cell.fill.end_color
                )
            if ref_cell.border:
                new_cell.border = Border(
                    left=ref_cell.border.left,
                    right=ref_cell.border.right,
                    top=ref_cell.border.top,
                    bottom=ref_cell.border.bottom
                )
            if ref_cell.alignment:
                new_cell.alignment = Alignment(
                    horizontal=ref_cell.alignment.horizontal,
                    vertical=ref_cell.alignment.vertical,
                    wrap_text=ref_cell.alignment.wrap_text
                )
            if col in val_cols:
                new_cell.value = ref_cell.value

def shift_template_groups(template_groups, after_row, shift_amt):
    for key, r_list in list(template_groups.items()):
        new_list = []
        for r in r_list:
            if r >= after_row:
                new_list.append(r + shift_amt)
            else:
                new_list.append(r)
        template_groups[key] = new_list

def unmerge_and_fill_all_template_company_and_ca_cells(ws):
    for col_idx in [1, 2, 5, 6]:
        ranges_to_unmerge = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= 3 and rng.min_col == col_idx and rng.max_col == col_idx:
                ranges_to_unmerge.append(rng)

        for rng in ranges_to_unmerge:
            val = ws.cell(rng.min_row, rng.min_col).value
            ws.unmerge_cells(start_row=rng.min_row, start_column=rng.min_col, end_row=rng.max_row, end_column=rng.max_col)
            for r in range(rng.min_row, rng.max_row + 1):
                key = (r, col_idx)
                if key in ws._cells:
                    if type(ws._cells[key]).__name__ == "MergedCell":
                        del ws._cells[key]
                ws.cell(r, col_idx).value = val

def should_process_group_in_panel(sheet_name, c_key, k_key, panel_idx, template_groups):
    if sheet_name in ["Bếp trung tâm", "Bếp trung tâm 2"]:
        if k_key == "" and not (sheet_name == "Bếp trung tâm 2" and panel_idx == 2):
            return False

    has_template_rows = any(k[0] == c_key for k in template_groups.keys())
    return has_template_rows

def set_existing_cell_value(ws, row, col, value):
    if row < 1 or col < 1 or row > ws.max_row or col > ws.max_column:
        return False
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True

def find_bt2_veggie_start_row(ws):
    for row in range(3, ws.max_row + 1):
        row_text = "".join(str(ws.cell(row, col).value or "") for col in range(1, ws.max_column + 1))
        if "monchay" in strip_accents(row_text.lower()).replace(" ", ""):
            return row
    return None

def get_fill_only_panel_specs(sheet_name):
    if sheet_name == "Bếp trung tâm 2":
        return [
            {"panel_idx": 1, "company_col": 1, "shift_col": 2, "data_cols": (3, 4, 5)},
            {"panel_idx": 2, "company_col": 6, "shift_col": 7, "data_cols": (7, 8, 9)},
        ]
    return [
        {"panel_idx": 1, "company_col": 1, "shift_col": 2, "data_cols": (3, 4, 5)},
        {"panel_idx": 2, "company_col": 6, "shift_col": 7, "data_cols": (8, 9, 10)},
    ]

def group_format2_records_for_fill_only(all_parsed_rows):
    grouped_data = defaultdict(list)
    central_veggie_rows = []

    for r in all_parsed_rows:
        bep = str(r.get("bep") or "").strip()
        cust = str(r.get("ma_kh") or "").strip()
        ca = str(r.get("ca") or "").strip()
        if not bep or not cust:
            continue

        is_chay_rec = is_vegetarian_row(r)
        if is_chay_rec and bep in ["Bếp trung tâm", "Bếp trung tâm 2"]:
            central_veggie_rows.append(r)
            continue

        rule = get_rule_for_customer(cust)
        loai_mon_resolved = rule.get_loai_mon_val(r)
        co_cau_str = str(r.get("co_cau") or "").lower()
        lm_str = str(loai_mon_resolved or "").lower()
        cust_norm = normalize_comp(cust)

        is_com_khach = "khach" in strip_accents(co_cau_str) or "khach" in strip_accents(lm_str)
        if is_com_khach:
            price = None
            for p_val in ["50", "70", "40"]:
                if p_val in co_cau_str or p_val in lm_str:
                    price = p_val
                    break
            if cust_norm == "scavi" and price == "50":
                comp_key = "scavicomkhach50k"
            elif cust_norm == "scavi" and price == "70":
                comp_key = "scavicomkhach70k"
            elif cust_norm == "sofa" and price == "40":
                comp_key = "sofacomql40k"
            else:
                comp_key = cust_norm
            key_norm = ""
        else:
            comp_key = cust_norm
            lm_clean = strip_accents(str(loai_mon_resolved).lower().strip()).replace(" ", "")
            lm_clean_base = re.sub(r"\(.*?\)", "", lm_clean).strip()
            is_main = lm_clean_base in ["monman", "suatman", "suatcom", "com"] or \
                "suatcom" in lm_clean or "suatman" in lm_clean or "monman" in lm_clean

            if is_main:
                key_norm = normalize_template_shift_or_category(ca)
            else:
                cat = normalize_template_shift_or_category(loai_mon_resolved)
                if not cat:
                    cat = normalize_template_shift_or_category(ca)

                if cat in ["nuoc", "chao", "trangmieng", "cai_thien", "chay"]:
                    ca_clean = normalize_template_shift_or_category(ca)
                    if ca_clean not in ["ca1", "ca2", "ca3"]:
                        ca_clean = "ca1"
                    key_norm = f"{cat}_{ca_clean}"
                else:
                    key_norm = cat

        grouped_data[(bep, comp_key, key_norm)].append(r)

    return grouped_data, central_veggie_rows

def build_fill_only_template_groups(ws, sheet_name, spec):
    lookup = merged_value_lookup(ws)
    groups = defaultdict(list)
    end_row = ws.max_row
    if sheet_name == "Bếp trung tâm 2":
        veggie_start = find_bt2_veggie_start_row(ws)
        if veggie_start:
            end_row = veggie_start - 1

    frame_cols = [spec["company_col"], spec["shift_col"], *spec["data_cols"]]
    for row in range(3, end_row + 1):
        if not row_has_format2_panel_frame(ws, row, frame_cols):
            continue

        cty_val = get_row_company(ws, lookup, row, spec["company_col"])
        if not cty_val:
            continue

        shift_val = get_row_shift(ws, lookup, row, spec["shift_col"], sheet_name, spec["panel_idx"])
        comp_norm = normalize_comp(cty_val)
        cat = normalize_template_shift_or_category(shift_val) if shift_val else ""
        hint_cat = infer_template_category_from_dish_hint(
            effective_cell_value(ws, lookup, row, spec["data_cols"][1])
        )
        if hint_cat and cat in ["ca1", "ca2", "ca3"]:
            key_norm = f"{hint_cat}_{cat}"
        elif cat in ["nuoc", "chao", "trangmieng", "cai_thien", "chay"]:
            active_main = get_row_active_main_shift(ws, lookup, row, spec["shift_col"])
            key_norm = f"{cat}_{active_main}"
        else:
            key_norm = cat

        groups[(comp_norm, key_norm)].append(row)

    return groups

def build_fill_only_template_group_specs(ws, sheet_name):
    return [
        (spec, build_fill_only_template_groups(ws, sheet_name, spec))
        for spec in get_fill_only_panel_specs(sheet_name)
    ]

def clear_fill_only_data_slots(ws, sheet_name, panel_groups=None):
    if panel_groups is None:
        panel_groups = build_fill_only_template_group_specs(ws, sheet_name)

    for spec, groups in panel_groups:
        for row_list in groups.values():
            for row in row_list:
                for col in spec["data_cols"]:
                    set_existing_cell_value(ws, row, col, None)

    if sheet_name == "Bếp trung tâm 2":
        veggie_start = find_bt2_veggie_start_row(ws)
        if veggie_start:
            for col in (2, 5, 6, 9):
                set_existing_cell_value(ws, veggie_start, col, None)
            for row in range(veggie_start + 1, ws.max_row + 1):
                for col in range(2, 10):
                    set_existing_cell_value(ws, row, col, None)

def fill_template_header_dates_only(ws, all_parsed_rows):
    dates = []
    seen = set()
    for row in all_parsed_rows:
        value = str(row.get("ngay") or "").strip()
        if value and value not in seen:
            seen.add(value)
            dates.append(value)
    date_str = ", ".join(dates)
    if not date_str:
        return
    weekday_str = get_vietnamese_weekday(date_str).strip().rstrip("-").strip() + " -"
    set_existing_cell_value(ws, 1, 4, weekday_str)
    set_existing_cell_value(ws, 1, 6, date_str)

def fill_standard_template_slots_only(ws, sheet_name, grouped_data, panel_groups=None):
    if panel_groups is None:
        panel_groups = build_fill_only_template_group_specs(ws, sheet_name)

    for spec, groups in panel_groups:
        qty_col, dish_col, weight_col = spec["data_cols"]
        for (comp_norm, key_norm), row_list in groups.items():
            data_rows = find_format2_fallback_data_rows(grouped_data, sheet_name, comp_norm, key_norm)
            if not data_rows:
                continue
            dishes = format_group_dishes(data_rows[0].get("ma_kh"), data_rows)
            dishes = compact_format2_dishes_for_slots(dishes, len(row_list))
            for row, dish in zip(row_list, dishes):
                set_existing_cell_value(ws, row, qty_col, dish["qty"])
                set_existing_cell_value(ws, row, dish_col, dish["mon"])
                set_existing_cell_value(ws, row, weight_col, dish["recipe"])

def fill_central_veggie_template_slots_only(ws, central_veggie_rows):
    veggie_start = find_bt2_veggie_start_row(ws)
    if not veggie_start or not central_veggie_rows:
        return

    grouped = defaultdict(list)
    ordered_keys = []
    for row in central_veggie_rows:
        shift_label = str(row.get("ca") or "").strip() or "Ca 1"
        customer = str(row.get("ma_kh") or "").strip()
        key = (shift_label, customer)
        if key not in grouped:
            ordered_keys.append(key)
        grouped[key].append(row)

    ordered_keys.sort(key=lambda item: (shift_sort_key(item[0]), normalize_comp(item[1]), item[1]))

    row_idx = veggie_start + 1
    for shift_label, customer in ordered_keys:
        dishes = format_group_dishes(customer, grouped[(shift_label, customer)])
        for dish in dishes:
            if row_idx > ws.max_row:
                return
            set_existing_cell_value(ws, row_idx, 2, shift_label)
            set_existing_cell_value(ws, row_idx, 3, customer)
            set_existing_cell_value(ws, row_idx, 4, dish["qty"])
            set_existing_cell_value(ws, row_idx, 5, dish["mon"])
            set_existing_cell_value(ws, row_idx, 6, dish["recipe"])
            row_idx += 1

def preserve_blank_styled_cells(ws, original_cell_keys):
    for key in original_cell_keys:
        cell = ws._cells.get(key)
        if cell is None:
            continue
        if isinstance(cell, MergedCell):
            continue
        if cell.value is None:
            cell.value = ""

def capitalize_first_alpha_per_line(value):
    if not isinstance(value, str) or not value or value.startswith("="):
        return value

    result = []
    for line in value.split("\n"):
        chars = list(line)
        for idx, char in enumerate(chars):
            if char.isspace():
                continue
            if char.isalpha():
                chars[idx] = char.upper()
            break
        result.append("".join(chars))
    return "\n".join(result)

def capitalize_format2_text_cells(ws, sheet_name):
    max_col = 9 if sheet_name == "Bếp trung tâm 2" else 10
    for row in range(1, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = capitalize_first_alpha_per_line(cell.value)

def get_format2_print_data_columns(sheet_name):
    if sheet_name == "Bếp trung tâm 2":
        return (2, 3, 4, 5, 6, 7, 8, 9)
    return (3, 4, 5, 8, 9, 10)

def tune_format2_print_columns(ws, sheet_name):
    if sheet_name == "Bếp tại chỗ":
        widths = {
            "A": 6, "B": 7, "C": 9, "D": 38, "E": 40,
            "F": 6, "G": 7, "H": 18, "I": 38, "J": 40,
        }
    elif sheet_name == "Bếp trung tâm":
        widths = {"D": 58, "E": 24, "I": 58, "J": 24}
    elif sheet_name == "Bếp trung tâm 2":
        widths = {
            "A": 9, "B": 13, "C": 14, "D": 42, "E": 26,
            "F": 14, "G": 10, "H": 42, "I": 26,
        }
    else:
        widths = {}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def apply_format2_readable_fonts_and_wrap(ws, sheet_name):
    max_col = 9 if sheet_name == "Bếp trung tâm 2" else 10
    center_cols = {1, 2, 3, 6, 7, 8}
    if sheet_name == "Bếp trung tâm 2":
        center_cols = {1, 2, 3, 4, 6, 7}

    for row in range(1, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue

            font = copy(cell.font)
            font_name = "Aptos Narrow"
            if row == 1:
                font_size = 11
            elif row == 2:
                font_size = 10
            elif sheet_name == "Bếp tại chỗ":
                font_size = 9.5
            elif sheet_name == "Bếp trung tâm 2":
                font_size = 9 if row >= 41 and col in {5, 6} else 9.5
            else:
                font_size = 9

            cell.font = Font(
                name=font_name,
                size=font_size,
                bold=font.bold,
                italic=font.italic,
                color=font.color,
            )

            horizontal = "center" if col in center_cols else "left"
            vertical = "center" if row <= 2 or col in center_cols else "top"
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical=vertical,
                wrap_text=True,
                text_rotation=cell.alignment.text_rotation,
            )

def optimize_format2_row_heights_for_print(ws, sheet_name):
    data_cols = get_format2_print_data_columns(sheet_name)
    empty_height = 4 if sheet_name == "Bếp tại chỗ" else 6

    if ws.max_row >= 1:
        ws.row_dimensions[1].height = 18 if sheet_name == "Bếp tại chỗ" else 20
    if ws.max_row >= 2:
        ws.row_dimensions[2].height = 20

    for row in range(3, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in data_cols]
        if not any(value not in (None, "") for value in values):
            ws.row_dimensions[row].height = empty_height
            continue

        max_lines = 1
        for col, value in zip(data_cols, values):
            if value in (None, ""):
                continue
            cell = ws.cell(row, col)
            cell.alignment = copy(cell.alignment)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or "top",
                wrap_text=True,
                text_rotation=cell.alignment.text_rotation,
            )
            col_letter = get_column_letter(col)
            width = ws.column_dimensions[col_letter].width or 12
            font_size = float(cell.font.sz or 11)
            chars_per_line = max(6, int(width * 11 / font_size * 1.05))
            wrapped_lines = 0
            for line in str(value).split("\n"):
                wrapped_lines += max(1, (len(line) + chars_per_line - 1) // chars_per_line)
            max_lines = max(max_lines, wrapped_lines)

        if sheet_name == "Bếp tại chỗ":
            ws.row_dimensions[row].height = min(46, max(20, max_lines * 12 + 6))
        elif sheet_name == "Bếp trung tâm 2" and row >= 41:
            ws.row_dimensions[row].height = min(58, max(20, max_lines * 9 + 6))
        else:
            ws.row_dimensions[row].height = min(38, max(18, max_lines * 9 + 6))

def configure_format2_one_page_print(ws, sheet_name):
    max_print_col = 9 if sheet_name == "Bếp trung tâm 2" else 10
    max_print_row = ws.max_row
    while max_print_row > 1:
        has_content_or_style = False
        for col in range(1, max_print_col + 1):
            cell = ws.cell(max_print_row, col)
            if cell.value not in (None, "") or cell.has_style:
                has_content_or_style = True
                break
        if has_content_or_style:
            break
        max_print_row -= 1

    ws.print_area = f"A1:{get_column_letter(max_print_col)}{max_print_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_margins.left = 0.12
    ws.page_margins.right = 0.12
    ws.page_margins.top = 0.18
    ws.page_margins.bottom = 0.18
    ws.page_margins.header = 0
    ws.page_margins.footer = 0
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.sheet_view.showGridLines = False

def populate_template_workbook_fill_only(all_parsed_rows, selected_kitchen=None):
    template_path = Path("Mẫu/Mẫu format 2.xlsx")
    if not template_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file mẫu tại {template_path}")

    wb = openpyxl.load_workbook(template_path)
    original_cell_keys_by_sheet = {
        sheet_name: set(wb[sheet_name]._cells.keys())
        for sheet_name in wb.sheetnames
    }
    grouped_data, central_veggie_rows = group_format2_records_for_fill_only(all_parsed_rows)

    for sheet_name in wb.sheetnames:
        if sheet_name not in ["Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"]:
            continue
        ws = wb[sheet_name]
        panel_groups = build_fill_only_template_group_specs(ws, sheet_name)
        clear_fill_only_data_slots(ws, sheet_name, panel_groups)
        fill_template_header_dates_only(ws, all_parsed_rows)
        fill_standard_template_slots_only(ws, sheet_name, grouped_data, panel_groups)
        if sheet_name == "Bếp trung tâm 2":
            fill_central_veggie_template_slots_only(ws, central_veggie_rows)
        preserve_blank_styled_cells(ws, original_cell_keys_by_sheet[sheet_name])
        capitalize_format2_text_cells(ws, sheet_name)
        tune_format2_print_columns(ws, sheet_name)
        apply_format2_readable_fonts_and_wrap(ws, sheet_name)
        optimize_format2_row_heights_for_print(ws, sheet_name)
        configure_format2_one_page_print(ws, sheet_name)

    if selected_kitchen and selected_kitchen != "Đầy đủ dữ liệu":
        for name in list(wb.sheetnames):
            if name in ["Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"] and name != selected_kitchen:
                wb.remove(wb[name])
    return wb

def populate_template_workbook(all_parsed_rows, selected_kitchen=None):
    # Load template
    template_path = Path("Mẫu/Mẫu format 2.xlsx")
    if not template_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file mẫu tại {template_path}")

    wb = openpyxl.load_workbook(template_path)
    for name in ["Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"]:
        if name in wb.sheetnames:
            unmerge_and_fill_all_template_company_and_ca_cells(wb[name])

    # Group input data by (Kitchen, Company, Shift/Category)
    grouped_data = defaultdict(list)
    for r in all_parsed_rows:
        bep = str(r.get("bep") or "").strip()
        cust = str(r.get("ma_kh") or "").strip()
        ca = str(r.get("ca") or "").strip()

        # Use customer specific rule for Loại món ăn (to resolve key)
        rule = get_rule_for_customer(cust)
        loai_mon_resolved = rule.get_loai_mon_val(r)

        co_cau_str = str(r.get("co_cau") or "").lower()
        lm_str = str(loai_mon_resolved or "").lower()

        is_chay_rec = is_vegetarian_row(r)
        if is_chay_rec and bep in ["Bếp trung tâm", "Bếp trung tâm 2"]:
            continue

        cust_norm = normalize_comp(cust)

        if is_chay_rec and bep == "Bếp trung tâm 2":
            comp_key = get_chay_record_group_key(cust_norm, normalize_template_shift_or_category(ca))
            key_norm = ""
        else:
            is_com_khach = "khach" in strip_accents(co_cau_str) or "khach" in strip_accents(lm_str)
            if is_com_khach:
                price = None
                for p_val in ["50", "70", "40"]:
                    if p_val in co_cau_str or p_val in lm_str:
                        price = p_val
                        break

                if cust_norm == "scavi" and price == "50":
                    comp_key = "scavicomkhach50k"
                    key_norm = ""
                elif cust_norm == "scavi" and price == "70":
                    comp_key = "scavicomkhach70k"
                    key_norm = ""
                elif cust_norm == "sofa" and price == "40":
                    comp_key = "sofacomql40k"
                    key_norm = ""
                else:
                    comp_key = cust_norm
                    key_norm = ""
            else:
                comp_key = cust_norm
                lm_clean = strip_accents(str(loai_mon_resolved).lower().strip()).replace(" ", "")
                lm_clean_base = re.sub(r"\(.*?\)", "", lm_clean).strip()

                is_main = lm_clean_base in ["monman", "suatman", "suatcom", "com"] or \
                          "suatcom" in lm_clean or "suatman" in lm_clean or "monman" in lm_clean

                if is_main:
                    key_norm = normalize_template_shift_or_category(ca)
                else:
                    cat = normalize_template_shift_or_category(loai_mon_resolved)
                    if not cat:
                        cat = normalize_template_shift_or_category(ca)

                    if cat in ["nuoc", "chao", "trangmieng", "cai_thien", "chay"]:
                        ca_clean = normalize_template_shift_or_category(ca)
                        if ca_clean not in ["ca1", "ca2", "ca3"]:
                            ca_clean = "ca1"
                        key_norm = f"{cat}_{ca_clean}"
                    else:
                        key_norm = cat

        grouped_data[(bep, comp_key, key_norm)].append(r)

    # Get Date
    dates = sorted(list(set(str(r.get("ngay", "")).strip() for r in all_parsed_rows if r.get("ngay"))))
    date_str = ", ".join(dates) if dates else ""
    weekday_str = get_vietnamese_weekday(date_str) if date_str else ""

    # Loop over sheets in template
    for sheet_name in wb.sheetnames:
        if sheet_name not in ["Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"]:
            continue

        ws = wb[sheet_name]
        if sheet_name == "Bếp trung tâm 2":
            clear_bt2_veggie_template(ws)

        is_btc = (sheet_name == "Bếp tại chỗ")

        # The current template already has SL and Khối lượng columns for all
        # kitchen sheets, so do not insert columns into central kitchen sheets.
        is_btc = True

        # Configure A4 Portrait printing with fit-to-width
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Update Date at D1/E1/F1
        ws["D1"] = weekday_str.strip().rstrip("-").strip() + " -"
        ws["E1"] = date_str
        if is_btc:
            ws["F1"] = date_str

        p1_start = 1
        p2_start = 6  # for both Bếp tại chỗ and central kitchens (which have a col inserted at pos 5)

        for panel_idx, start_col in enumerate([p1_start, p2_start], start=1):
            lookup = merged_value_lookup(ws)
            template_groups = defaultdict(list)
            frame_cols = list(range(start_col, start_col + 5))
            for r in range(3, ws.max_row + 1):
                if not row_has_format2_panel_frame(ws, r, frame_cols):
                    continue

                cty_val = get_row_company(ws, lookup, r, start_col)
                shift_val = get_row_shift(ws, lookup, r, start_col + 1, sheet_name, panel_idx)

                if cty_val:
                    comp_norm = normalize_comp(cty_val)
                    cat = normalize_template_shift_or_category(shift_val) if shift_val else ""
                    hint_cat = infer_template_category_from_dish_hint(
                        effective_cell_value(ws, lookup, r, start_col + 3)
                    )
                    if hint_cat and cat in ["ca1", "ca2", "ca3"]:
                        key_norm = f"{hint_cat}_{cat}"
                    elif cat in ["nuoc", "chao", "trangmieng", "cai_thien", "chay"]:
                        active_main = get_row_active_main_shift(ws, lookup, r, start_col + 1)
                        key_norm = f"{cat}_{active_main}"
                    else:
                        key_norm = cat

                    template_groups[(comp_norm, key_norm)].append(r)

            original_group_keys = set(template_groups.keys())
            original_company_keys = set(k[0] for k in original_group_keys)

            # Dynamic empty rows allocation to prevent skipped categories/shifts
            free_rows = defaultdict(list)

            # Add all truly empty template rows first
            for c_key in list(set(k[0] for k in template_groups.keys())):
                empty_rows = template_groups.get((c_key, ""))
                if empty_rows:
                    free_rows[c_key].extend(empty_rows)
                    template_groups[(c_key, "")] = [] # clear it temporarily

            # For each company and key, check if we have daily records.
            # If so, reclaim any excess template rows.
            # If we don't have daily records, reclaim ALL template rows for that key!
            for (c_key, k_key), row_list in list(template_groups.items()):
                if k_key == "":
                    continue
                # Get daily records for this slot
                records = grouped_data.get((sheet_name, c_key, k_key), [])
                if records:
                    dishes_data = format_group_dishes(records[0].get("ma_kh"), records)
                    n_needed = len(dishes_data)
                else:
                    n_needed = 0

                n_avail = len(row_list)
                if n_avail > n_needed:
                    # Pop excess rows
                    excess = row_list[n_needed:]
                    template_groups[(c_key, k_key)] = row_list[:n_needed]
                    free_rows[c_key].extend(excess)

            # 2. Allocate the free rows to daily slots that were skipped or need more rows.
            # If we run out of free rows, we insert new rows!
            ca_display_map = {}
            for (b_key, c_key, k_key), records in list(grouped_data.items()):
                if b_key == sheet_name:
                    if k_key is not None and should_process_group_in_panel(sheet_name, c_key, k_key, panel_idx, template_groups):
                        row_list = template_groups.get((c_key, k_key), [])
                        n_avail = len(row_list)

                        dishes_data = format_group_dishes(records[0].get("ma_kh"), records)
                        n_needed = len(dishes_data)

                        n_still_needed = n_needed - n_avail
                        if n_still_needed > 0:
                            # Sort free rows so they are in top-to-bottom order
                            free_rows[c_key].sort()

                            allocated = []
                            # Take what we can from free_rows
                            for _ in range(min(n_still_needed, len(free_rows[c_key]))):
                                allocated.append(free_rows[c_key].pop(0))

                            # If we still need more rows, we insert them!
                            n_still_needed_after_free = n_still_needed - len(allocated)
                            if n_still_needed_after_free > 0:
                                # Find the last row number of this company block
                                company_rows = []
                                for key, r_list in template_groups.items():
                                    if key[0] == c_key:
                                        company_rows.extend(r_list)
                                # Also include any remaining free rows for this company
                                company_rows.extend(free_rows[c_key])
                                # Also include already allocated rows for this new key
                                company_rows.extend(allocated)
                                # Also include original rows for this key
                                company_rows.extend(row_list)

                                if company_rows:
                                    last_row = max(company_rows)
                                    insert_at = last_row + 1
                                else:
                                    insert_at = ws.max_row + 1

                                # Insert rows in the sheet
                                insert_template_rows(ws, insert_at, n_still_needed_after_free, start_col)

                                # Shift row indices in template_groups
                                shift_template_groups(template_groups, insert_at, n_still_needed_after_free)

                                # Shift row indices in free_rows
                                for comp_key in free_rows:
                                    free_rows[comp_key] = [r + n_still_needed_after_free if r >= insert_at else r for r in free_rows[comp_key]]

                                # Shift already allocated rows if any
                                allocated = [r + n_still_needed_after_free if r >= insert_at else r for r in allocated]

                                # Shift original row_list if any of them are at or below insert_at
                                row_list = [r + n_still_needed_after_free if r >= insert_at else r for r in row_list]

                                # The new rows start at insert_at!
                                for r_new in range(insert_at, insert_at + n_still_needed_after_free):
                                    allocated.append(r_new)

                            new_row_list = sorted(row_list + allocated)
                            template_groups[(c_key, k_key)] = new_row_list
                            ca_display_map[(c_key, k_key)] = records[0].get("co_cau") or records[0].get("ca")

            # 3. Put any remaining unused free rows back into the empty slot so they get cleared
            for c_key, rows in free_rows.items():
                if rows:
                    template_groups[(c_key, "")].extend(rows)

            for (comp_norm, key_norm), row_list in template_groups.items():
                data_rows = find_format2_fallback_data_rows(grouped_data, sheet_name, comp_norm, key_norm)

                dishes_data = []
                if data_rows:
                    dishes_data = format_group_dishes(data_rows[0].get("ma_kh"), data_rows)
                    dishes_data = compact_format2_dishes_for_slots(dishes_data, len(row_list))

                # Determine custom ca_display
                ca_display = ca_display_map.get((comp_norm, key_norm))
                if not ca_display:
                    for r_idx in row_list:
                        orig = effective_cell_value(ws, lookup, r_idx, start_col + 1)
                        if orig:
                            ca_display = orig
                            break
                if not ca_display:
                    if key_norm.startswith("chay"):
                        ca_display = "Chay"
                    elif key_norm.startswith("chao"):
                        ca_display = "M.cháo"
                    elif key_norm.startswith("nuoc"):
                        ca_display = "M.nước"
                    elif key_norm.startswith("trangmieng"):
                        ca_display = "T.miệng"
                    elif key_norm.startswith("cai_thien"):
                        ca_display = "Cải thiện"
                    elif key_norm.startswith("ca"):
                        m = re.search(r'ca\d', key_norm)
                        if m:
                            ca_display = m.group(0).upper().replace("CA", "Ca ")
                        else:
                            ca_display = key_norm.upper()

                def safe_write_cell(row, col, val, align=None):
                    c = ws.cell(row, col)
                    if type(c).__name__ == "MergedCell":
                        return
                    c.value = val
                    if align:
                        c.alignment = align

                for idx, row_num in enumerate(row_list):
                    if idx < len(dishes_data):
                        dish = dishes_data[idx]

                        # Col 1: Cty (only for new blocks that had no template rows)
                        if (comp_norm, key_norm) not in original_group_keys and comp_norm not in original_company_keys:
                            comp_display = data_rows[0].get("ma_kh")
                            if comp_display:
                                safe_write_cell(row_num, start_col, comp_display, Alignment(horizontal="center", vertical="center", wrap_text=True))

                        # Col 2: Ca
                        ca_val = ca_display or dish.get("ca_display")
                        if ca_val:
                            ca_col = start_col + 1
                            # Dynamically unmerge if the Ca cell is part of any merged cells range
                            for rng in list(ws.merged_cells.ranges):
                                if rng.min_col <= ca_col <= rng.max_col and rng.min_row <= row_num <= rng.max_row:
                                    ws.unmerge_cells(start_row=rng.min_row, start_column=rng.min_col,
                                                     end_row=rng.max_row, end_column=rng.max_col)
                                    break
                            safe_write_cell(row_num, ca_col, ca_val, Alignment(horizontal="center", vertical="center", wrap_text=True))

                        # Col 3: Quantity
                        safe_write_cell(row_num, start_col + 2, dish["qty"], Alignment(horizontal="center", vertical="center", wrap_text=True))

                        # Col 4: Món ăn
                        safe_write_cell(row_num, start_col + 3, dish["mon"], Alignment(horizontal="left", vertical="center", wrap_text=True))

                        # Col 5: Khối lượng đi chợ (only Bếp tại chỗ)
                        if is_btc:
                            safe_write_cell(row_num, start_col + 4, dish["recipe"], Alignment(horizontal="left", vertical="center", wrap_text=True))

                        # Dynamically estimate lines for text columns to prevent vertical cutoff
                        col_letter_mon = get_column_letter(start_col + 3)
                        col_width_mon = ws.column_dimensions[col_letter_mon].width or 65
                        chars_per_line_mon = max(10, int(col_width_mon * 1.1))
                        str_mon = str(dish["mon"])
                        lines_mon = sum(max(1, (len(line) + chars_per_line_mon - 1) // chars_per_line_mon) for line in str_mon.split("\n"))

                        lines_qty = str(dish["qty"]).count("\n") + 1

                        if is_btc:
                            col_letter_rec = get_column_letter(start_col + 4)
                            col_width_rec = ws.column_dimensions[col_letter_rec].width or 35
                            chars_per_line_rec = max(10, int(col_width_rec * 1.2))
                            str_rec = str(dish["recipe"])
                            lines_rec = sum(max(1, (len(line) + chars_per_line_rec - 1) // chars_per_line_rec) for line in str_rec.split("\n"))
                        else:
                            lines_rec = 1

                        max_lines = max(lines_qty, lines_mon, lines_rec)
                        curr_height = ws.row_dimensions[row_num].height or 15
                        ws.row_dimensions[row_num].height = max(curr_height, max_lines * 15 + 5)
                    else:
                        # Clear unused row
                        safe_write_cell(row_num, start_col + 2, None)
                        safe_write_cell(row_num, start_col + 3, None)
                        if is_btc:
                            safe_write_cell(row_num, start_col + 4, None)

        # Clean up unused and empty rows for this sheet to prevent empty spaces when printing
        delete_unused_rows(ws, sheet_name)
        if sheet_name == "Bếp trung tâm 2":
            append_central_veggie_table(ws, all_parsed_rows)
        apply_reference_menu_style(ws, sheet_name)
        capitalize_format2_text_cells(ws, sheet_name)
        tune_format2_print_columns(ws, sheet_name)
        apply_format2_readable_fonts_and_wrap(ws, sheet_name)
        optimize_format2_row_heights_for_print(ws, sheet_name)
        configure_format2_one_page_print(ws, sheet_name)

    if selected_kitchen and selected_kitchen != "Đầy đủ dữ liệu":
        for name in list(wb.sheetnames):
            if name in ["Bếp tại chỗ", "Bếp trung tâm", "Bếp trung tâm 2"] and name != selected_kitchen:
                wb.remove(wb[name])
    return wb

def process_sheet_format2(ws_source, file_path, date_mode="auto", selected_kitchen=None, quantity_mode=None):
    # Try to find all Excel files in Excel/ folder
    excel_dir = file_path.parent
    if excel_dir.name != "Excel":
        excel_dir = Path("Excel")

    excel_files = []
    if excel_dir.exists():
        day_du_files = list(excel_dir.glob("*IN (đầy đủ).xlsx")) + list(excel_dir.glob("*IN (đầy đủ).xlsm"))
        if day_du_files:
            excel_files = day_du_files
        else:
            excel_files = sorted(excel_dir.glob("*.xlsx")) + sorted(excel_dir.glob("*.xlsm"))
            excel_files = [f for f in excel_files if not f.name.startswith("~$")]

    all_parsed_rows = []
    if excel_files:
        print(f"Format 2: Phát hiện {len(excel_files)} file trong {excel_dir}. Tiến hành gộp dữ liệu...")
        for p in excel_files:
            try:
                wb_temp = openpyxl.load_workbook(p, data_only=True)
                ws_temp = wb_temp.active
                parsed = parse_source_sheet(ws_temp, p, quantity_mode)
                all_parsed_rows.extend(parsed)
            except Exception as e:
                print(f"Lỗi khi đọc file {p.name}: {e}")
    else:
        parsed = parse_source_sheet(ws_source, file_path, quantity_mode)
        all_parsed_rows.extend(parsed)

    return populate_template_workbook_fill_only(all_parsed_rows, selected_kitchen)
