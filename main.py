from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote
import argparse
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding='utf-8')
import webbrowser
import html
import csv

# Cho phép chạy script trong thư mục này dù virtualenv được tạo ở máy khác.
for site_packages in (Path(__file__).resolve().parent / ".venv" / "lib").glob(
    "python*/site-packages"
):
    if site_packages.exists():
        sys.path.insert(0, str(site_packages))

from openpyxl import load_workbook, Workbook

# Import components
from components.common import (
    ensure_visible_worksheet,
    safe_filename,
    output_filename_for_format,
    content_disposition_filename,
    max_upload_bytes,
    input_dir,
    output_dir,
    APPROVAL_FORMAT_MODE,
    QUANTITY_MODE_DEFAULT,
    normalize_quantity_mode,
)
from components.format_bizen_po import BIZEN_IDENTIFIER, BIZEN_FORMAT_MODE, can_process_bizen_po
from components.format_bizen_po_export import BIZEN_EXPORT_IDENTIFIER, BIZEN_EXPORT_FORMAT_MODE, can_process_bizen_po_export
from components import (
    process_sheet_format1_a4,
    process_sheet_format2,
    process_sheet_approval,
    process_sheet_bizen_po,
    process_sheet_bizen_po_export,
)


def _detect_bizen_file(filename):
    """Trả về True nếu tên file chứa chuỗi nhận diện BIZEN PO."""
    return BIZEN_IDENTIFIER in (filename or "")


def _detect_bizen_export_file(filename):
    """Trả về True nếu tên file chứa chuỗi nhận diện BIZEN PO Xuất."""
    return BIZEN_EXPORT_IDENTIFIER in (filename or "")


def format_workbook_bytes(
    file_bytes,
    filename,
    date_mode="auto",
    format_mode="format1",
    selected_kitchen=None,
    quantity_mode=QUANTITY_MODE_DEFAULT,
):
    if not file_bytes:
        raise ValueError("File Excel đang trống.")

    safe_name = safe_filename(filename)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".csv"}:
        raise ValueError("Vui lòng chọn file Excel .xlsx, .xlsm hoặc .csv.")

    # Auto-detect BIZEN PO files by filename
    if _detect_bizen_export_file(safe_name):
        format_mode = BIZEN_EXPORT_FORMAT_MODE
    elif _detect_bizen_file(safe_name):
        format_mode = BIZEN_FORMAT_MODE

    if suffix == ".csv":
        from io import StringIO
        workbook = Workbook()
        worksheet = workbook.active
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(StringIO(text))
        for row in reader:
            worksheet.append(row)
    else:
        workbook = load_workbook(BytesIO(file_bytes))
        worksheet = ensure_visible_worksheet(workbook, workbook.active)

    # Auto-detect by headers: nếu chưa chọn BIZEN mà file có đủ cột → tự chuyển
    if format_mode not in (BIZEN_FORMAT_MODE, BIZEN_EXPORT_FORMAT_MODE):
        if can_process_bizen_po_export(worksheet):
            format_mode = BIZEN_EXPORT_FORMAT_MODE
        elif can_process_bizen_po(worksheet):
            format_mode = BIZEN_FORMAT_MODE

    quantity_mode = normalize_quantity_mode(quantity_mode)

    if format_mode == BIZEN_EXPORT_FORMAT_MODE:
        workbook = process_sheet_bizen_po_export(worksheet)
    elif format_mode == BIZEN_FORMAT_MODE:
        workbook = process_sheet_bizen_po(worksheet)
    elif format_mode == "format2":
        workbook = process_sheet_format2(
            worksheet,
            Path(safe_name),
            date_mode,
            selected_kitchen,
            quantity_mode,
        )
    elif format_mode == APPROVAL_FORMAT_MODE:
        workbook = process_sheet_approval(worksheet)
    else:
        workbook = process_sheet_format1_a4(worksheet, quantity_mode)

    output = BytesIO()
    ensure_visible_worksheet(workbook)
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def parse_multipart_form(headers, body):
    content_type = headers.get("Content-Type", "")
    marker = "boundary="
    if marker not in content_type:
        raise ValueError("Form upload không hợp lệ.")

    boundary = content_type.split(marker, 1)[1].split(";", 1)[0].strip().strip('"')
    if not boundary:
        raise ValueError("Không đọc được boundary của file upload.")

    delimiter = b"--" + boundary.encode("utf-8")
    fields = {}
    file_bytes = None
    filename = None

    for raw_part in body.split(delimiter):
        part = raw_part.strip()
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2].strip()
        if b"\r\n\r\n" not in part:
            continue

        raw_headers, val_data = part.split(b"\r\n\r\n", 1)
        header_text = raw_headers.decode("utf-8", "replace")
        disposition_line = next(
            (
                line
                for line in header_text.splitlines()
                if line.lower().startswith("content-disposition:")
            ),
            "",
        )
        disposition_params = {}
        for segment in disposition_line.split(";")[1:]:
            if "=" not in segment:
                continue
            key, value = segment.split("=", 1)
            disposition_params[key.strip().lower()] = value.strip().strip('"')

        name = disposition_params.get("name")
        if not name:
            continue

        if val_data.endswith(b"\r\n"):
            val_data = val_data[:-2]

        if "filename" in disposition_params or "filename*" in disposition_params:
            if name == "excel_file":
                file_bytes = val_data
                fname = "file.xlsx"
                if "filename*" in disposition_params:
                    encoded_filename = disposition_params["filename*"]
                    if "''" in encoded_filename:
                        encoded_filename = encoded_filename.split("''", 1)[1]
                    fname = unquote(encoded_filename)
                elif "filename" in disposition_params:
                    fname = disposition_params["filename"]
                filename = safe_filename(fname)
        else:
            fields[name] = val_data.decode("utf-8", "ignore").strip()

    return fields, filename, file_bytes


index_html = """<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Định dạng BOM đi chợ</title>
  <style>
    :root {
      --bg: #f4f6f8;
      --panel: #ffffff;
      --text: #17202a;
      --muted: #5b6673;
      --line: #d8dee6;
      --accent: #145c52;
      --accent-strong: #0d4039;
      --soft: #e8f3f0;
      --danger: #9f2d2d;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      font-family: Arial, Helvetica, sans-serif;
      color: var(--text);
      background:
        linear-gradient(180deg, #ffffff 0, var(--bg) 260px),
        var(--bg);
    }
    main {
      width: min(920px, calc(100% - 32px));
      margin: 0 auto;
      padding: 40px 0;
    }
    .top {
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 22px;
    }
    h1 {
      margin: 0;
      font-size: 30px;
      line-height: 1.2;
      letter-spacing: 0;
    }
    .badge {
      border: 1px solid var(--line);
      background: var(--panel);
      color: var(--muted);
      padding: 9px 12px;
      border-radius: 6px;
      font-size: 14px;
      white-space: nowrap;
    }
    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: 0 16px 36px rgba(31, 42, 55, 0.08);
      overflow: hidden;
    }
    .form {
      padding: 26px;
      display: grid;
      gap: 20px;
    }
    .drop {
      display: grid;
      place-items: center;
      min-height: 260px;
      border: 2px dashed #aab5c2;
      border-radius: 8px;
      background: #fbfcfd;
      cursor: pointer;
      transition: border-color .18s ease, background .18s ease;
    }
    .drop:hover, .drop.is-dragover {
      border-color: var(--accent);
      background: var(--soft);
    }
    .drop input {
      position: absolute;
      opacity: 0;
      pointer-events: none;
    }
    .drop-inner {
      display: grid;
      justify-items: center;
      gap: 12px;
      text-align: center;
      padding: 22px;
    }
    .icon {
      width: 58px;
      height: 58px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      background: var(--soft);
      color: var(--accent);
    }
    .icon svg {
      width: 30px;
      height: 30px;
    }
    .file-title {
      font-size: 20px;
      font-weight: 700;
    }
    .file-subtitle {
      color: var(--muted);
      font-size: 15px;
    }
    .actions {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      flex-wrap: wrap;
    }
    .status {
      min-height: 24px;
      color: var(--muted);
      font-size: 15px;
    }
    .status.error {
      color: var(--danger);
      font-weight: 700;
    }
    button {
      min-width: 190px;
      border: 0;
      border-radius: 6px;
      background: var(--accent);
      color: #ffffff;
      padding: 14px 20px;
      font-size: 16px;
      font-weight: 700;
      cursor: pointer;
    }
    button:hover { background: var(--accent-strong); }
    button:disabled {
      background: #9aa7b2;
      cursor: not-allowed;
    }
    .steps {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      border-top: 1px solid var(--line);
      background: #f9fafb;
    }
    .step {
      padding: 18px 20px;
      border-right: 1px solid var(--line);
      color: var(--muted);
      font-size: 14px;
    }
    .step:last-child { border-right: 0; }
    .step strong {
      display: block;
      color: var(--text);
      font-size: 15px;
      margin-bottom: 4px;
    }
    .options-row {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
      margin-bottom: 8px;
    }
    .control-group {
      display: flex;
      flex-direction: column;
      gap: 6px;
    }
    .control-group label {
      font-size: 12px;
      font-weight: 700;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: 0.5px;
    }
    .control-group select {
      padding: 12px;
      font-size: 14px;
      font-family: inherit;
      border: 1px solid var(--line);
      border-radius: 6px;
      background-color: var(--panel);
      color: var(--text);
      outline: none;
      cursor: pointer;
      transition: border-color .15s ease, box-shadow .15s ease;
    }
    .control-group select:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(20, 92, 82, 0.15);
    }
    @media (max-width: 720px) {
      .options-row { grid-template-columns: 1fr; }
      main { width: min(100% - 20px, 920px); padding: 24px 0; }
      .top { align-items: flex-start; flex-direction: column; }
      h1 { font-size: 25px; }
      .drop { min-height: 220px; }
      .actions { align-items: stretch; flex-direction: column; }
      button { width: 100%; }
      .steps { grid-template-columns: 1fr; }
      .step { border-right: 0; border-bottom: 1px solid var(--line); }
      .step:last-child { border-bottom: 0; }
    }
  </style>
</head>
<body>
  <main>
    <div class="top">
      <h1>Định dạng BOM đi chợ</h1>
      <div class="badge">A4 dọc/ngang · dễ in · gộp dữ liệu</div>
    </div>
    <section class="panel">
      <form class="form" action="/format" method="post" enctype="multipart/form-data">
        <label class="drop" id="drop">
          <input id="file" name="excel_file" type="file" accept=".xlsx,.xlsm" required>
          <span class="drop-inner">
            <span class="icon" aria-hidden="true">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
                <path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"></path>
                <path d="M14 2v6h6"></path>
                <path d="M12 18v-6"></path>
                <path d="m9 15 3 3 3-3"></path>
              </svg>
            </span>
            <span class="file-title" id="fileTitle">Chọn file Excel</span>
            <span class="file-subtitle">Kéo thả file vào đây hoặc bấm để chọn</span>
          </span>
        </label>
        <div class="options-row">
          <div class="control-group">
            <label for="format_mode">Bố cục xuất file</label>
            <select id="format_mode" name="format_mode">
              <option value="format1" selected>Format 1 (Bảng A4 dọc 13 cột + dòng ngày)</option>
              <option value="format2">Format 2 (Theo mẫu)</option>
              <option value="duyet_dinh_muc">Duyệt định mức (A4 ngang, gộp món liền kề)</option>
              <option value="bizen_po">BIZEN PO Lưới (Đặt hàng 9 cột)</option>
              <option value="bizen_po_export">BIZEN PO Xuất (Có mẫu)</option>
            </select>
          </div>
          <div class="control-group" id="quantityGroup">
            <label for="quantity_mode">Số lượng để in</label>
            <select id="quantity_mode" name="quantity_mode">
              <option value="approved" selected>Số lượng cô Nga duyệt</option>
              <option value="forecast">Số lượng dự báo</option>
            </select>
          </div>
        </div>
        <div class="actions">
          <div class="status" id="status">Chưa chọn file</div>
          <button id="submit" type="submit" disabled>Xuất file đã định dạng</button>
        </div>
      </form>
      <div class="steps">
        <div class="step"><strong>1. Chọn file</strong>File BOM Excel cần in.</div>
        <div class="step"><strong>2. Xuất file</strong>Chọn bố cục cần dùng rồi tải file.</div>
        <div class="step"><strong>3. In A4</strong>Mở file tải về rồi in theo thiết lập sẵn.</div>
      </div>
    </section>
  </main>
  <script>
    const drop = document.getElementById('drop');
    const file = document.getElementById('file');
    const title = document.getElementById('fileTitle');
    const status = document.getElementById('status');
    const submit = document.getElementById('submit');
    const form = document.querySelector('form');
    const formatMode = document.getElementById('format_mode');
    const quantityGroup = document.getElementById('quantityGroup');

    function updateQuantityVisibility() {
      const usesQuantityMode = ['format1', 'format2'].includes(formatMode.value);
      quantityGroup.hidden = !usesQuantityMode;
    }

    function updateFileName() {
      const selected = file.files && file.files[0];
      title.textContent = selected ? selected.name : 'Chọn file Excel';
      status.textContent = selected ? 'Sẵn sàng xuất file' : 'Chưa chọn file';
      status.classList.remove('error');
      submit.disabled = !selected;
    }

    file.addEventListener('change', updateFileName);
    formatMode.addEventListener('change', updateQuantityVisibility);
    drop.addEventListener('dragover', event => {
      event.preventDefault();
      drop.classList.add('is-dragover');
    });
    drop.addEventListener('dragleave', () => drop.classList.remove('is-dragover'));
    drop.addEventListener('drop', event => {
      event.preventDefault();
      drop.classList.remove('is-dragover');
      if (event.dataTransfer.files.length) {
        file.files = event.dataTransfer.files;
        updateFileName();
      }
    });
    form.addEventListener('submit', () => {
      submit.disabled = true;
      status.textContent = 'Đang xử lý file...';
      window.setTimeout(() => {
        submit.disabled = false;
        status.textContent = file.files[0] ? 'Có thể xuất lại file' : 'Chưa chọn file';
      }, 1800);
    });
    updateQuantityVisibility();
  </script>
</body>
</html>
"""


def error_html(message):
    escaped_message = html.escape(message)
    return f"""<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Không xử lý được file</title>
  <style>
    body {{
      margin: 0;
      min-height: 100vh;
      display: grid;
      place-items: center;
      font-family: Arial, Helvetica, sans-serif;
      background: #f4f6f8;
      color: #17202a;
    }}
    .box {{
      width: min(560px, calc(100% - 32px));
      background: #fff;
      border: 1px solid #d8dee6;
      border-radius: 8px;
      padding: 26px;
      box-shadow: 0 16px 36px rgba(31, 42, 55, 0.08);
    }}
    h1 {{ margin: 0 0 12px; font-size: 24px; }}
    p {{ margin: 0 0 18px; color: #5b6673; line-height: 1.5; }}
    a {{
      display: inline-block;
      background: #145c52;
      color: #fff;
      text-decoration: none;
      border-radius: 6px;
      padding: 12px 16px;
      font-weight: 700;
    }}
  </style>
</head>
<body>
  <div class="box">
    <h1>Không xử lý được file</h1>
    <p>{escaped_message}</p>
    <a href="/">Quay lại</a>
  </div>
</body>
</html>"""


class BomFormatterHandler(BaseHTTPRequestHandler):
    server_version = "BomFormatter/1.0"

    def send_text(self, status, content, content_type="text/html; charset=utf-8"):
        payload = content.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(payload)

    def do_GET(self):
        if self.path in {"/", "/index.html"}:
            self.send_text(200, index_html)
            return
        if self.path == "/health":
            self.send_text(200, "OK", "text/plain; charset=utf-8")
            return
        self.send_text(404, error_html("Không tìm thấy trang."))

    def do_POST(self):
        if self.path != "/format":
            self.send_text(404, error_html("Không tìm thấy trang."))
            return

        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            if content_length <= 0:
                raise ValueError("Vui lòng chọn file Excel trước khi xuất.")
            if content_length > max_upload_bytes:
                raise ValueError("File quá lớn. Giới hạn hiện tại là 50 MB.")

            body = self.rfile.read(content_length)
            fields, filename, file_bytes = parse_multipart_form(self.headers, body)

            if not file_bytes:
                raise ValueError("Vui lòng chọn file Excel trước khi xuất.")

            format_opt = fields.get("format_mode", "format1")
            quantity_opt = fields.get("quantity_mode", QUANTITY_MODE_DEFAULT)

            # ---------------- DIAGNOSTIC LOGGING ----------------
            try:
                import datetime
                from openpyxl import load_workbook
                import io
                wb_test = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
                ws_test = wb_test.active
                headers_test = []
                for c in range(1, min(ws_test.max_column + 1, 20)):
                    val = ws_test.cell(1, c).value
                    if val is not None:
                        headers_test.append(str(val).strip())
                with open("diagnostic_log.txt", "a", encoding="utf-8") as f:
                    f.write(f"\n--- {datetime.datetime.now()} ---\n")
                    f.write(f"Uploaded file: {filename}\n")
                    f.write(f"Format mode selected: {format_opt}\n")
                    f.write(f"Quantity mode selected: {quantity_opt}\n")
                    f.write(f"Headers row 1: {headers_test}\n")
            except Exception as e:
                with open("diagnostic_log.txt", "a", encoding="utf-8") as f:
                    f.write(f"Error logging headers: {e}\n")
            # ----------------------------------------------------

            output_bytes = format_workbook_bytes(
                file_bytes,
                filename,
                date_mode="auto",
                format_mode=format_opt,
                quantity_mode=quantity_opt,
            )
            download_name = output_filename_for_format(filename, format_opt)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header(
                "Content-Disposition", content_disposition_filename(download_name)
            )
            self.send_header("Content-Length", str(len(output_bytes)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(output_bytes)
        except Exception as exc:
            self.send_text(400, error_html(str(exc)))

    def log_message(self, format, *args):
        print(f"{self.address_string()} - {format % args}")


def run_web_server(host="127.0.0.1", port=8100, open_browser=True):
    last_error = None
    for candidate_port in range(port, port + 20):
        try:
            server = ThreadingHTTPServer((host, candidate_port), BomFormatterHandler)
            break
        except OSError as exc:
            last_error = exc
    else:
        raise RuntimeError(f"Không mở được cổng web: {last_error}")

    url = f"http://{host}:{server.server_port}"
    browser_url = f"http://127.0.0.1:{server.server_port}" if host == "0.0.0.0" else url
    print(f"Web đã sẵn sàng: {url}")
    print("Nhấn Ctrl+C để dừng.")
    if open_browser:
        webbrowser.open(browser_url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nĐã dừng web.")
    finally:
        server.server_close()


def run_batch():
    output_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(input_dir.glob("*.xlsx"))
    files += sorted(input_dir.glob("*.xlsm"))
    files += sorted(input_dir.glob("*.csv"))
    files = [f for f in files if not f.name.startswith("~$")]

    if not files:
        print(f"Không tìm thấy file Excel/CSV trong thư mục: {input_dir}")
        return

    for file_path in files:
        file_bytes = file_path.read_bytes()

        # Nếu file là BIZEN PO Lưới → chỉ chạy format BIZEN
        if _detect_bizen_export_file(file_path.name):
            outputs = (
                (BIZEN_EXPORT_FORMAT_MODE, f"{file_path.stem} - xuat po.xlsx", "BIZEN PO Xuất"),
            )
        elif _detect_bizen_file(file_path.name):
            outputs = (
                (BIZEN_FORMAT_MODE, f"{file_path.stem} - bizen po.xlsx", "BIZEN PO"),
            )
        else:
            outputs = (
                ("format1", f"{file_path.stem} - da dinh dang.xlsx", "Format 1"),
                ("format2", f"{file_path.stem} - format 2.xlsx", "Format 2"),
                (APPROVAL_FORMAT_MODE, f"{file_path.stem} - duyet dinh muc.xlsx", "Duyệt định mức"),
            )
        for format_mode, output_name, label in outputs:
            try:
                output_bytes = format_workbook_bytes(
                    file_bytes,
                    file_path.name,
                    date_mode="auto",
                    format_mode=format_mode,
                )
                out_path = output_dir / output_name
                with open(out_path, "wb") as f_out:
                    f_out.write(output_bytes)
                print(f"Đã gộp xong ({label}): {out_path}")
            except Exception as exc:
                print(f"Lỗi khi xử lý {label} cho {file_path.name}: {exc}")


def main():
    parser = argparse.ArgumentParser(description="Định dạng file BOM Excel để in đi chợ.")
    parser.add_argument(
        "--batch",
        action="store_true",
        help="Xử lý tất cả file trong thư mục Excel và lưu vào thư mục Đã gộp.",
    )
    parser.add_argument("--host", default="127.0.0.1", help="Địa chỉ web local.")
    parser.add_argument("--port", type=int, default=8100, help="Cổng web local.")
    parser.add_argument(
        "--no-browser",
        action="store_true",
        help="Không tự mở trình duyệt sau khi bật web.",
    )
    args = parser.parse_args()

    if args.batch:
        run_batch()
    else:
        run_web_server(args.host, args.port, open_browser=not args.no_browser)


if __name__ == "__main__":
    main()
