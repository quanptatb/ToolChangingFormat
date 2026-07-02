import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()

blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
pink_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

white_bold = Font(color="FFFFFF", bold=True)
black_bold = Font(color="000000", bold=True)
red_bold = Font(color="FF0000", bold=True)

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
align_rotate = Alignment(horizontal="center", vertical="center", text_rotation=90, wrap_text=True)

def apply_common_formatting(ws, max_row, rotate_col_a_f=False):
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 20

    for r in range(1, max_row + 1):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            
            if r >= 3 and c in [5, 10]:
                if getattr(cell.fill, 'fill_type', None) is None:
                    cell.fill = pink_fill
                    
            if rotate_col_a_f and r >= 3 and c in [1, 6]:
                cell.alignment = align_rotate
            elif getattr(cell.alignment, 'horizontal', None) is None:
                if c in [4, 9]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

def build_data(ws, data):
    for r, c, val, rd, cd, fill in data:
        cell = ws.cell(row=r, column=c)
        if val != "":
            try:
                cell.value = val
            except AttributeError:
                pass
        if fill:
            try:
                cell.fill = fill
            except AttributeError:
                pass
        if rd > 0 or cd > 0:
            ws.merge_cells(start_row=r, start_column=c, end_row=r+rd, end_column=c+cd)
            if fill:
                for row_idx in range(r, r+rd+1):
                    for col_idx in range(c, c+cd+1):
                        try:
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                        except AttributeError:
                            pass

def sheet1():
    ws = wb.active
    ws.title = "Bếp tại chỗ"
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Thứ 2 -"
    ws['A1'].fill = blue_fill
    ws['A1'].font = white_bold
    ws['A1'].alignment = align_center

    ws.merge_cells('E1:G1')
    ws['E1'] = "18/05/26"
    ws['E1'].fill = blue_fill
    ws['E1'].font = white_bold
    ws['E1'].alignment = align_center

    ws.merge_cells('H1:I1')
    ws['H1'] = "BẾP TẠI CHỖ"
    ws['H1'].fill = yellow_fill
    ws['H1'].font = black_bold
    ws['H1'].alignment = align_center

    ws['J1'] = "328"
    ws['J1'].fill = yellow_fill
    ws['J1'].font = red_bold
    ws['J1'].alignment = align_center
    
    for c, val in enumerate(["Cty", "", "Ca", "Món ăn", "", "Cty", "", "Ca", "Món ăn", ""], start=1):
        cell = ws.cell(row=2, column=c)
        cell.value = val
        cell.fill = blue_fill
        cell.font = white_bold
        cell.alignment = align_center

    data = [
        (3, 1, "ZAMIL", 6, 0, None),
        (3, 2, "335", 0, 0, None), (3, 3, "CA1", 2, 0, None), (3, 4, "Thịt kho trứng 100g+2 cút", 0, 0, None), (3, 5, "33.5k+670 cút", 0, 0, None),
        (4, 2, "285", 0, 0, None), (4, 4, "Cá chiên sốt mắm S8", 0, 0, None), (4, 5, "285p cá+ 50p tép: 2.5k tép", 0, 0, None),
        (5, 4, "0", 0, 0, None), 
        (6, 3, "M.nước", 0, 0, None), (6, 4, "Ổi", 0, 0, None), (6, 5, "35k ổi", 0, 0, None),
        (7, 3, "T.miệng", 0, 0, None), (7, 4, "Đậu bắp luộc", 0, 0, None),
        (8, 3, "Chay", 1, 0, None), (8, 4, "Lagu chay", 0, 0, None),
        (9, 4, "Đùi gà chay rang muối", 0, 0, None),

        (10, 1, "WACOAL", 6, 0, None),
        (10, 2, "1140+500", 3, 0, None), (10, 3, "CA1", 3, 0, None), (10, 4, "Heo quay kho dưa 40g", 0, 0, None), (10, 5, "45.6+20k", 0, 0, None),
        (11, 4, "Chả cá nha trang rim 40g", 0, 0, None), (11, 5, "45.6k+20k", 0, 0, None),
        (12, 4, "Canh bầu 100g", 0, 0, None),
        (13, 4, "Rau muống luộc 90g", 0, 0, None),
        (14, 2, "60+20", 2, 0, None), (14, 3, "T.miệng", 0, 0, None), (14, 4, "Ổi", 0, 0, None), (14, 5, "108+47k ổi", 0, 0, None),
        (15, 3, "Chay", 1, 0, None), (15, 4, "Lagu chay: 1/2 đậu, 25g tàu hũ+100g cà rốt+100g khoai tây", 0, 0, None),
        (16, 4, "Đùi gà chay rang muối 70g", 0, 0, None), (16, 5, "4.2+1.4k", 0, 0, None),

        (17, 1, "QUADRILLE", 7, 0, None),
        (17, 2, "485", 5, 0, None), (17, 3, "CA1", 3, 0, None), (17, 4, "Heo quay kho dưa 40g", 0, 0, None), (17, 5, "19.4k", 0, 0, None),
        (18, 4, "Chả cá nha trang rim 40g", 0, 0, None), (18, 5, "19.4", 0, 0, None),
        (19, 4, "Canh bầu", 0, 0, None),
        (20, 4, "Rau muống xào tỏi", 0, 0, None),
        (21, 3, "M.nước", 0, 0, None), (21, 4, "Ổi", 1, 0, None), (21, 5, "45k", 1, 0, None),
        (22, 3, "T.miệng", 0, 0, None),
        (23, 2, "12", 1, 0, None), (23, 3, "Chay", 1, 0, None), (23, 4, "Lagu chay: 1/2 đậu, 25g tàu hũ+100g cà rốt+100g khoai tây", 0, 0, None),
        (24, 4, "Đùi gà chay rang muối 70g", 0, 0, None), (24, 5, "0.8k", 0, 0, None),

        (25, 1, "SMC", 15, 0, None),
        (25, 2, "670", 2, 0, None), (25, 3, "CA1", 2, 0, None), (25, 4, "Sườn non kho dưa 70g", 0, 0, None), (25, 5, "47k sườn", 0, 0, None),
        (26, 4, "Cá chiên + mắm xoài S8 + 30g xoà", 0, 0, None), (26, 5, "s8+20k xoài xanh", 0, 0, None),
        (27, 4, "Trứng chưng nấm 0.7 trứng+20g xay+2g miến+1g nấm mèo", 0, 0, None), (27, 5, "470 trứng+13.4k xay+1.3k miến+0.7k nấm mèo", 0, 0, None),
        (28, 2, "320", 1, 0, None), (28, 3, "T.miệng", 0, 0, None), (28, 4, "Chuối", 0, 0, None), (28, 5, "118k chuối", 0, 0, None),
        (29, 3, "M.nước", 0, 0, None), (29, 4, "Nui nấu gà 300g gtu", 0, 0, None), (29, 5, "96 góc tư", 0, 0, None),
        (30, 2, "150", 2, 0, None), (30, 3, "CA2", 2, 0, None), (30, 4, "Đùi gà nướng", 0, 0, None), (30, 5, "1 đùi", 0, 0, None),
        (31, 4, "Trứng sốt thịt", 0, 0, None), (31, 5, "1 trứng+15g xay", 0, 0, None),
        (32, 4, "Đậu hũ tứ xuyên", 0, 0, None), (32, 5, "1 đậu trắng+15g xay", 0, 0, None),
        (33, 2, "35", 3, 0, None), (33, 3, "M.nước", 0, 0, None), (33, 4, "Bánh canh chả cá", 0, 0, None),
        (34, 3, "Chay", 2, 0, None), (34, 4, "Lagu chay: 1/2 đậu, 25g tàu hũ+100g cà rốt+100g khoai tây", 0, 0, None), (34, 5, "18 đậu+0.9k tàu hũ ki.....", 0, 0, None),
        (35, 4, "Đùi gà chay rang muối 70g", 0, 0, None), (35, 5, "2.5k đùi gà chay", 0, 0, None),
        (36, 4, "Cá cơm chay rim tỏi ớt 40g", 0, 0, None), (36, 5, "1.4k cá cơm chay", 0, 0, None),
        (37, 2, "0", 0, 0, None),

        (3, 6, "LIXIL", 12, 0, None),
        (3, 7, "120+120+150", 0, 0, None), (3, 8, "CA1", 4, 0, None), (3, 9, "Cá ngừ kho cà 150g", 0, 0, None), (3, 10, "18k + 18k + 22.5k", 0, 0, None),
        (4, 7, "125+145+160", 0, 0, None), (4, 9, "Đậu hũ dồn thịt sốt mắm 25g xay", 0, 0, None), (4, 10, "3k + 3.6k + 4k thịt xay", 0, 0, None),
        (5, 7, "245+265+310", 0, 0, None), (5, 9, "Sườn non kho dưa 100g", 0, 0, None), (5, 10, "24.5k + 26.5k + 31k", 0, 0, None),
        (8, 7, "98+89+125+16=328", 1, 0, yellow_fill), (8, 8, "T.miệng", 0, 0, None), (8, 9, "Chuối 100g", 0, 0, yellow_fill), (8, 10, "130k chuối", 0, 0, yellow_fill),
        (9, 8, "M.nước", 0, 0, None), (9, 9, "Bún bò gân 20g tái+30g gàu+30g gân+1c chả cây", 0, 0, yellow_fill), (9, 10, "6.7k tái + 9.8k gàu + 9.8k gân + 328c chả", 0, 0, yellow_fill),
        (10, 7, "82", 2, 0, None), (10, 8, "CA2", 2, 0, None), (10, 9, "Chả giò chiên (4 cái)", 0, 0, None), (10, 10, "11 bịch", 0, 0, None),
        (11, 9, "Cá hồi sốt bơ tỏi 100g", 0, 0, None), (11, 10, "8.2k", 0, 0, None),
        (12, 9, "0", 0, 0, None),
        (13, 7, "36", 0, 0, None), (13, 8, "CA3", 0, 0, None), (13, 9, "Phở bò 50g tái + 30g nạm + 20 bò viên + 10g xương ống", 0, 0, None), (13, 10, "1.8k tái + 1k nạm + 0.7k bò viên + 0.4k xương ống", 0, 0, None),
        (14, 7, "19+5+8+1=33", 1, 0, yellow_fill), (14, 8, "M.chay", 1, 0, None), (14, 9, "Đậu hũ kho tương hột 1.5 đậu+10g tương", 0, 0, yellow_fill), (14, 10, "50 đậu + 0.3k tương", 0, 0, yellow_fill),
        (15, 9, "Củ sen xào ớt chuông 70g sen+50g ớt chuông+30g rốt", 0, 0, None), (15, 10, "2.2k sen + 1.6k ớt chuông + 1k rốt", 0, 0, None),

        (16, 6, "ODSV", 30, 0, None),
        (16, 7, "450", 1, 0, None), (16, 8, "CA1", 5, 0, None), (16, 9, "Cá hường chiên sả S8", 0, 0, None),
        (17, 9, "Vịt kho gừng 130g", 0, 0, None), (17, 10, "58.5k", 0, 0, None),
        (18, 7, "370", 1, 0, None), (18, 9, "Tôm rim hành 2 con s50", 0, 0, None),
        (19, 9, "Vịt kho gừng 130g", 0, 0, None), (19, 10, "48.1k", 0, 0, None),
        (20, 7, "", 1, 0, None), (20, 9, "Đậu bắp luộc", 0, 0, None),
        (21, 9, "Canh đậu hũ hẹ", 0, 0, None),
        (22, 8, "T.miệng", 0, 0, None), (22, 9, "Chuối", 0, 0, yellow_fill), (22, 10, "162k chuối", 0, 0, yellow_fill),
        (23, 7, "15", 1, 0, None), (23, 8, "Cải thiện", 1, 0, None), (23, 9, "Mỹ xào hải sản (10p) - chuyên gia", 0, 0, None),
        (24, 9, "cháo chay (10p) mỗi ngày 1 đậu + 100g nấm rơm", 0, 0, None), (24, 10, "Cháo chay: 10 đậu + 1k nấm rơm", 0, 0, None),
        (25, 7, "320", 0, 0, None), (25, 8, "M.Cháo", 0, 0, None), (25, 9, "Cháo gà 140g TH + 40g file", 0, 0, None), (25, 10, "2.1g TH + 0.6k file", 0, 0, None),
        (26, 8, "M.nước", 0, 0, None), (26, 9, "Bánh canh bột xắt 20g càng ghẹ + 60g nạc + 1 trứng cút + 1 tôm s60", 0, 0, None), (26, 10, "6.4k càng ghẹ + 19.2k nạc + 320 trứng cút + 320 con tôm", 0, 0, None),
        (27, 7, "55", 1, 0, None), (27, 8, "Chay", 1, 0, None), (27, 9, "Lagu chay", 0, 0, None),
        (28, 9, "Đùi gà chay rang muối", 0, 0, None),
        (29, 7, "130", 1, 0, None), (29, 8, "CA2", 6, 0, None), (29, 9, "Cá hường chiên sả S8", 0, 0, None),
        (30, 9, "Vịt kho gừng 130g", 0, 0, None), (30, 10, "3.9k", 0, 0, None),
        (31, 7, "100", 4, 0, None), (31, 9, "Tôm rim hành 2 con s50", 0, 0, None),
        (32, 9, "Vịt kho gừng 130g", 0, 0, None), (32, 10, "13k", 0, 0, None),
        (33, 9, "Đậu bắp luộc", 0, 0, None),
        (34, 9, "Canh đậu hũ hẹ", 0, 0, None),
        (35, 9, "Chuối", 0, 0, None),
        (36, 8, "Cải thiện", 0, 0, None), (36, 9, "0", 0, 0, None),
        (37, 7, "5", 0, 0, None), (37, 8, "M.Cháo", 0, 0, None), (37, 9, "Cháo gà 140g TH + 40g file", 0, 0, None), (37, 10, "0.7k TH + 0.2k file", 0, 0, None),
        (38, 7, "80", 0, 0, None), (38, 8, "M.nước", 0, 0, None), (38, 9, "Bánh canh bột xắt 20g càng ghẹ + 60g nạc + 1 trứng cút + 1 tôm s60", 0, 0, None), (38, 10, "1.6k càng ghẹ + 4.8k nạc + 80 trứng cút + 80 con tôm", 0, 0, None),
        (39, 7, "5", 1, 0, None), (39, 8, "Chay", 1, 0, None), (39, 9, "Lagu chay", 0, 0, None),
        (40, 9, "Đùi gà chay rang muối", 0, 0, None),
        (41, 7, "35", 0, 0, None), (41, 8, "CA3", 1, 0, None), (41, 9, "Miến gà 150g TH + 50g file + 50g măng", 0, 0, None), (41, 10, "5.3k TH + 1.8k file + 1.8k măng", 0, 0, None),
        (42, 7, "45", 0, 0, None), (42, 9, "Cháo gà 140g TH + 40g file", 0, 0, None), (42, 10, "6.3k TH + 1.8k file", 0, 0, None),
        (43, 7, "", 1, 0, None), (43, 8, "Ăn sáng", 1, 0, None), (43, 9, "Mì gói xào bò", 0, 0, None),
        (44, 9, "Nui xương", 0, 0, None),
        (45, 7, "20", 1, 0, None), (45, 8, "Nhập trước ăn sáng T3", 1, 0, None), (45, 9, "Xôi mặn 25g lạp xưởng + 40g chả + 2 cút", 0, 0, None), (45, 10, "0.5k lạp xưởng + 0.8k chả + 40 cút", 0, 0, None),
        (46, 9, "Hủ tiếu gà 90g TH + 40g file", 0, 0, None), (46, 10, "1.8k TH + 0.8k file", 0, 0, None),
    ]
    build_data(ws, data)
    for r, c in [(5,4), (12,9), (36,9)]: ws.cell(row=r, column=c).alignment = align_right
    apply_common_formatting(ws, max_row=46, rotate_col_a_f=True)


def sheet2():
    ws = wb.create_sheet(title="Bếp trung tâm")
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "Thứ 2 -"
    ws['A1'].fill = blue_fill
    ws['A1'].font = white_bold
    ws['A1'].alignment = align_center

    ws.merge_cells('E1:G1')
    ws['E1'] = "18/05/26"
    ws['E1'].fill = blue_fill
    ws['E1'].font = white_bold
    ws['E1'].alignment = align_center

    ws.merge_cells('H1:I1')
    ws['H1'] = "BẾP TRUNG TÂM"
    ws['H1'].fill = yellow_fill
    ws['H1'].font = black_bold
    ws['H1'].alignment = align_center
    
    for c, val in enumerate(["Cty", "", "Ca", "Món ăn", "", "Cty", "", "Ca", "Món ăn", ""], start=1):
        cell = ws.cell(row=2, column=c)
        cell.value = val
        cell.fill = blue_fill
        cell.font = white_bold
        cell.alignment = align_center

    data = [
        (3, 1, "Watabe", 2, 0, None),
        (3, 2, "165", 1, 0, None), (3, 3, "Ca 1", 1, 0, None), (3, 4, "Tôm rang lá chanh 40g", 0, 0, None), (3, 5, "6.6k", 0, 0, None),
        (4, 4, "Thịt xào rau củ 50g", 0, 0, None), (4, 5, "8.3k nạc xắt", 0, 0, None),
        (5, 3, "T.miệng", 0, 0, None), (5, 4, "Ổi", 0, 0, None),
        
        (6, 1, "Vikyno", 2, 0, None),
        (6, 2, "140", 1, 0, None), (6, 3, "Ca 1", 1, 0, None), (6, 4, "Tôm rang lá chanh 40g", 0, 0, None), (6, 5, "5.6k", 0, 0, None),
        (7, 4, "Thịt xào rau củ 50g", 0, 0, None), (7, 5, "7k", 0, 0, None),
        (8, 4, "0", 0, 0, None),
        
        (9, 1, "Shine", 2, 0, None),
        (9, 2, "45", 1, 0, None), (9, 3, "Ca 1", 1, 0, None), (9, 4, "Tôm rang lá chanh", 0, 0, None), (9, 5, "1.8k", 0, 0, None),
        (10, 4, "Thịt xào rau củ 50g nạc xắt", 0, 0, None), (10, 5, "2.3k", 0, 0, None),
        (11, 3, "T.miệng", 0, 0, None), (11, 4, "Ổi", 0, 0, None),
        
        (12, 1, "Cát Thái", 4, 0, None),
        (12, 2, "100+210+95+300", 1, 0, None), (12, 3, "CA 1", 1, 0, None), (12, 4, "Tôm rang lá chanh 40g", 0, 0, None), (12, 5, "4k + 8.4k + 3.8k + 12k", 0, 0, None),
        (13, 4, "Thịt xào rau củ 50g", 0, 0, None), (13, 5, "5k + 10.5k + 4.8k + 15k", 0, 0, None),
        (14, 3, "T.miệng", 0, 0, None), (14, 4, "Trái cây", 0, 0, None),
        (15, 2, "180", 1, 0, None), (15, 3, "CA 2", 1, 0, None), (15, 4, "Cá basa kho tiêu 100g", 0, 0, None), (15, 5, "18k", 0, 0, None),
        (16, 4, "Mực trứng rim hành tây 50g", 0, 0, None), (16, 5, "9k", 0, 0, None),
        
        (17, 1, "Maspro", 2, 0, None),
        (17, 2, "120", 1, 0, None), (17, 3, "CA 1", 1, 0, None), (17, 4, "Cánh gà chiên nước mắm", 0, 0, None), (17, 5, "1 cái", 0, 0, None),
        (18, 4, "Chả lụa rim ớt chuông 40G", 0, 0, None), (18, 5, "4.8k", 0, 0, None),
        (19, 3, "T.miệng", 0, 0, None), (19, 4, "Ổi", 0, 0, None),
        
        (20, 1, "Medic", 4, 0, None),
        (20, 2, "95", 2, 0, None), (20, 3, "Ca 1", 2, 0, None), (20, 4, "Bánh canh mọc 70g nạc+40g mọc", 0, 0, None), (20, 5, "6.6k nạc+3.8k mọc", 0, 0, None),
        (21, 4, "0", 0, 0, None),
        (22, 4, "0", 0, 0, None),
        (23, 4, "0", 0, 0, None),
        (24, 4, "Ổi", 0, 0, None),
        
        (25, 1, "OSV", 12, 0, None),
        (25, 2, "200", 0, 0, yellow_fill), (25, 3, "Ca 1", 3, 0, None), (25, 4, "Vịt kho gừng 120g", 0, 0, yellow_fill), (25, 5, "24k", 0, 0, yellow_fill),
        (26, 2, "300", 2, 0, yellow_fill), (26, 4, "Thịt xào khoai tây 50g nạc + 50g ktây", 0, 0, yellow_fill), (26, 5, "10k nạc xắt + 10K Ktây", 0, 0, yellow_fill),
        (27, 4, "Chả cá thì là 20g cá sống+10g mọc+50g xay", 0, 0, yellow_fill), (27, 5, "20k(6k cá sống+3k mọc+15k xay)", 0, 0, yellow_fill),
        (28, 4, "Thịt kho trứng 80g+2 cút", 0, 0, yellow_fill), (28, 5, "24k + 2 cút", 0, 0, yellow_fill),
        (29, 2, "300", 1, 0, None), (29, 3, "Mnuoc", 1, 0, None), (29, 4, "Bánh canh bột xắt 20g càng ghẹ + 50g nạc + 1 cút + 1 tôm 60", 1, 0, None), (29, 5, "6 càng ghẹ+15 nạc+1 cút+1 tôm size 60", 1, 0, None),
        (31, 3, "T.miệng", 0, 0, None), (31, 4, "Ổi", 0, 0, None), (31, 5, "74+33k ổi", 0, 0, None),
        (32, 2, "160", 1, 0, None), (32, 3, "Ca 2", 3, 0, None), (32, 4, "Cá hường chiên sả", 0, 0, None), (32, 5, "s8", 0, 0, None),
        (33, 4, "Thịt kho đậu hũ 80g + 1/2 đậu", 0, 0, None), (33, 5, "12.8k+1/2 đậu", 0, 0, None),
        (34, 2, "80", 1, 0, yellow_fill), (34, 4, "Gỏi hoa chuối 50g nạc", 0, 0, yellow_fill), (34, 5, "4k nạc", 0, 0, yellow_fill),
        (35, 4, "Ếch xào lăn 140g", 0, 0, yellow_fill), (35, 5, "11.2k ếch", 0, 0, yellow_fill),
        (36, 2, "120", 0, 0, yellow_fill), (36, 3, "Mnuoc", 0, 0, yellow_fill), (36, 4, "Bún bò 60g gàu+ 1 cây", 0, 0, yellow_fill), (36, 5, "7.2k gàu+122 cây chả", 0, 0, yellow_fill),
        (37, 2, "130", 0, 0, None), (37, 3, "Ca 3", 0, 0, None), (37, 4, "Nui giò heo 120g giò+40g mọc", 0, 0, None), (37, 5, "15.5k giò+5.2k mọc", 0, 0, None),

        (3, 6, "Saitex D13", 4, 0, None),
        (3, 7, "420+480", 1, 0, None), (3, 8, "Ca 1", 1, 0, None), (3, 9, "Cánh gà chiên nước mắm", 0, 0, None), (3, 10, "1 cái", 0, 0, None),
        (4, 9, "Chả lụa rim ớt chuông 40g", 0, 0, None), (4, 10, "16.8k + 19.2k", 0, 0, None),
        (5, 7, "100+100", 0, 0, None), (5, 8, "Nước C1", 0, 0, None), (5, 9, "Bánh canh thập cẩm 80g nạc + 10g xay + 1 cút + 1 tôm s60", 0, 0, yellow_fill), (5, 10, "", 0, 0, yellow_fill),
        (6, 7, "180", 0, 0, None), (6, 8, "Ca 2", 1, 0, None), (6, 9, "Thịt rim tôm 70g+1 tôm", 0, 0, None), (6, 10, "12.6k+1 tôm size 100", 0, 0, None),
        (7, 9, "Gỏi gà hành tây 50g file", 0, 0, None), (7, 10, "9k", 0, 0, None),

        (8, 6, "Saitex 4", 2, 0, None),
        (8, 7, "150+200", 0, 0, None), (8, 8, "Ca 1", 1, 0, None), (8, 9, "Cánh gà chiên nước mắm", 0, 0, None),
        (9, 9, "Chả lụa rim ớt chuông 40g", 0, 0, None), (9, 10, "6k + 8k", 0, 0, None),
        (10, 7, "140", 0, 0, None), (10, 8, "Nước C1", 0, 0, None), (10, 9, "Bánh canh thập cẩm 80g nạc + 10g xay + 1 cút + 1 tôm s60", 0, 0, yellow_fill), (10, 10, "11.2k nạc + 1.4k xay + 140 cút + 1 tôm size 60", 0, 0, yellow_fill),

        (11, 6, "Saitex 6, ICD", 2, 0, None),
        (11, 7, "610", 0, 0, None), (11, 8, "Ca 1", 1, 0, None), (11, 9, "Cánh gà chiên nước mắm", 0, 0, None), (11, 10, "24.4k", 0, 0, None),
        (12, 9, "Chả lụa rim ớt chuông 40g", 0, 0, None),
        (13, 7, "260", 0, 0, None), (13, 8, "Nước C1", 0, 0, None), (13, 9, "Bánh canh thập cẩm 80g nạc + 10g xay + 1 cút + 1 tôm s60", 0, 0, yellow_fill), (13, 10, "20.8k nạc + 2.6k xay + 260 cút + 1 tôm size 60", 0, 0, yellow_fill),

        (14, 6, "Figla", 4, 0, None),
        (14, 7, "300", 1, 0, None), (14, 8, "Ca 1", 1, 0, None), (14, 9, "Cánh gà chiên nước mắm", 0, 0, None), (14, 10, "1 cái", 0, 0, None),
        (15, 9, "Cá cơm rim 25g", 0, 0, None), (15, 10, "7.5k", 0, 0, None),
        (16, 7, "100", 0, 0, None), (16, 8, "M.nước", 0, 0, None), (16, 9, "Bánh canh thập cẩm 80g nạc + 10g xay + 1 cút + 1 tôm s60", 0, 0, yellow_fill), (16, 10, "8k nạc + 1k xay + 100 cút + 1 tôm s60", 0, 0, yellow_fill),
        (17, 8, "Ca 2 + Maspro", 1, 0, None),

        (19, 6, "Shiseido, 7A", 4, 0, None),
        (19, 7, "240+20", 0, 0, None), (19, 8, "T.miệng", 0, 0, None), (19, 9, "Ổi", 0, 0, None),
        (20, 7, "60+20vp", 1, 0, None), (20, 8, "Ca 1", 1, 0, None), (20, 9, "Đùi gà nướng", 0, 0, None), (20, 10, "1 cái", 0, 0, None),
        (21, 9, "Gỏi tai heo 30g tai+20g nạc", 0, 0, None), (21, 10, "(7.8k tai+5.2k nạc)+ (2.4k tai+1.6 nạc)", 0, 0, None),
        (22, 8, "T.miệng", 0, 0, None), (22, 9, "Sữa chua vinamilk", 0, 0, None),
        (23, 7, "40", 0, 0, None), (23, 8, "M.nước", 0, 0, None), (23, 9, "Phở tái + nạm 35 tái+20g nạm+20 viên", 0, 0, None), (23, 10, "1.4k tái+0.8k nạm+0.8k viên", 0, 0, None),

        (24, 6, "JFE", 4, 0, None),
        (24, 7, "65", 1, 0, None), (24, 8, "Ca 1", 1, 0, None), (24, 9, "Vịt kho gừng 120g", 0, 0, None), (24, 10, "7.8K", 0, 0, None),
        (25, 9, "Thịt xào khoai tây 50g nạc xắt + 40g ktây", 0, 0, None), (25, 10, "3.3K nạc xắt", 0, 0, None),
        (26, 9, "Ổi", 0, 0, None),
        (27, 7, "32", 1, 0, None), (27, 8, "Ca 2", 1, 0, None), (27, 9, "Cá hường chiên sả", 0, 0, None), (27, 10, "s8", 0, 0, None),
        (28, 9, "Thịt kho đậu hũ 80g + 1/2 đậu", 0, 0, None), (28, 10, "2.5k+1/2 đậu", 0, 0, None),

        (29, 6, "Briskheat", 8, 0, None),
        (29, 7, "420", 2, 0, None), (29, 8, "Ca 1", 2, 0, None), (29, 9, "Gà kho xả ớt 140g", 0, 0, None), (29, 10, "58k", 0, 0, None),
        (30, 9, "Salad trứng 1 trứng + 20 caron", 0, 0, None), (30, 10, "1 trứng+8 caron", 0, 0, None),
        (32, 9, "Rau câu", 0, 0, None),
        (33, 8, "Mnuoc", 0, 0, None), (33, 9, "0", 0, 0, None),
        (34, 7, "70", 2, 0, None), (34, 8, "Chay", 2, 0, None), (34, 9, "Lagu chay: 1/2 đậu, 25g tàu hũ+100g cà rốt+100g khoai tây", 0, 0, None),
        (35, 9, "Nấm rơm kho thơm", 0, 0, None),
    ]

    build_data(ws, data)
    for r, c in [(8,4), (21,4), (22,4), (23,4), (33,9)]: ws.cell(row=r, column=c).alignment = align_right
    apply_common_formatting(ws, max_row=37, rotate_col_a_f=False)


def sheet3():
    ws = wb.create_sheet(title="Bếp trung tâm 2")
    
    ws.merge_cells('A1:C1')
    ws['A1'] = "Thứ 2 -"
    ws['A1'].fill = blue_fill
    ws['A1'].font = white_bold
    ws['A1'].alignment = align_center

    ws['D1'] = "217.5"
    ws['D1'].fill = blue_fill
    ws['D1'].font = white_bold
    ws['D1'].alignment = align_center

    ws.merge_cells('E1:G1')
    ws['E1'] = "18/05/26"
    ws['E1'].fill = blue_fill
    ws['E1'].font = white_bold
    ws['E1'].alignment = align_center

    ws.merge_cells('H1:I1')
    ws['H1'] = "BẾP TRUNG TÂM"
    ws['H1'].fill = yellow_fill
    ws['H1'].font = black_bold
    ws['H1'].alignment = align_center
    
    for c, val in enumerate(["Cty", "", "Ca", "Món ăn", "", "Cty", "", "Ca", "Món ăn", ""], start=1):
        cell = ws.cell(row=2, column=c)
        cell.value = val
        cell.fill = blue_fill
        cell.font = white_bold
        cell.alignment = align_center

    data = [
        (3, 1, "Scavi", 6, 0, None),
        (3, 2, "815+75", 1, 0, None), (3, 3, "Ca 1", 1, 0, None), (3, 4, "Gà kho gừng 150g", 0, 0, None), (3, 5, "122.2k + 11.2k", 0, 0, None),
        (4, 4, "Cá thu nhật kho cà 70g + 20g", 0, 0, None), (4, 5, "57k +5.3k cá", 0, 0, None),
        (5, 2, "435", 1, 0, None), (5, 3, "Mnuoc", 1, 0, None), (5, 4, "Canh bún 30g nạc + 5g tôm + 10g cua + 45g xay + 20g mọc + 15g chả + 1/2 đậu", 1, 0, None), (5, 5, "13.1k nạc + 2.2k tôm + 4.4k cua + 20g xay + 8.7k mọc 6.5k chả + 218 đậu", 1, 0, None),
        (7, 3, "T.miệng", 0, 0, None), (7, 4, "0", 0, 0, None),
        (8, 2, "480", 1, 0, None), (8, 3, "Ca 1", 1, 0, None), (8, 4, "Gà hấp rau răm 150g", 0, 0, None), (8, 5, "72k má", 0, 0, None),
        (9, 4, "Cá lóc kho tiêu 100g", 0, 0, None), (9, 5, "48k", 0, 0, None),

        (10, 1, "Sofa", 5, 0, None),
        (10, 4, "0", 0, 0, None),
        (11, 2, "400", 1, 0, None), (11, 3, "Ca 2", 1, 0, None), (11, 4, "Bún bò huế 58g gàu+1c chả cây", 0, 0, None), (11, 5, "23.2k gàu + 400c chả cây", 0, 0, None),
        (13, 2, "20", 2, 0, None), (13, 3, "Ca2 CƠM QUẢN LÝ 40K", 2, 0, None), (13, 4, "Bò xào ớt chuông 60g", 0, 0, None), (13, 5, "1.2k", 0, 0, None),
        (14, 4, "Vịt kho gừng 200g", 0, 0, None), (14, 5, "4k", 0, 0, None),
        (15, 4, "Canh cà chua trứng", 0, 0, None),

        (16, 1, "Dona Quế Bằng", 8, 0, None),
        (16, 4, "Su su xào", 0, 0, None),
        (17, 2, "320", 2, 0, None), (17, 3, "Ca 1", 2, 0, None), (17, 4, "Gà hấp rau răm: 150g", 0, 0, None), (17, 5, "48k", 0, 0, None),
        (18, 4, "Cá thu nhật kho cà: 70g", 0, 0, None), (18, 5, "22.4k", 0, 0, None),
        (20, 2, "80", 1, 0, None), (20, 3, "Ca 2", 1, 0, None), (20, 4, "0", 0, 0, None),
        (21, 4, "Ổi", 0, 0, None), (21, 5, "43K ỔI", 0, 0, None),
        (22, 4, "Thịt kho dưa chua 70g", 0, 0, None), (22, 5, "5.6k", 0, 0, None),
        (23, 4, "Khô đù rim ngọt", 0, 0, None), (23, 5, "6.8k", 0, 0, None),
        (24, 2, "60", 0, 0, None), (24, 4, "Bún bò huế 65g gàu+20g chả+1 cây", 0, 0, None), (24, 5, "3.9k gàu+1.2k chả+65 cây chả", 0, 0, None),

        (25, 1, "Ca 1:", 0, 4, None),
        (26, 1, "Watabe: 10", 0, 0, None), (26, 4, "Saitex D13: 55", 0, 0, None),
        (27, 1, "Cát Thái: 110", 0, 0, None), (27, 4, "Saitex 4: 10", 0, 0, None),
        (28, 1, "Maspro: 07", 0, 0, None), (28, 4, "Saitex 6: 60", 0, 0, None),
        (29, 1, "Medic: 05", 0, 0, None), (29, 4, "Figla: 45", 0, 0, None),
        (30, 1, "Vikyno: 12", 0, 0, None), (30, 4, "Zamil: 15", 0, 0, None),
        (31, 1, "Dona QB: 10", 0, 0, None), (31, 4, "OSV: 20", 0, 0, None),
        (32, 1, "Ca 2:", 0, 4, None),
        (33, 1, "Saitex", 0, 0, None), (33, 4, "Sofa", 0, 0, None),
        (34, 1, "Figla", 0, 0, None), (34, 4, "OSV : 10", 0, 0, None), (34, 6, "Dona QB", 0, 0, None),
        (35, 1, "Ca 1: Scavi + Sofa+ Art + Shiseido: .....60+30+20+15=125", 0, 4, None),
        (36, 1, "Bánh hỏi chay: 1.5 đậu+3 cái chả giò+50g heo quay chay", 0, 4, None),
        (37, 1, "0", 0, 4, None),
        (39, 1, "Ca 1: Briskheat: 70", 0, 4, None),
        (40, 1, "Lagu chay: 1/2 đậu, 25g tàu hũ+100g cà rốt+100g khoai tây", 0, 4, None),
        (41, 1, "Nấm rơm kho thơm", 0, 4, None),

        (3, 6, "Scavi cơm khách 50k", 4, 0, None),
        (3, 7, "12", 4, 0, None), (3, 8, "", 4, 0, None), (3, 9, "Cá lăng kho gừng sả 170g", 0, 0, None), (3, 10, "2k", 0, 0, None),
        (4, 9, "Tôm cháy tỏi 3 con size 50", 0, 0, None),
        (5, 9, "Canh bí xanh", 0, 0, None),
        (6, 9, "Rau muống xào tỏi", 0, 0, None),
        (7, 9, "Táo", 0, 0, None),

        (8, 6, "Scavi cơm khách 70k", 5, 0, None),
        (8, 7, "9", 5, 0, None), (8, 8, "", 5, 0, None), (8, 9, "Cơm trắng gạo ST25", 0, 0, None),
        (9, 9, "Cá lăng kho gừng sả 250g", 0, 0, None), (9, 10, "2.3k", 0, 0, None),
        (10, 9, "Tôm cháy tỏi 5 con size 50", 0, 0, None), (10, 10, "45 con size 50", 0, 0, None),
        (11, 9, "Canh bí xanh", 0, 0, None),
        (12, 9, "Rau muống xào tỏi", 0, 0, None),
        (13, 9, "Táo", 0, 0, None),

        (14, 6, "Sofa cơm QL 40K", 6, 0, None),
        (14, 7, "32", 6, 0, None), (14, 8, "", 6, 0, None), (14, 9, "Cơm trắng gạo ST25", 0, 0, None),
        (15, 9, "Cá lăng kho gừng sả 170g", 0, 0, None), (15, 10, "5.4k", 0, 0, None),
        (16, 9, "Cánh gà chiên nước mắm", 0, 0, yellow_fill), (16, 10, "", 0, 0, yellow_fill),
        (17, 9, "Canh rong biển", 0, 0, None),
        (18, 9, "Rau bí xào tỏi", 0, 0, None),
        (19, 9, "Dưa Gang", 0, 0, yellow_fill), (19, 10, "", 0, 0, yellow_fill),
        (20, 9, "Cơm trắng gạo ST25", 0, 0, None),

        (21, 6, "Artwell", 3, 0, None),
        (21, 7, "270", 2, 0, None), (21, 8, "Ca 1", 2, 0, None), (21, 9, "Gà hấp rau răm 140g", 0, 0, None), (21, 10, "37.8k má", 0, 0, None),
        (22, 9, "Cá thu nhật kho cà 70g", 0, 0, None), (22, 10, "18.9k", 0, 0, None),
        (23, 9, "0", 0, 0, None), (23, 10, "0", 0, 0, None),
        (24, 7, "100", 0, 0, None), (24, 8, "Ca 2", 0, 0, None), (24, 9, "Bún bò huế 58g gàu+1c chả cây", 0, 0, None), (24, 10, "5.8k gàu+100 cây chả", 0, 0, None),

        (25, 7, "M,chay", 2, 1, yellow_fill), (25, 9, "Lagu chay 1/2 đậu+25g tàu hũ ki+100g ktay+100g rốt", 1, 1, yellow_fill),
        (27, 9, "Đùi gà chay rang muối 80g", 0, 1, yellow_fill),
        (32, 8, "Ca 1", 0, 0, None), (32, 9, "360", 0, 0, None),
        (33, 8, "Lagu chay 1/2 đậu+25g tàu hũ ki+100g ktay+100g rốt", 0, 2, None),
        (34, 8, "Đùi gà chay rang muối 80g", 0, 2, None),
        (36, 8, "Ca 2", 0, 0, None), (36, 9, "35", 0, 0, None),
        (37, 8, "Đậu hũ kho tiêu 2", 0, 2, None),
        (38, 8, "Củ hủ dừa kho nấm 100g +50g nấm", 0, 2, None),
        (40, 8, "Ca 3", 0, 0, None), (40, 9, "35", 0, 0, None),
        (41, 8, "Đậu hũ kho tiêu 2", 0, 2, None),
        (42, 8, "Củ hủ dừa kho nấm 100g +50g nấm", 0, 2, None),
    ]

    build_data(ws, data)
    for r, c in [(7,4), (10,4), (20,4), (23,9), (23,10), (24,9), (24,10)]: ws.cell(row=r, column=c).alignment = align_right
        
    apply_common_formatting(ws, max_row=42, rotate_col_a_f=False)

    for r in range(25, 43):
        for c in range(1, 7):
            if type(ws.cell(row=r, column=c).fill) == PatternFill and ws.cell(row=r, column=c).fill.start_color.rgb == 'FFFCE4D6':
                ws.cell(row=r, column=c).fill = PatternFill()
            if ws.cell(row=r, column=c).value:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="left", vertical="center")
        if type(ws.cell(row=r, column=10).fill) == PatternFill and ws.cell(row=r, column=10).fill.start_color.rgb == 'FFFCE4D6':
            ws.cell(row=r, column=10).fill = PatternFill()

    for r in [25, 32, 35, 39]:
        ws.cell(row=r, column=1).font = black_bold

sheet1()
sheet2()
sheet3()

wb.save("BaoCao.xlsx")
print("Tạo file BaoCao.xlsx thành công!")
